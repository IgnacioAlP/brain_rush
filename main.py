"""
=====================================================================
BRAIN RUSH - Sistema de Gamificación Educativa
=====================================================================
Aplicación Flask para crear y gestionar cuestionarios interactivos
con mecánicas de juego en tiempo real.

Autor: Brain Rush Team
Fecha: 2025
=====================================================================
"""

# ==================== IMPORTS ====================
# Librerías estándar de Python
import os
import json
import random
import traceback
import re
from io import BytesIO
from datetime import datetime, timedelta

# Flask y extensiones
from flask import (
    Flask, render_template, request, redirect, url_for, flash, 
    jsonify, Response, session, send_file, current_app, make_response
)
from werkzeug.exceptions import InternalServerError
from dotenv import load_dotenv
from itsdangerous import URLSafeTimedSerializer
import pymysql.cursors
import requests

# Configuración y base de datos
from config import config
from bd import verificar_conexion, obtener_conexion, inicializar_usuarios_prueba
from extensions import mail

# Utilidades de autenticación
from utils_auth import (
    login_required,
    jwt_or_session_required,
    docente_required,
    estudiante_required,
    crear_token_jwt,
    verificar_token_jwt,
    establecer_cookies_usuario,
    limpiar_cookies_usuario,
    obtener_usuario_cookies
)

# Controladores de negocio
from controladores import (
    controlador_salas,
    controlador_usuario,
    controlador_cuestionarios,
    controlador_juego,
    controlador_preguntas,
    controlador_participaciones,
    controlador_ranking,
    controlador_recompensas,
    controlador_respuestas,
    controlador_opciones,
    controlador_xp
)

# APIs CRUD
from api_crud import api_crud

# Verificar disponibilidad de MSAL para OneDrive
try:
    import msal
    MSAL_AVAILABLE = True
except ImportError:
    MSAL_AVAILABLE = False
    print("⚠️  WARNING: msal no está instalado. La funcionalidad de OneDrive no estará disponible.")
    print("   Para instalar: pip install msal")

# Cargar variables de entorno
load_dotenv()

# ==================== CONFIGURACIÓN DE LA APLICACIÓN ====================

# Crear instancia de Flask
app = Flask(__name__)

# Cargar configuración según el entorno
env = os.getenv('FLASK_ENV', 'development')
app_config = config.get(env, config['default'])
app.config.from_object(app_config)
app.secret_key = app_config.SECRET_KEY

# Inicializar extensiones
mail.init_app(app)

# Serializador para tokens de email (usa la misma SECRET_KEY)
serializer = URLSafeTimedSerializer(app.config['SECRET_KEY'])

# Registrar blueprints de APIs
app.register_blueprint(api_crud)

# Configuración de sesiones seguras
app.config['SESSION_COOKIE_SECURE'] = app.config.get('ENV') == 'production'
app.config['SESSION_COOKIE_HTTPONLY'] = True
app.config['SESSION_COOKIE_SAMESITE'] = 'Lax'
app.config['PERMANENT_SESSION_LIFETIME'] = timedelta(days=7)


# ==================== CONTEXTO DE JINJA ====================

@app.context_processor
def utility_processor():
    """Proporciona funciones útiles al contexto de Jinja2"""
    def now():
        return datetime.now()
    return dict(now=now)


# ==================== MIDDLEWARE Y HOOKS ====================

@app.before_request
def fix_session():
    """
    Middleware para gestionar la limpieza de sesión en rutas API
    y corregir inconsistencias en los datos de sesión
    """
    # Lista de rutas API que permiten autenticación por sesión (además de JWT)
    api_rutas_con_sesion = [
        '/api/cuestionarios/',
        '/api/cuestionario/',
        '/api/sala/',
        '/api/exportar-',
        '/api/participante',
        '/api/perfil-xp/',
        '/api/insignias',
        '/api/tienda-insignias',
        '/api/comprar-insignia',
        '/api/recompensas',
        '/api/otorgar_recompensas_automaticas',
        '/api/ranking-xp',
        '/unirse-juego'
    ]
    
    # Limpiar sesión en rutas API que NO están en la lista blanca
    if request.path.startswith('/api/'):
        permitir_sesion = any(ruta in request.path for ruta in api_rutas_con_sesion)
        
        if not permitir_sesion:
            session.clear()
            response = make_response()
            response.delete_cookie('session', path='/')
            response.delete_cookie('user_id', path='/')
            response.delete_cookie('user_name', path='/')
    
    # Corregir inconsistencia en sesión (tipo_usuario vs usuario_tipo)
    elif 'tipo_usuario' in session and 'usuario_tipo' not in session:
        session['usuario_tipo'] = session['tipo_usuario']


# ==================== FUNCIONES HELPER ====================

def es_sala_automatica(pin_sala):
    """
    Verifica si un PIN corresponde a una sala en modo automático.
    
    Args:
        pin_sala: Código PIN de la sala
        
    Returns:
        bool: True si es sala automática, False en caso contrario
        
    Formato:
        - Salas automáticas: AUTOXXXX (8 caracteres)
        - Salas normales: 6 dígitos numéricos
    """
    if not pin_sala:
        return False
    return pin_sala.startswith('AUTO') and len(pin_sala) == 8


def admin_required(f):
    """
    Decorador para requerir permisos de administrador.
    Es un alias de docente_required para mantener compatibilidad.
    """
    return docente_required(f)


# ==================== FUNCIONES DE SALAS ====================

def verificar_y_crear_tabla_salas():
    """
    Verifica si existe la tabla salas_juego y la crea si no existe.
    
    Returns:
        bool: True si la tabla existe o se creó exitosamente, False en caso de error
    """
    try:
        conexion = obtener_conexion()
        cursor = conexion.cursor()

        # Verificar si la tabla existe
        cursor.execute("SHOW TABLES LIKE 'salas_juego'")
        tabla_existe = cursor.fetchone()

        if not tabla_existe:
            print("DEBUG: Tabla salas_juego no existe, creándola...")
            create_table_query = """
            CREATE TABLE salas_juego (
                id_sala INT AUTO_INCREMENT PRIMARY KEY,
                pin_sala VARCHAR(6) UNIQUE NOT NULL,
                id_cuestionario INT,
                modo_juego ENUM('individual', 'grupo') DEFAULT 'individual',
                estado ENUM('esperando', 'en_curso', 'finalizada') DEFAULT 'esperando',
                tiempo_por_pregunta INT DEFAULT 30,
                fecha_creacion DATETIME DEFAULT CURRENT_TIMESTAMP,
                FOREIGN KEY (id_cuestionario) REFERENCES cuestionarios(id_cuestionario)
            )
            """
            cursor.execute(create_table_query)
            conexion.commit()
            print("DEBUG: Tabla salas_juego creada exitosamente")
        else:
            print("DEBUG: Tabla salas_juego ya existe")

        conexion.close()
        return True

    except Exception as e:
        print(f"DEBUG: Error al verificar/crear tabla salas_juego: {str(e)}")
        if 'conexion' in locals():
            conexion.close()
        return False

def crear_sala_simple(cuestionario_id):
    """
    Crea una sala de juego simple para un cuestionario.
    
    Args:
        cuestionario_id: ID del cuestionario asociado
        
    Returns:
        tuple: (sala_id, pin_sala) si tiene éxito, (None, None) si falla
    """
    try:
        print(f"DEBUG: crear_sala_simple - Iniciando con cuestionario_id: {cuestionario_id}")

        # Verificar que la tabla exista
        if not verificar_y_crear_tabla_salas():
            return None, None

        pin_sala = str(random.randint(100000, 999999))
        print(f"DEBUG: crear_sala_simple - PIN generado: {pin_sala}")

        conexion = obtener_conexion()
        print(f"DEBUG: crear_sala_simple - Conexión obtenida: {conexion}")

        cursor = conexion.cursor()
        query = "INSERT INTO salas_juego (pin_sala, id_cuestionario, modo_juego, estado) VALUES (%s, %s, %s, %s)"
        valores = (pin_sala, cuestionario_id, 'individual', 'esperando')
        print(f"DEBUG: crear_sala_simple - Query: {query}")
        print(f"DEBUG: crear_sala_simple - Valores: {valores}")

        cursor.execute(query, valores)
        conexion.commit()
        sala_id = cursor.lastrowid

        print(f"DEBUG: crear_sala_simple - Sala creada exitosamente con ID: {sala_id}")
        conexion.close()

        return sala_id, pin_sala

    except Exception as e:
        print(f"DEBUG: ERROR en crear_sala_simple: {str(e)}")
        import traceback
        traceback.print_exc()
        if 'conexion' in locals():
            conexion.close()
        return None, None

def crear_grupos_para_sala(sala_id, num_grupos):
    """
    Crea grupos para una sala de juego.

    Args:
        sala_id: ID de la sala
        num_grupos: Número de grupos a crear (2-6)

    Returns:
        True si se crearon exitosamente, False en caso de error
    """
    try:
        print(f"DEBUG: crear_grupos_para_sala - Iniciando para sala {sala_id} con {num_grupos} grupos")

        if num_grupos < 2 or num_grupos > 6:
            print(f"DEBUG: crear_grupos_para_sala - Número de grupos inválido: {num_grupos}")
            return False

        conexion = obtener_conexion()
        cursor = conexion.cursor()

        # Verificar que la sala exista
        cursor.execute("SELECT id_sala FROM salas_juego WHERE id_sala = %s", (sala_id,))
        if not cursor.fetchone():
            print(f"DEBUG: crear_grupos_para_sala - Sala {sala_id} no encontrada")
            conexion.close()
            return False

        # Crear los grupos
        for numero_grupo in range(1, num_grupos + 1):
            nombre_grupo = f"Grupo {numero_grupo}"
            cursor.execute("""
                INSERT INTO grupos_sala (id_sala, numero_grupo, nombre_grupo)
                VALUES (%s, %s, %s)
            """, (sala_id, numero_grupo, nombre_grupo))
            print(f"DEBUG: crear_grupos_para_sala - Grupo {numero_grupo} creado: {nombre_grupo}")

        conexion.commit()
        print(f"DEBUG: crear_grupos_para_sala - {num_grupos} grupos creados exitosamente")
        conexion.close()

        return True

    except Exception as e:
        print(f"DEBUG: ERROR en crear_grupos_para_sala: {str(e)}")
        import traceback
        traceback.print_exc()
        if 'conexion' in locals():
            try:
                conexion.rollback()
                conexion.close()
            except:
                pass
        return False

def obtener_sala_por_id_simple(id_sala):
    """
    Obtiene información completa de una sala por su ID.
    
    Args:
        id_sala: ID de la sala a buscar
        
    Returns:
        dict: Información de la sala o None si no existe
    """
    conexion = obtener_conexion()
    try:
        cursor = conexion.cursor()
        # Debug: verificar estado directamente de BD
        cursor.execute("""
            SELECT id_sala, pin_sala, id_cuestionario, modo_juego, estado, max_participantes, fecha_creacion, total_preguntas
            FROM salas_juego
            WHERE id_sala = %s
        """, (id_sala,))
        print(f"🔍 [obtener_sala_por_id_simple] Consultando sala {id_sala}")
        sala = cursor.fetchone()
        if sala:
            # Determinar si es sala con docente (modo manual) o juego individual (modo automático)
            # REGLA DEFINITIVA:
            #   - Si tiene PIN de sala (modo colaborativo) = Sala con docente
            #   - Si estado es 'esperando' = Sala con docente (esperando que docente inicie)
            #   - Si modo_juego es 'colaborativo' = Sala con docente
            #   - Si hay múltiples participantes = Sala con docente (clase)
            #   - En cualquier otro caso = Juego individual

            modo_juego = sala[3]
            estado = sala[4]
            pin_sala = sala[1]

            # Debug logs
            print(f"\n🔍 DEBUG - Detectando modo de sala {id_sala}:")
            print(f"   PIN: {pin_sala} (tipo: {type(pin_sala)}, len: {len(str(pin_sala)) if pin_sala else 0})")
            print(f"   Modo juego: {modo_juego}")
            print(f"   Estado ACTUAL desde BD: '{estado}' (tipo: {type(estado)})")

            # REGLA 1: Salas colaborativas siempre tienen docente
            if modo_juego == 'colaborativo':
                tiene_docente = True
                print(f"   ✅ REGLA 1: modo_juego='colaborativo' → tiene_docente = True")
            # REGLA 2: Si está esperando, significa que el docente aún no inició
            elif estado == 'esperando':
                tiene_docente = True
                print(f"   ✅ REGLA 2: estado='esperando' → tiene_docente = True")
            # REGLA 3: Verificar si es sala automática (PIN formato AUTOXXXX)
            elif pin_sala and es_sala_automatica(pin_sala):
                tiene_docente = False
                print(f"   ✅ REGLA 3: PIN automático ({pin_sala}) → tiene_docente = False (modo automático)")
            # REGLA 4: Si tiene PIN normal (6 dígitos), es sala de clase con docente
            elif pin_sala and len(str(pin_sala)) == 6 and str(pin_sala).isdigit():
                tiene_docente = True
                print(f"   ✅ REGLA 4: PIN normal de 6 dígitos → tiene_docente = True")
            else:
                # REGLA 5: Verificar si hay múltiples participantes (indica sala de clase)
                cursor.execute("""
                    SELECT COUNT(*) FROM participantes_sala
                    WHERE id_sala = %s AND estado != 'desconectado'
                """, (id_sala,))
                num_participantes = cursor.fetchone()[0]
                tiene_docente = (num_participantes > 1)
                print(f"   🔍 REGLA 5: {num_participantes} participantes → tiene_docente = {tiene_docente}")

            print(f"   🎯 RESULTADO FINAL: tiene_docente = {tiene_docente}\n")

            return {
                'id': sala[0],
                'pin_sala': sala[1],
                'id_cuestionario': sala[2],
                'modo_juego': sala[3],
                'estado': sala[4],
                'max_participantes': sala[5],
                'fecha_creacion': sala[6],
                'total_preguntas': sala[7] if len(sala) > 7 and sala[7] is not None else 0,
                'tiene_docente': tiene_docente  # True = modo manual, False = modo automático
            }
        return None
    finally:
        conexion.close()

def obtener_cuestionario_por_id_simple(id_cuestionario):
    """
    Obtiene información básica de un cuestionario por su ID.
    
    Args:
        id_cuestionario: ID del cuestionario
        
    Returns:
        tuple: (id, titulo, descripcion) o None si no existe
    """
    conexion = obtener_conexion()
    try:
        cursor = conexion.cursor()
        cursor.execute("SELECT id_cuestionario, titulo, descripcion FROM cuestionarios WHERE id_cuestionario = %s", (id_cuestionario,))
        return cursor.fetchone()
    finally:
        conexion.close()

def obtener_preguntas_por_cuestionario_simple(id_cuestionario):
    """
    Obtiene el conteo de preguntas de un cuestionario.
    
    Args:
        id_cuestionario: ID del cuestionario
        
    Returns:
        list: Lista con objetos de preguntas
    """
    conexion = obtener_conexion()
    try:
        cursor = conexion.cursor()
        cursor.execute("""
            SELECT COUNT(*) FROM cuestionario_preguntas cp
            JOIN preguntas p ON cp.id_pregunta = p.id_pregunta
            WHERE cp.id_cuestionario = %s
        """, (id_cuestionario,))
        count = cursor.fetchone()[0]
        return [{'id': i} for i in range(count)]  # Fake questions for count
    finally:
        conexion.close()

def obtener_cuestionarios_por_docente_simple(id_docente):
    """
    Obtiene todos los cuestionarios creados por un docente.
    
    Args:
        id_docente: ID del docente
        
    Returns:
        list: Lista de diccionarios con información de cuestionarios
    """
    conexion = obtener_conexion()
    try:
        cursor = conexion.cursor()
        cursor.execute("""
            SELECT c.id_cuestionario, c.titulo, c.descripcion, c.id_docente, c.estado,
                   c.fecha_creacion, c.fecha_programada, c.fecha_publicacion,
                   COUNT(DISTINCT cp.id_pregunta) as num_preguntas
            FROM cuestionarios c
            LEFT JOIN cuestionario_preguntas cp ON c.id_cuestionario = cp.id_cuestionario
            WHERE c.id_docente = %s
            GROUP BY c.id_cuestionario, c.titulo, c.descripcion, c.id_docente, c.estado,
                     c.fecha_creacion, c.fecha_programada, c.fecha_publicacion
            ORDER BY c.fecha_creacion DESC
        """, (id_docente,))
        resultados = cursor.fetchall()

        cuestionarios = []
        for resultado in resultados:
            cuestionarios.append({
                'id_cuestionario': resultado[0],
                'titulo': resultado[1],
                'descripcion': resultado[2],
                'id_docente': resultado[3],
                'estado': resultado[4],
                'fecha_creacion': resultado[5],  # Mantenemos como datetime object
                'fecha_programada': resultado[6],  # Mantenemos como datetime object
                'fecha_publicacion': resultado[7],  # Mantenemos como datetime object
                'num_preguntas': resultado[8]  # Ahora cuenta las preguntas reales
            })

        return cuestionarios
    finally:
        conexion.close()


# ==================== FUNCIONES DE ESTADÍSTICAS ====================

def obtener_partidas_recientes_estudiante(usuario_id, limit=5):
    """
    Obtiene las partidas recientes de un estudiante.
    
    Args:
        usuario_id: ID del usuario estudiante
        limit: Número máximo de partidas a retornar (default: 5)
        
    Returns:
        list: Lista de diccionarios con información de partidas
    """
    conexion = obtener_conexion()
    cursor = conexion.cursor()

    try:
        query = """
            SELECT
                ps.id_participante,
                rs.puntaje_total,
                rs.tiempo_total_respuestas,
                ps.fecha_union,
                s.id_sala,
                s.pin_sala,
                s.modo_juego,
                c.id_cuestionario,
                c.titulo as cuestionario_titulo,
                c.descripcion as cuestionario_descripcion,
                (SELECT COUNT(*) FROM cuestionario_preguntas cp WHERE cp.id_cuestionario = c.id_cuestionario) as total_preguntas
            FROM participantes_sala ps
            INNER JOIN salas_juego s ON ps.id_sala = s.id_sala
            INNER JOIN cuestionarios c ON s.id_cuestionario = c.id_cuestionario
            LEFT JOIN ranking_sala rs ON ps.id_participante = rs.id_participante AND ps.id_sala = rs.id_sala
            WHERE ps.id_usuario = %s
            ORDER BY ps.fecha_union DESC
            LIMIT %s
        """

        cursor.execute(query, (usuario_id, limit))
        partidas = cursor.fetchall()

        partidas_formateadas = []
        for partida in partidas:
            es_individual = (partida[6] == 'individual')

            partidas_formateadas.append({
                'id_participante': partida[0],
                'puntaje': partida[1] if partida[1] is not None else 0,
                'tiempo': partida[2] if partida[2] is not None else 0,
                'fecha': partida[3],
                'sala_id': partida[4],
                'sala_pin': partida[5],
                'modo_juego': partida[6],
                'es_individual': es_individual,
                'cuestionario_id': partida[7],
                'cuestionario_titulo': partida[8],
                'cuestionario_descripcion': partida[9],
                'total_preguntas': partida[10],
                'tipo': 'Individual' if es_individual else 'Multijugador'
            })

        return partidas_formateadas

    except Exception as e:
        print(f"❌ Error en obtener_partidas_recientes_estudiante: {e}")
        traceback.print_exc()
        return []
    finally:
        cursor.close()
        conexion.close()


def obtener_estadisticas_estudiante(usuario_id):
    """
    Obtiene estadísticas completas del estudiante.
    
    Args:
        usuario_id: ID del usuario estudiante
        
    Returns:
        dict: Diccionario con estadísticas del estudiante
    """
    conexion = obtener_conexion()
    cursor = conexion.cursor()

    try:
        # Total de participaciones
        cursor.execute("""
            SELECT COUNT(*)
            FROM participantes_sala
            WHERE id_usuario = %s
        """, (usuario_id,))
        total_participaciones = cursor.fetchone()[0] or 0

        # Promedio de puntaje
        cursor.execute("""
            SELECT AVG(rs.puntaje_total)
            FROM ranking_sala rs
            INNER JOIN participantes_sala ps ON rs.id_participante = ps.id_participante
            WHERE ps.id_usuario = %s AND rs.puntaje_total IS NOT NULL
        """, (usuario_id,))
        promedio_puntaje = cursor.fetchone()[0] or 0

        # Mejor posición
        cursor.execute("""
            SELECT MIN(rs.posicion)
            FROM ranking_sala rs
            INNER JOIN participantes_sala ps ON rs.id_participante = ps.id_participante
            WHERE ps.id_usuario = %s AND rs.posicion IS NOT NULL
        """, (usuario_id,))
        mejor_posicion = cursor.fetchone()[0] or 'N/A'

        # Recompensas obtenidas
        cursor.execute("""
            SELECT COUNT(*)
            FROM recompensas_otorgadas
            WHERE id_estudiante = %s
        """, (usuario_id,))
        recompensas_obtenidas = cursor.fetchone()[0] or 0

        return {
            'total_participaciones': total_participaciones,
            'promedio_puntaje': round(promedio_puntaje, 1) if promedio_puntaje else 0,
            'mejor_posicion': mejor_posicion,
            'recompensas_obtenidas': recompensas_obtenidas
        }

    except Exception as e:
        print(f"❌ Error en obtener_estadisticas_estudiante: {e}")
        traceback.print_exc()
        return {
            'total_participaciones': 0,
            'promedio_puntaje': 0,
            'mejor_posicion': 'N/A',
            'recompensas_obtenidas': 0
        }
    finally:
        cursor.close()
        conexion.close()


# =====================================================================
# RUTAS DE LA APLICACIÓN
# =====================================================================

# =====================================================================
# RUTAS DE LA APLICACIÓN
# =====================================================================

# ==================== RUTAS PRINCIPALES ====================

@app.route('/')
def index():
    """
    Ruta principal - Redirige según estado de autenticación.
    """
    if 'logged_in' in session and session['logged_in']:
        if session.get('usuario_tipo') == 'estudiante':
            return redirect(url_for('dashboard_estudiante'))
        else:
            return redirect(url_for('dashboard_admin'))
    else:
        return redirect(url_for('login'))


@app.route('/errorsistema')
def error_sistema_page():
    """Página de error del sistema"""
    return render_template('ErrorSistema.html')


# ==================== RUTAS LEGACY ====================

@app.route('/crear_sala')
def crear_sala():
    """Legacy route - Usar /crear-sala en su lugar"""
    return render_template('CrearSala.html')


@app.route('/generar_preguntas', methods=['POST'])
def generar_preguntas():
    """Legacy route - Generación de preguntas"""
    nombre_sala = request.form['nombre_sala']
    return render_template('GenerarPreguntas.html', nombre_sala=nombre_sala)


# ==================== RUTAS DE AUTENTICACIÓN ====================

@app.route('/registrarse', methods=['GET', 'POST'])
def registrarse():
    if request.method == 'POST':
        data = request.get_json(silent=True) or request.form.to_dict()
        try:
            nombre = (data.get('nombre') or '').strip()
            apellidos = (data.get('apellidos') or '').strip()
            email = (data.get('email') or '').strip().lower()
            password = data.get('password') or ''
            tipo_usuario = (data.get('tipo_usuario') or 'estudiante').strip().lower()

            # Validar dominio del correo electrónico
            dominios_permitidos = ['@usat.pe', '@usat.edu.pe']
            dominio_valido = any(email.endswith(dominio) for dominio in dominios_permitidos)

            if not dominio_valido:
                error_msg = 'Solo se permiten correos institucionales con dominio @usat.pe o @usat.edu.pe'
                flash(error_msg, 'error')
                return render_template('Registrarse.html')

            success, result = controlador_usuario.crear_usuario(nombre, apellidos, email, password, tipo_usuario)
            if not success:
                error_msg = result or 'No se pudo registrar el usuario'
                flash(error_msg, 'error')
                return render_template('Registrarse.html')

            # Solo intentar enviar correo si está habilitado en la configuración
            if app.config.get('MAIL_ENABLED', False):
                # Intentar enviar el correo de confirmación
                correo_enviado, mensaje_correo = controlador_usuario.enviar_correo_confirmacion(email)

                if correo_enviado:
                    flash('¡Registro exitoso! Por favor, revisa tu correo para activar tu cuenta. Verifica también la carpeta de spam.', 'success')
                else:
                    flash('Registro exitoso, pero no se pudo enviar el correo de confirmación. Por favor, contacta a soporte para activar tu cuenta.', 'warning')
                    print(f"DEBUG: Error al enviar correo: {mensaje_correo}")
            else:
                # Correo desactivado - usuario creado directamente como activo
                flash('¡Registro exitoso! Ya puedes iniciar sesión con tu cuenta.', 'success')

            return redirect(url_for('login'))
        except Exception as e:
            print(f"DEBUG: Excepción en registro: {str(e)}")
            flash('Error interno del sistema', 'error')
            return render_template('Registrarse.html')
    return render_template('Registrarse.html')

@app.route('/confirmar/<token>')
def confirmar_email(token):
    """Confirmar email de usuario mediante token"""
    try:
        print(f"🔐 Intentando confirmar con token: {token[:20]}...")

        # Intentar decodificar el token (válido por 1 hora)
        email = serializer.loads(token, salt='email-confirm-salt', max_age=3600)
        print(f"✅ Token válido para email: {email}")

        # Activar la cuenta
        success, message = controlador_usuario.activar_cuenta_usuario(email)

        if success:
            flash(message, 'success')
            print(f"✅ Cuenta activada exitosamente: {email}")
        else:
            flash(f"Error al activar la cuenta: {message}", 'error')
            print(f"⚠️ Error al activar cuenta: {message}")

        return redirect(url_for('login'))

    except Exception as e:
        error_type = type(e).__name__
        print(f"❌ Error al confirmar email ({error_type}): {e}")
        import traceback
        traceback.print_exc()

        # Mensajes de error más específicos
        if 'expired' in str(e).lower() or 'Signature expired' in str(e):
            flash('El enlace de confirmación ha expirado (válido por 1 hora). Por favor, solicita un nuevo correo de confirmación.', 'error')
        elif 'bad' in str(e).lower() or 'invalid' in str(e).lower():
            flash('El enlace de confirmación es inválido. Por favor, verifica que copiaste la URL completa o solicita un nuevo correo.', 'error')
        else:
            flash('Error al procesar la confirmación. Por favor, intenta nuevamente o solicita un nuevo correo de confirmación.', 'error')

        return redirect(url_for('login'))

# ========== RECUPERACIÓN DE CONTRASEÑA ==========

@app.route('/recuperar-contrasena', methods=['GET', 'POST'])
def recuperar_contrasena():
    """Página para solicitar recuperación de contraseña"""
    if request.method == 'GET':
        return render_template('RecuperarContrasena.html')

    try:
        # Obtener email del formulario
        email = request.form.get('email', '').strip().lower()

        if not email:
            return render_template('RecuperarContrasena.html',
                                 mensaje_error='Por favor ingresa tu correo electrónico.')

        # Validar formato de email
        import re
        email_regex = r'^[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}$'
        if not re.match(email_regex, email):
            return render_template('RecuperarContrasena.html',
                                 mensaje_error='Por favor ingresa un correo electrónico válido.')

        # Llamar a la función de solicitar recuperación
        success, mensaje = controlador_usuario.solicitar_recuperacion_contrasena(email)

        # Siempre mostrar mensaje de éxito por seguridad (no revelar si el email existe)
        return render_template('RecuperarContrasena.html',
                             mensaje_exito='Si el correo existe en nuestro sistema, recibirás instrucciones para restablecer tu contraseña.')

    except Exception as e:
        print(f"❌ Error en recuperar_contrasena: {e}")
        import traceback
        traceback.print_exc()
        return render_template('RecuperarContrasena.html',
                             mensaje_error='Hubo un error al procesar tu solicitud. Por favor intenta nuevamente.')

@app.route('/restablecer/<token>', methods=['GET', 'POST'])
def restablecer_contrasena(token):
    """Página para restablecer contraseña con token"""

    # Validar el token primero
    success, resultado = controlador_usuario.validar_token_recuperacion(token)

    if not success:
        # Token inválido o expirado
        return render_template('RestablecerContrasena.html',
                             token=token,
                             mensaje_error=resultado)

    # Token válido, resultado contiene el email
    email = resultado

    if request.method == 'GET':
        # Mostrar formulario de nueva contraseña
        return render_template('RestablecerContrasena.html', token=token)

    # POST: Procesar nueva contraseña
    try:
        password = request.form.get('password', '')
        confirm_password = request.form.get('confirm_password', '')

        # Validaciones
        if not password or not confirm_password:
            return render_template('RestablecerContrasena.html',
                                 token=token,
                                 mensaje_error='Por favor completa todos los campos.')

        if password != confirm_password:
            return render_template('RestablecerContrasena.html',
                                 token=token,
                                 mensaje_error='Las contraseñas no coinciden.')

        # Validar fortaleza de contraseña
        if len(password) < 8:
            return render_template('RestablecerContrasena.html',
                                 token=token,
                                 mensaje_error='La contraseña debe tener al menos 8 caracteres.')

        import re
        if not re.search(r'[A-Z]', password):
            return render_template('RestablecerContrasena.html',
                                 token=token,
                                 mensaje_error='La contraseña debe contener al menos una mayúscula.')

        if not re.search(r'[a-z]', password):
            return render_template('RestablecerContrasena.html',
                                 token=token,
                                 mensaje_error='La contraseña debe contener al menos una minúscula.')

        if not re.search(r'[0-9]', password):
            return render_template('RestablecerContrasena.html',
                                 token=token,
                                 mensaje_error='La contraseña debe contener al menos un número.')

        # Restablecer la contraseña
        success, mensaje = controlador_usuario.restablecer_contrasena(email, password)

        if success:
            flash('✅ Tu contraseña ha sido actualizada exitosamente. Ya puedes iniciar sesión.', 'success')
            return redirect(url_for('login'))
        else:
            return render_template('RestablecerContrasena.html',
                                 token=token,
                                 mensaje_error=mensaje)

    except Exception as e:
        print(f"❌ Error en restablecer_contrasena POST: {e}")
        import traceback
        traceback.print_exc()
        return render_template('RestablecerContrasena.html',
                             token=token,
                             mensaje_error='Hubo un error al actualizar tu contraseña. Por favor intenta nuevamente.')

@app.route('/unirse_a_sala', methods=['GET', 'POST'])
def unirse_a_sala():
    if request.method == 'GET':
        return render_template('UnirseASala.html')

    # POST - Procesar unión a sala
    try:
        data = request.get_json(silent=True) or request.form.to_dict()
        codigo_sala = data.get('codigo_sala', '').strip().upper()

        print(f"\n{'='*80}")
        print(f"🎮 ESTUDIANTE INTENTANDO UNIRSE (UnirseASala.html)")
        print(f"Código: {codigo_sala}")
        print(f"Datos recibidos: {data}")
        print(f"{'='*80}\n")

        # Validar código (debe ser 6 caracteres)
        if not codigo_sala or len(codigo_sala) != 6:
            error_msg = 'El código debe tener 6 caracteres'
            print(f"❌ Error: {error_msg}")
            if request.is_json:
                return jsonify({'success': False, 'error': error_msg}), 400
            flash(error_msg, 'error')
            return render_template('UnirseASala.html')

        # Buscar la sala por PIN
        print(f"🔍 Buscando sala con código: {codigo_sala}")
        sala = controlador_salas.obtener_sala_por_codigo(codigo_sala)

        if not sala:
            error_msg = f'No se encontró ninguna sala con el código {codigo_sala}'
            print(f"❌ Error: {error_msg}")
            if request.is_json:
                return jsonify({'success': False, 'error': error_msg}), 404
            flash(error_msg, 'error')
            return render_template('UnirseASala.html')

        print(f"✅ Sala encontrada - ID: {sala['id']}, Estado: {sala['estado']}")

        # Verificar estado de la sala
        if sala['estado'] != 'esperando':
            error_msg = 'Esta sala no está disponible para unirse'
            if sala['estado'] == 'en_curso':
                error_msg = 'Esta sala ya está en curso'
            elif sala['estado'] == 'finalizada':
                error_msg = 'Esta sala ya ha finalizado'

            print(f"❌ Error: {error_msg}")
            if request.is_json:
                return jsonify({'success': False, 'error': error_msg}), 400
            flash(error_msg, 'error')
            return render_template('UnirseASala.html')

        # Guardar información en sesión y redirigir a página para ingresar nombre
        session['pin_sala_temp'] = codigo_sala
        session['sala_id_temp'] = sala['id']

        print(f"✅ Redirigiendo a /unirse-juego con PIN: {codigo_sala}")
        print(f"{'='*80}\n")

        # Si es JSON, devolver URL de redirección
        if request.is_json:
            return jsonify({
                'success': True,
                'redirect_url': url_for('unirse_juego') + f'?pin={codigo_sala}'
            })

        # Si es form normal, redirigir directamente
        return redirect(url_for('unirse_juego') + f'?pin={codigo_sala}')

    except Exception as e:
        error_msg = f'Error del sistema: {str(e)}'
        print(f"❌ ERROR en unirse_a_sala: {e}")
        import traceback
        traceback.print_exc()

        if request.is_json:
            return jsonify({'success': False, 'error': 'Error del servidor'}), 500
        flash('Error al procesar la solicitud', 'error')
        return render_template('UnirseASala.html')


# ==================== RUTAS DE API JWT ====================

@app.route('/api/auth', methods=['POST'])
def jwt_login():
    """Login que devuelve token JWT para acceso API"""
    try:
        data = request.get_json()
        if not data:
            return jsonify({'success': False, 'error': 'JSON requerido'}), 400

        email = (data.get('email') or '').strip().lower()
        password = data.get('password') or ''

        ok, resultado = controlador_usuario.autenticar_usuario(email, password)
        if not ok:
            return jsonify({'success': False, 'error': resultado}), 401

        usuario = resultado
        token = crear_token_jwt(str(usuario['id_usuario']))

        return jsonify({
            'success': True,
            'message': 'Autenticación exitosa',
            'access_token': token,
            'usuario': {
                'id_usuario': usuario['id_usuario'],
                'email': usuario['email'],
                'nombre': usuario['nombre'],
                'tipo_usuario': usuario['tipo_usuario']
            }
        }), 200

    except Exception as e:
        print(f"❌ ERROR en jwt_login: {e}")
        return jsonify({'success': False, 'error': 'Error del servidor'}), 500


@app.route('/login', methods=['GET', 'POST'])
def login():
    """
    Ruta de inicio de sesión con soporte para web y API.
    
    GET: Muestra formulario de login
    POST: Procesa credenciales y crea sesión
    """
    if request.method == 'GET':
        return render_template('Login.html')

    try:
        data = request.get_json(silent=True) or request.form.to_dict()
        email = (data.get('email') or '').strip().lower()
        password = data.get('password') or ''

        print(f"DEBUG: Intento de login para email: {email}")

        if not email or not password:
            error_msg = 'Email y contraseña son requeridos'
            if request.is_json:
                return jsonify({ 'success': False, 'error': error_msg }), 400
            flash(error_msg, 'error')
            return render_template('Login.html')

        ok, resultado = controlador_usuario.autenticar_usuario(email, password)
        if not ok:
            error_msg = resultado if isinstance(resultado, str) else 'Email o contraseña incorrectos'
            print(f"DEBUG: Login fallido para {email}: {error_msg}")
            if request.is_json:
                return jsonify({ 'success': False, 'error': error_msg }), 400
            flash(error_msg, 'error')
            return render_template('Login.html')

        usuario = resultado

        # Guardar información del usuario en la sesión
        session['usuario_id'] = usuario['id_usuario']
        session['usuario_email'] = usuario['email']
        session['usuario_nombre'] = usuario['nombre']
        session['usuario_apellidos'] = usuario['apellidos']
        session['usuario_tipo'] = usuario['tipo_usuario']
        session['logged_in'] = True
        session.permanent = True  # Hacer la sesión permanente (usa PERMANENT_SESSION_LIFETIME)

        print(f"DEBUG: Login exitoso para {email}, tipo: {usuario['tipo_usuario']}")

        # Determinar ruta de redirección
        if usuario['tipo_usuario'] == 'estudiante':
            redirect_url = url_for('dashboard_estudiante')
        elif usuario['tipo_usuario'] == 'docente':
            redirect_url = url_for('dashboard_docente')
        else:
            redirect_url = url_for('dashboard_admin')

        # Para peticiones web (no API), establecer cookies encriptadas
        if not request.is_json:
            response = make_response(redirect(redirect_url))
            establecer_cookies_usuario(response, usuario['id_usuario'], usuario['nombre'])
            return response
        else:
            return jsonify({ 'success': True, 'redirect': redirect_url })

    except Exception as e:
        error_msg = f'Error del sistema: {str(e)}'
        print(f"DEBUG: Error en login: {error_msg}")
        if request.is_json:
            return jsonify({ 'success': False, 'error': 'Error del sistema. Intenta nuevamente.' }), 500
        flash('Error del sistema. Intenta nuevamente.', 'error')
        return render_template('Login.html')

@app.route('/logout')
def logout():
    """
    Cierra la sesión del usuario y limpia cookies.
    """
    session.clear()
    flash('Has cerrado sesión correctamente', 'success')
    response = make_response(redirect(url_for('login')))
    limpiar_cookies_usuario(response)
    return response


# ==================== RUTAS DE DASHBOARDS ====================

def obtener_partidas_recientes_estudiante(usuario_id, limit=5):
    """Obtiene las partidas recientes de un estudiante (salas y juegos individuales)"""
    conexion = obtener_conexion()
    cursor = conexion.cursor()

    try:
        # Consulta que une participantes_sala con información del cuestionario
        query = """
            SELECT
                ps.id_participante,
                rs.puntaje_total,
                rs.tiempo_total_respuestas,
                ps.fecha_union,
                s.id_sala,
                s.pin_sala,
                s.modo_juego,
                c.id_cuestionario,
                c.titulo as cuestionario_titulo,
                c.descripcion as cuestionario_descripcion,
                (SELECT COUNT(*) FROM cuestionario_preguntas cp WHERE cp.id_cuestionario = c.id_cuestionario) as total_preguntas
            FROM participantes_sala ps
            INNER JOIN salas_juego s ON ps.id_sala = s.id_sala
            INNER JOIN cuestionarios c ON s.id_cuestionario = c.id_cuestionario
            LEFT JOIN ranking_sala rs ON ps.id_participante = rs.id_participante AND ps.id_sala = rs.id_sala
            WHERE ps.id_usuario = %s
            ORDER BY ps.fecha_union DESC
            LIMIT %s
        """

        cursor.execute(query, (usuario_id, limit))
        partidas = cursor.fetchall()

        # Formatear las partidas (PyMySQL devuelve tuplas, no diccionarios)
        partidas_formateadas = []
        for partida in partidas:
            es_individual = (partida[6] == 'individual')  # modo_juego

            partidas_formateadas.append({
                'id_participante': partida[0],
                'puntaje': partida[1] if partida[1] is not None else 0,
                'tiempo': partida[2] if partida[2] is not None else 0,
                'fecha': partida[3],
                'sala_id': partida[4],
                'sala_pin': partida[5],
                'modo_juego': partida[6],
                'es_individual': es_individual,
                'cuestionario_id': partida[7],
                'cuestionario_titulo': partida[8],
                'cuestionario_descripcion': partida[9],
                'total_preguntas': partida[10],
                'tipo': 'Individual' if es_individual else 'Multijugador'
            })

        return partidas_formateadas

    except Exception as e:
        print(f"Error en obtener_partidas_recientes_estudiante: {e}")
        import traceback
        traceback.print_exc()
        return []
    finally:
        cursor.close()
        conexion.close()

def obtener_estadisticas_estudiante(usuario_id):
    """Obtiene estadísticas reales del estudiante desde las tablas del sistema de juego"""
    conexion = obtener_conexion()
    cursor = conexion.cursor()

    try:
        # Total de participaciones
        cursor.execute("""
            SELECT COUNT(*)
            FROM participantes_sala
            WHERE id_usuario = %s
        """, (usuario_id,))
        total_participaciones = cursor.fetchone()[0] or 0

        # Promedio de puntaje
        cursor.execute("""
            SELECT AVG(rs.puntaje_total)
            FROM ranking_sala rs
            INNER JOIN participantes_sala ps ON rs.id_participante = ps.id_participante
            WHERE ps.id_usuario = %s AND rs.puntaje_total IS NOT NULL
        """, (usuario_id,))
        promedio_puntaje = cursor.fetchone()[0] or 0

        # Mejor posición
        cursor.execute("""
            SELECT MIN(rs.posicion)
            FROM ranking_sala rs
            INNER JOIN participantes_sala ps ON rs.id_participante = ps.id_participante
            WHERE ps.id_usuario = %s AND rs.posicion IS NOT NULL
        """, (usuario_id,))
        mejor_posicion = cursor.fetchone()[0] or 'N/A'

        # Recompensas obtenidas (de la tabla recompensas_otorgadas)
        cursor.execute("""
            SELECT COUNT(*)
            FROM recompensas_otorgadas
            WHERE id_estudiante = %s
        """, (usuario_id,))
        recompensas_obtenidas = cursor.fetchone()[0] or 0

        return {
            'total_participaciones': total_participaciones,
            'promedio_puntaje': round(promedio_puntaje, 1) if promedio_puntaje else 0,
            'mejor_posicion': mejor_posicion,
            'recompensas_obtenidas': recompensas_obtenidas
        }

    except Exception as e:
        print(f"Error en obtener_estadisticas_estudiante: {e}")
        import traceback
        traceback.print_exc()
        return {
            'total_participaciones': 0,
            'promedio_puntaje': 0,
            'mejor_posicion': 'N/A',
            'recompensas_obtenidas': 0
        }
    finally:
        cursor.close()
        conexion.close()


@app.route('/estudiante')
@login_required
@estudiante_required
def dashboard_estudiante():
    """
    Dashboard principal para estudiantes.
    Muestra estadísticas, cuestionarios disponibles y partidas recientes.
    """

    usuario = {
        'id_usuario': session['usuario_id'],
        'nombre': session['usuario_nombre'],
        'apellidos': session['usuario_apellidos'],
        'email': session['usuario_email'],
        'tipo_usuario': session['usuario_tipo']
    }

    # Obtener estadísticas del estudiante
    try:
        estadisticas = obtener_estadisticas_estudiante(session['usuario_id'])
    except Exception as e:
        print(f"Error obteniendo estadísticas: {e}")
        estadisticas = {
            'total_participaciones': 0,
            'promedio_puntaje': 0,
            'mejor_posicion': 'N/A',
            'recompensas_obtenidas': 0
        }

    # Obtener todos los cuestionarios publicados
    try:
        todos_cuestionarios = controlador_cuestionarios.obtener_cuestionarios_publicados()
    except:
        todos_cuestionarios = []

    # Obtener partidas recientes del estudiante
    try:
        partidas_recientes = obtener_partidas_recientes_estudiante(session['usuario_id'])
    except Exception as e:
        print(f"Error obteniendo partidas recientes: {e}")
        partidas_recientes = []

    return render_template('DashboardEstudiante.html',
                         usuario=usuario,
                         estadisticas=estadisticas,
                         cuestionarios=todos_cuestionarios,
                         partidas_recientes=partidas_recientes)

@app.route('/admin')
@login_required
@admin_required
def dashboard_admin():
    """Dashboard para administradores, maestros, etc."""
    usuario = {
        'id_usuario': session['usuario_id'],
        'nombre': session['usuario_nombre'],
        'apellidos': session['usuario_apellidos'],
        'email': session['usuario_email'],
        'tipo_usuario': session['usuario_tipo']
    }

    # Si es docente, redirigir al dashboard de docente
    if session['usuario_tipo'] == 'docente':
        return redirect(url_for('dashboard_docente'))

    # Obtener estadísticas del sistema para administradores
    try:
        estadisticas = {
            'total_usuarios': len(controlador_usuario.obtener_todos_usuarios()),
            'total_cuestionarios': len(controlador_cuestionarios.obtener_cuestionarios()),
            'salas_activas': len(controlador_salas.obtener_salas_activas()),
            'total_participaciones': len(controlador_participaciones.obtener_participaciones()),
            'total_preguntas': len(controlador_preguntas.obtener_preguntas())
        }
    except:
        estadisticas = {
            'total_usuarios': 0,
            'total_cuestionarios': 0,
            'salas_activas': 0,
            'total_participaciones': 0,
            'total_preguntas': 0
        }

    return render_template('DashboardAdmin.html', usuario=usuario, estadisticas=estadisticas)


@app.route('/docente')
@login_required
@docente_required
def dashboard_docente():
    """
    Dashboard principal para docentes.
    Muestra cuestionarios, estadísticas y ranking de estudiantes.
    """
    usuario = {
        'id_usuario': session['usuario_id'],
        'nombre': session['usuario_nombre'],
        'apellidos': session['usuario_apellidos'],
        'email': session['usuario_email'],
        'tipo_usuario': session['usuario_tipo']
    }

    # Obtener cuestionarios del docente
    try:
        id_docente = session['usuario_id']
        cuestionarios = obtener_cuestionarios_por_docente_simple(id_docente)

        # Obtener ranking global de los cuestionarios del docente
        ranking_global = controlador_ranking.obtener_ranking_global_por_docente(id_docente)

        # Estadísticas específicas del docente
        estadisticas = {
            'total_cuestionarios': len(cuestionarios),
            'cuestionarios_activos': len([c for c in cuestionarios if c.get('estado') == 'activo']),
            'total_estudiantes': len(ranking_global),  # Estudiantes únicos que han participado
            'total_participaciones': sum(e.get('total_participaciones', 0) for e in ranking_global),
            'total_respuestas': sum(e.get('total_respuestas', 0) for e in ranking_global),
            'promedio_calificaciones': round(sum(e.get('promedio_puntaje', 0) for e in ranking_global) / len(ranking_global), 1) if ranking_global else 0
        }

    except Exception as e:
        print(f"Error obteniendo datos del docente: {e}")
        import traceback
        traceback.print_exc()
        cuestionarios = []
        ranking_global = []
        estadisticas = {
            'total_cuestionarios': 0,
            'cuestionarios_activos': 0,
            'total_estudiantes': 0,
            'total_participaciones': 0,
            'total_respuestas': 0,
            'promedio_calificaciones': 0
        }

    return render_template('DashboardDocente.html',
                         usuario=usuario,
                         estadisticas=estadisticas,
                         cuestionarios=cuestionarios,
                         ranking_global=ranking_global)

# ==================== RUTAS DE EXPORTACIÓN PARA DASHBOARD DOCENTE ====================

@app.route('/api/exportar-dashboard-docente/excel', methods=['POST'])
@login_required
@docente_required
def exportar_dashboard_docente_excel():
    """Exportar ranking global del docente a Excel"""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        
        id_docente = session['usuario_id']
        
        # Obtener ranking global del docente
        ranking_global = controlador_ranking.obtener_ranking_global_por_docente(id_docente)
        
        # Obtener información del docente
        conexion = obtener_conexion()
        cursor = conexion.cursor()
        cursor.execute("""
            SELECT nombre, apellidos FROM usuarios WHERE id_usuario = %s
        """, (id_docente,))
        docente = cursor.fetchone()
        nombre_docente = f"{docente[0]} {docente[1]}" if docente else "Docente"
        cursor.close()
        conexion.close()
        
        # Crear workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Ranking Global"
        
        # Estilos
        header_fill = PatternFill(start_color="667eea", end_color="667eea", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=12)
        title_font = Font(bold=True, size=16, color="667eea")
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Título
        ws.merge_cells('A1:F1')
        title_cell = ws['A1']
        title_cell.value = f"📊 Ranking Global de Estudiantes - {nombre_docente}"
        title_cell.font = title_font
        title_cell.alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[1].height = 30
        
        # Fecha
        ws.merge_cells('A2:F2')
        fecha_cell = ws['A2']
        fecha_cell.value = f"Fecha: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
        fecha_cell.alignment = Alignment(horizontal='center')
        ws.row_dimensions[2].height = 20
        
        # Encabezados
        headers = ['Posición', 'Nombre Completo', 'Puntaje Total', 'Participaciones', 'Promedio', 'Precisión']
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=4, column=col)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border
        
        ws.row_dimensions[4].height = 25
        
        # Datos
        for idx, estudiante in enumerate(ranking_global, start=5):
            row_data = [
                estudiante['posicion'],
                estudiante['nombre_completo'],
                estudiante['puntaje_acumulado'],
                estudiante['total_participaciones'],
                f"{estudiante['promedio_puntaje']:.1f}",
                f"{estudiante['precision_global']:.1f}%"
            ]
            
            for col, value in enumerate(row_data, start=1):
                cell = ws.cell(row=idx, column=col)
                cell.value = value
                cell.border = border
                cell.alignment = Alignment(horizontal='center', vertical='center')
                
                # Colorear top 3
                if estudiante['posicion'] <= 3:
                    colors = {1: "FFD700", 2: "C0C0C0", 3: "CD7F32"}
                    cell.fill = PatternFill(start_color=colors[estudiante['posicion']],
                                          end_color=colors[estudiante['posicion']],
                                          fill_type="solid")
        
        # Ajustar anchos de columna
        column_widths = [12, 30, 18, 18, 15, 15]
        for col, width in enumerate(column_widths, start=1):
            ws.column_dimensions[chr(64 + col)].width = width
        
        # Guardar en memoria
        excel_file = BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)
        
        return send_file(
            excel_file,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=f'ranking_global_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        )
    
    except Exception as e:
        print(f"ERROR exportar_dashboard_docente_excel: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/exportar-dashboard-docente/onedrive', methods=['POST'])
@login_required
@docente_required
def exportar_dashboard_docente_onedrive():
    """Exportar ranking global del docente a OneDrive"""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        
        # Verificar que msal está disponible
        if not MSAL_AVAILABLE:
            return jsonify({
                'success': False,
                'error': 'La integración con OneDrive no está disponible. Instala msal: pip install msal'
            }), 500
        
        id_docente = session['usuario_id']
        
        # Obtener ranking global del docente
        ranking_global = controlador_ranking.obtener_ranking_global_por_docente(id_docente)
        
        # Obtener información del docente
        conexion = obtener_conexion()
        cursor = conexion.cursor()
        cursor.execute("""
            SELECT nombre, apellidos, email FROM usuarios WHERE id_usuario = %s
        """, (id_docente,))
        docente = cursor.fetchone()
        nombre_docente = f"{docente[0]} {docente[1]}" if docente else "Docente"
        email_docente = docente[2] if docente else None
        cursor.close()
        conexion.close()
        
        # Crear archivo Excel
        wb = Workbook()
        ws = wb.active
        ws.title = "Ranking Global"
        
        # Estilos
        header_fill = PatternFill(start_color="667eea", end_color="667eea", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=12)
        title_font = Font(bold=True, size=16, color="667eea")
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Título
        ws.merge_cells('A1:F1')
        title_cell = ws['A1']
        title_cell.value = f"📊 Ranking Global de Estudiantes - {nombre_docente}"
        title_cell.font = title_font
        title_cell.alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[1].height = 30
        
        # Fecha
        ws.merge_cells('A2:F2')
        fecha_cell = ws['A2']
        fecha_cell.value = f"Fecha: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
        fecha_cell.alignment = Alignment(horizontal='center')
        ws.row_dimensions[2].height = 20
        
        # Encabezados y datos (igual que la versión Excel)
        headers = ['Posición', 'Nombre Completo', 'Puntaje Total', 'Participaciones', 'Promedio', 'Precisión']
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=4, column=col)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border
        
        for idx, estudiante in enumerate(ranking_global, start=5):
            row_data = [
                estudiante['posicion'],
                estudiante['nombre_completo'],
                estudiante['puntaje_acumulado'],
                estudiante['total_participaciones'],
                f"{estudiante['promedio_puntaje']:.1f}",
                f"{estudiante.get('precision_global', 0):.1f}%"
            ]
            
            for col, value in enumerate(row_data, start=1):
                cell = ws.cell(row=idx, column=col)
                cell.value = value
                cell.border = border
                cell.alignment = Alignment(horizontal='center', vertical='center')
                
                if estudiante['posicion'] <= 3:
                    colors = {1: "FFD700", 2: "C0C0C0", 3: "CD7F32"}
                    cell.fill = PatternFill(start_color=colors[estudiante['posicion']],
                                          end_color=colors[estudiante['posicion']],
                                          fill_type="solid")
        
        column_widths = [12, 30, 18, 18, 15, 15]
        for col, width in enumerate(column_widths, start=1):
            ws.column_dimensions[chr(64 + col)].width = width
        
        # Guardar en memoria
        excel_buffer = BytesIO()
        wb.save(excel_buffer)
        excel_content = excel_buffer.getvalue()
        
        # Subir a OneDrive
        filename = f"ranking_global_brainrush_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        
        # Obtener access token
        access_token = app.config.get('ONEDRIVE_ACCESS_TOKEN')
        if not access_token:
            return jsonify({
                'success': False,
                'error': 'OneDrive no está configurado. Por favor, configura la integración con OneDrive primero.'
            }), 400
        
        # Subir archivo a OneDrive
        upload_url = 'https://graph.microsoft.com/v1.0/me/drive/root:/BrainRUSH/' + filename + ':/content'
        headers = {
            'Authorization': f'Bearer {access_token}',
            'Content-Type': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        }
        
        response = requests.put(upload_url, headers=headers, data=excel_content)
        
        if response.status_code in [200, 201]:
            file_data = response.json()
            web_url = file_data.get('webUrl', '')
            
            # Enviar email con el link (si está configurado)
            if app.config.get('MAIL_ENABLED') and email_docente:
                try:
                    from flask_mail import Message
                    msg = Message(
                        '📊 Ranking Global exportado a OneDrive - Brain RUSH',
                        sender=app.config['MAIL_DEFAULT_SENDER'],
                        recipients=[email_docente]
                    )
                    msg.html = f"""
                    <div style="font-family: Arial, sans-serif; max-width: 600px; margin: 0 auto;">
                        <h2 style="color: #667eea;">📊 Ranking Global Exportado</h2>
                        <p>Hola {nombre_docente},</p>
                        <p>Tu ranking global de estudiantes ha sido exportado exitosamente a OneDrive.</p>
                        <p><strong>Archivo:</strong> {filename}</p>
                        <p><strong>Fecha:</strong> {datetime.now().strftime('%d/%m/%Y %H:%M')}</p>
                        <p style="margin: 30px 0;">
                            <a href="{web_url}" style="background: #667eea; color: white; padding: 12px 24px; text-decoration: none; border-radius: 5px; display: inline-block;">
                                📂 Abrir en OneDrive
                            </a>
                        </p>
                        <p style="color: #666; font-size: 12px;">Brain RUSH - Sistema de Evaluación Gamificada</p>
                    </div>
                    """
                    mail.send(msg)
                except Exception as e:
                    print(f"Error enviando email: {e}")
            
            return jsonify({
                'success': True,
                'message': 'Ranking exportado exitosamente a OneDrive',
                'web_url': web_url
            })
        else:
            return jsonify({
                'success': False,
                'error': f'Error al subir a OneDrive: {response.status_code}'
            }), 500
    
    except Exception as e:
        print(f"ERROR exportar_dashboard_docente_onedrive: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/exportar-dashboard-docente/pdf', methods=['POST'])
@login_required
@docente_required
def exportar_dashboard_docente_pdf():
    """Generar reporte PDF del dashboard del docente"""
    try:
        # Intentar importar reportlab
        try:
            from reportlab.lib.pagesizes import letter, A4
            from reportlab.lib import colors
            from reportlab.lib.units import inch
            from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            from reportlab.lib.enums import TA_CENTER, TA_LEFT
        except ImportError:
            return jsonify({
                'success': False,
                'error': 'Librería reportlab no instalada. Ejecuta: pip install reportlab'
            }), 500
        
        id_docente = session['usuario_id']
        
        # Obtener datos
        ranking_global = controlador_ranking.obtener_ranking_global_por_docente(id_docente)
        cuestionarios = obtener_cuestionarios_por_docente_simple(id_docente)
        
        # Obtener información del docente
        conexion = obtener_conexion()
        cursor = conexion.cursor()
        cursor.execute("""
            SELECT nombre, apellidos FROM usuarios WHERE id_usuario = %s
        """, (id_docente,))
        docente = cursor.fetchone()
        nombre_docente = f"{docente[0]} {docente[1]}" if docente else "Docente"
        cursor.close()
        conexion.close()
        
        # Crear PDF en memoria
        pdf_buffer = BytesIO()
        doc = SimpleDocTemplate(pdf_buffer, pagesize=A4)
        story = []
        
        # Estilos
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=24,
            textColor=colors.HexColor('#667eea'),
            spaceAfter=30,
            alignment=TA_CENTER
        )
        
        heading_style = ParagraphStyle(
            'CustomHeading',
            parent=styles['Heading2'],
            fontSize=16,
            textColor=colors.HexColor('#667eea'),
            spaceAfter=12
        )
        
        # Título
        story.append(Paragraph(f"📊 Reporte Brain RUSH", title_style))
        story.append(Paragraph(f"<b>Docente:</b> {nombre_docente}", styles['Normal']))
        story.append(Paragraph(f"<b>Fecha:</b> {datetime.now().strftime('%d/%m/%Y %H:%M')}", styles['Normal']))
        story.append(Spacer(1, 0.3*inch))
        
        # Estadísticas generales
        story.append(Paragraph("📈 Estadísticas Generales", heading_style))
        stats_data = [
            ['Cuestionarios Creados', len(cuestionarios)],
            ['Cuestionarios Activos', len([c for c in cuestionarios if c.get('estado') == 'activo' or c.get('estado') == 'publicado'])],
            ['Estudiantes Participantes', len(ranking_global)],
            ['Promedio General', f"{sum(e['promedio_puntaje'] for e in ranking_global) / len(ranking_global):.1f}%" if ranking_global else "0%"]
        ]
        
        stats_table = Table(stats_data, colWidths=[3*inch, 2*inch])
        stats_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, -1), colors.HexColor('#f0f0f0')),
            ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 12),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 12),
            ('GRID', (0, 0), (-1, -1), 1, colors.grey)
        ]))
        story.append(stats_table)
        story.append(Spacer(1, 0.3*inch))
        
        # Ranking Global
        story.append(Paragraph("🏆 Ranking Global de Estudiantes", heading_style))
        
        if ranking_global:
            # Encabezados de tabla
            ranking_data = [['Pos', 'Nombre', 'Puntaje', 'Partidas', 'Promedio', 'Precisión']]
            
            # Datos
            for estudiante in ranking_global[:20]:  # Top 20
                ranking_data.append([
                    str(estudiante['posicion']),
                    estudiante['nombre_completo'][:25],
                    str(estudiante['puntaje_acumulado']),
                    str(estudiante['total_participaciones']),
                    f"{estudiante['promedio_puntaje']:.1f}",
                    f"{estudiante.get('precision_global', 0):.1f}%"
                ])
            
            ranking_table = Table(ranking_data, colWidths=[0.5*inch, 2*inch, 1*inch, 1*inch, 1*inch, 1*inch])
            ranking_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#667eea')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 10),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.grey),
                ('FONTSIZE', (0, 1), (-1, -1), 9)
            ]))
            
            # Colorear top 3
            if len(ranking_data) > 1:
                ranking_table.setStyle(TableStyle([
                    ('BACKGROUND', (0, 1), (-1, 1), colors.HexColor('#FFD700')),
                ]))
            if len(ranking_data) > 2:
                ranking_table.setStyle(TableStyle([
                    ('BACKGROUND', (0, 2), (-1, 2), colors.HexColor('#C0C0C0')),
                ]))
            if len(ranking_data) > 3:
                ranking_table.setStyle(TableStyle([
                    ('BACKGROUND', (0, 3), (-1, 3), colors.HexColor('#CD7F32')),
                ]))
            
            story.append(ranking_table)
        else:
            story.append(Paragraph("No hay datos de ranking disponibles", styles['Normal']))
        
        # Construir PDF
        doc.build(story)
        pdf_buffer.seek(0)
        
        return send_file(
            pdf_buffer,
            mimetype='application/pdf',
            as_attachment=True,
            download_name=f'reporte_brainrush_{datetime.now().strftime("%Y%m%d_%H%M%S")}.pdf'
        )
    
    except Exception as e:
        print(f"ERROR exportar_dashboard_docente_pdf: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/dashboard')
@login_required
def dashboard():
    """
    Ruta genérica de dashboard que redirecciona según el tipo de usuario.
    Útil para enlaces genéricos que necesitan determinar el dashboard apropiado.
    """
    print(f"\n{'='*80}")
    print(f"🏠 DASHBOARD - Redirigiendo según tipo de usuario")
    print(f"   usuario_tipo: {session.get('usuario_tipo')}")
    print(f"   tipo_usuario: {session.get('tipo_usuario')}")
    print(f"{'='*80}\n")
    
    if session.get('usuario_tipo') == 'estudiante':
        print("   → Redirigiendo a dashboard_estudiante")
        return redirect(url_for('dashboard_estudiante'))
    elif session.get('usuario_tipo') == 'docente':
        print("   → Redirigiendo a dashboard_docente")
        return redirect(url_for('dashboard_docente'))
    else:
        print("   → Redirigiendo a dashboard_admin")
        return redirect(url_for('dashboard_admin'))


# ==================== RUTAS ADICIONALES PARA ESTUDIANTES ====================

@app.route('/historial/estudiante')
@login_required
def historial_estudiante():
    """Historial de participaciones del estudiante"""
    if session.get('usuario_tipo') != 'estudiante':
        return redirect(url_for('dashboard_admin'))

    try:
        import pymysql.cursors
        conexion = obtener_conexion()
        cursor = conexion.cursor(pymysql.cursors.DictCursor)

        # Obtener historial del estudiante desde participantes_sala y ranking_sala
        query = '''
            SELECT
                p.id_participante,
                p.nombre_participante,
                s.id_sala,
                s.pin_sala,
                s.estado as estado_sala,
                c.titulo as titulo_cuestionario,
                c.id_cuestionario,
                s.total_preguntas,
                s.fecha_creacion as fecha_inicio,
                r.puntaje_total,
                r.respuestas_correctas,
                r.tiempo_total_respuestas,
                r.posicion,
                g.nombre_grupo
            FROM participantes_sala p
            INNER JOIN salas_juego s ON p.id_sala = s.id_sala
            INNER JOIN cuestionarios c ON s.id_cuestionario = c.id_cuestionario
            LEFT JOIN ranking_sala r ON p.id_participante = r.id_participante AND p.id_sala = r.id_sala
            LEFT JOIN grupos_sala g ON p.id_grupo = g.id_grupo
            WHERE p.id_usuario = %s
            AND s.estado = 'finalizada'
            ORDER BY s.fecha_creacion DESC
        '''

        cursor.execute(query, (session['usuario_id'],))
        participaciones_raw = cursor.fetchall()

        # Procesar datos para calcular precisión y duración
        participaciones = []
        for part in participaciones_raw:
            # Calcular precisión
            if part['total_preguntas'] and part['total_preguntas'] > 0:
                precision = round((part['respuestas_correctas'] / part['total_preguntas']) * 100)
            else:
                precision = 0

            # Formatear duración
            if part['tiempo_total_respuestas']:
                minutos = int(part['tiempo_total_respuestas'] // 60)
                segundos = int(part['tiempo_total_respuestas'] % 60)
                duracion = f"{minutos}m {segundos}s"
            else:
                duracion = "N/A"

            participaciones.append({
                'id_participante': part['id_participante'],
                'nombre_participante': part['nombre_participante'],
                'titulo_cuestionario': part['titulo_cuestionario'],
                'fecha_inicio': part['fecha_inicio'],
                'puntaje_total': part['puntaje_total'] or 0,
                'respuestas_correctas': part['respuestas_correctas'] or 0,
                'total_preguntas': part['total_preguntas'] or 0,
                'precision': precision,
                'duracion': duracion,
                'posicion': f"#{part['posicion']}" if part['posicion'] else "N/A",
                'nombre_grupo': part['nombre_grupo'],
                'pin_sala': part['pin_sala']
            })

        cursor.close()
        conexion.close()

    except Exception as e:
        print(f"ERROR en historial_estudiante: {e}")
        import traceback
        traceback.print_exc()
        participaciones = []

    return render_template('HistorialEstudiante.html', participaciones=participaciones)

@app.route('/ranking/global')
@login_required
def ranking_global():
    """Ranking global del sistema"""
    try:
        ranking = controlador_ranking.obtener_ranking_global()

        # Si el usuario es estudiante, encontrar su posición
        usuario_actual = None
        if session.get('usuario_tipo') == 'estudiante':
            usuario_id = session.get('usuario_id')
            for estudiante in ranking:
                if estudiante['id_usuario'] == usuario_id:
                    usuario_actual = estudiante
                    break

        return render_template('RankingGlobal.html',
                             ranking=ranking,
                             usuario_actual=usuario_actual)
    except Exception as e:
        print(f"Error en ranking_global: {e}")
        import traceback
        traceback.print_exc()
        return render_template('RankingGlobal.html', ranking=[], usuario_actual=None)

# ==================== RUTAS DEL SISTEMA DE XP E INSIGNIAS ====================

@app.route('/api/perfil-xp/<int:usuario_id>')
def obtener_perfil_xp_api(usuario_id):
    """API: Obtiene el perfil completo de XP de un usuario"""
    try:
        perfil = controlador_xp.obtener_perfil_xp(usuario_id)
        if not perfil:
            return jsonify({'success': False, 'error': 'Usuario no encontrado'}), 404

        return jsonify({'success': True, 'perfil': perfil})
    except Exception as e:
        print(f"Error en obtener_perfil_xp_api: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/insignias/<int:usuario_id>')
def obtener_insignias_api(usuario_id):
    """API: Obtiene TODAS las insignias (desbloqueadas y bloqueadas) de un usuario"""
    try:
        insignias = controlador_xp.obtener_todas_insignias_usuario(usuario_id)
        return jsonify({'success': True, 'insignias': insignias})
    except Exception as e:
        print(f"Error en obtener_insignias_api: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/insignias-progreso/<int:usuario_id>')
def obtener_progreso_insignias_api(usuario_id):
    """API: Obtiene el progreso hacia insignias no desbloqueadas"""
    try:
        progreso = controlador_xp.obtener_progreso_insignias(usuario_id)
        return jsonify({'success': True, 'insignias_bloqueadas': progreso})
    except Exception as e:
        print(f"Error en obtener_progreso_insignias_api: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/ranking-xp')
def obtener_ranking_xp_api():
    """API: Obtiene el ranking global de XP"""
    try:
        limite = request.args.get('limite', 100, type=int)
        ranking = controlador_xp.obtener_ranking_global(limite)
        return jsonify({'success': True, 'ranking': ranking})
    except Exception as e:
        print(f"Error en obtener_ranking_xp_api: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/ranking-xp')
@login_required
def pagina_ranking_xp():
    """Página del ranking de XP"""
    try:
        ranking = controlador_xp.obtener_ranking_global(100)

        # Si es estudiante, encontrar su posición
        usuario_actual = None
        if session.get('usuario_tipo') == 'estudiante':
            usuario_id = session.get('usuario_id')
            for estudiante in ranking:
                if estudiante['id_usuario'] == usuario_id:
                    usuario_actual = estudiante
                    break

        return render_template('RankingXP.html',
                             ranking=ranking,
                             usuario_actual=usuario_actual)
    except Exception as e:
        print(f"Error en pagina_ranking_xp: {e}")
        import traceback
        traceback.print_exc()
        return render_template('RankingXP.html', ranking=[], usuario_actual=None)

@app.route('/mis-insignias')
@login_required
def mis_insignias():
    """Página de insignias del estudiante"""
    if session.get('usuario_tipo') != 'estudiante':
        flash('Solo los estudiantes tienen acceso a insignias', 'error')
        return redirect(url_for('maestra'))

    try:
        usuario_id = session.get('usuario_id')

        # Obtener perfil XP
        perfil = controlador_xp.obtener_perfil_xp(usuario_id)

        # Obtener insignias desbloqueadas
        insignias = controlador_xp.obtener_insignias_usuario(usuario_id)

        # Obtener progreso de insignias bloqueadas
        insignias_bloqueadas = controlador_xp.obtener_progreso_insignias(usuario_id)

        return render_template('MisInsignias.html',
                             perfil=perfil,
                             insignias=insignias,
                             insignias_bloqueadas=insignias_bloqueadas)
    except Exception as e:
        print(f"Error en mis_insignias: {e}")
        import traceback
        traceback.print_exc()
        flash('Error al cargar insignias. Por favor, intenta de nuevo.', 'error')
        return redirect(url_for('dashboard_estudiante'))

@app.route('/tienda-insignias')
@login_required
def tienda_insignias():
    """Página de tienda de insignias"""
    if session.get('usuario_tipo') != 'estudiante':
        flash('Solo los estudiantes tienen acceso a la tienda', 'error')
        return redirect(url_for('maestra'))

    return render_template('TiendaInsignias.html')

@app.route('/api/tienda-insignias')
@login_required
def obtener_tienda_insignias_api():
    """API: Obtiene insignias disponibles para comprar"""
    try:
        insignias = controlador_xp.obtener_insignias_tienda()
        usuario_id = session.get('usuario_id')

        # Obtener XP del usuario (usar XP total acumulado para la tienda)
        perfil = controlador_xp.obtener_perfil_xp(usuario_id)
        xp_total_acumulado = perfil['xp_total_acumulado'] if perfil else 0

        # Marcar cuáles ya tiene el usuario
        insignias_usuario = controlador_xp.obtener_insignias_usuario(usuario_id)
        ids_tenidas = [i['id'] for i in insignias_usuario]

        for insignia in insignias:
            insignia['ya_comprada'] = insignia['id_insignia'] in ids_tenidas
            insignia['puede_comprar'] = xp_total_acumulado >= insignia['precio_xp'] and not insignia['ya_comprada']

        return jsonify({
            'success': True,
            'insignias': insignias,
            'xp_actual': xp_total_acumulado  # Enviar XP total acumulado al frontend
        })
    except Exception as e:
        print(f"Error en obtener_tienda_insignias_api: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/comprar-insignia', methods=['POST'])
@login_required
@jwt_or_session_required
def comprar_insignia_api():
    """API: Compra una insignia con XP - Requiere autenticación"""
    try:
        data = request.get_json()
        id_insignia = data.get('id_insignia')
        usuario_id = session.get('usuario_id')

        if not id_insignia:
            return jsonify({'success': False, 'error': 'ID de insignia requerido'}), 400

        resultado = controlador_xp.comprar_insignia(usuario_id, id_insignia)

        if resultado['success']:
            return jsonify(resultado)
        else:
            return jsonify(resultado), 400

    except Exception as e:
        print(f"Error en comprar_insignia_api: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/mis-recompensas')
@login_required
def mis_recompensas():
    """Recompensas del estudiante"""
    if session.get('usuario_tipo') != 'estudiante':
        return redirect(url_for('dashboard_admin'))

    try:
        recompensas = controlador_recompensas.verificar_recompensas_disponibles(session['usuario_id'])
    except:
        recompensas = []

    return render_template('MisRecompensas.html', recompensas=recompensas)

# ==================== RUTAS PARA ADMINISTRADORES ====================

@app.route('/monitoreo/salas')
@login_required
@admin_required
def monitoreo_salas():
    """Monitoreo de salas activas"""
    try:
        salas = controlador_salas.obtener_salas_activas()
    except:
        salas = []

    return render_template('MonitoreoSalas.html', salas=salas)

@app.route('/reportes/sistema')
@login_required
@admin_required
def reportes_sistema():
    """Reportes y estadísticas del sistema"""
    try:
        estadisticas = controlador_cuestionarios.obtener_estadisticas_sistema()
    except:
        estadisticas = {}

    return render_template('ReportesSystem.html', estadisticas=estadisticas)

@app.route('/configuracion/sistema')
@login_required
@admin_required
def configuracion_sistema():
    """Configuración del sistema"""
    return render_template('ConfiguracionSistema.html')

@app.route('/backup/datos')
@login_required
@admin_required
def backup_datos():
    """Backup de datos"""
    return render_template('BackupDatos.html')

@app.route('/logs/sistema')
@login_required
@admin_required
def logs_sistema():
    """Logs del sistema"""
    return render_template('LogsSistema.html')

@app.route('/otorgar-recompensas')
@login_required
@admin_required
def otorgar_recompensas():
    """Otorgar recompensas"""
    return render_template('OtorgarRecompensas.html')


# ==================== RUTAS DE SALAS Y JUEGO ====================

@app.route('/crear-sala', methods=['GET', 'POST'])
@login_required
@docente_required
def crear_sala_route():
    # Validar que solo los docentes pueden crear salas
    if session.get('usuario_tipo') != 'docente':
        if request.method == 'GET':
            flash('Solo los docentes pueden crear salas de juego', 'error')
            return redirect(url_for('maestra'))
        return jsonify({'success': False, 'error': 'Solo los docentes pueden crear salas'}), 403

    if request.method == 'GET':
        return render_template('CrearSala.html')

    data = request.get_json(silent=True) or request.form.to_dict()
    try:
        nombre = data.get('nombre', '').strip()
        cuestionario_id = int(data.get('cuestionario_id'))
        tipo_sala = data.get('tipo_sala', 'individual')
        max_participantes = int(data.get('max_participantes', 30))
        tiempo_respuesta = int(data.get('tiempo_respuesta', 30))
        configuracion = data.get('configuracion', {})

        if not nombre:
            raise ValueError("El nombre de la sala es requerido")

        # Verificar que el cuestionario existe y pertenece al docente
        cuestionario = controlador_cuestionarios.obtener_cuestionario_por_id(cuestionario_id)
        if not cuestionario:
            raise ValueError("El cuestionario seleccionado no existe")

        # Crear la sala
        sala_id = controlador_salas.crear_sala(
            nombre=nombre,
            cuestionario_id=cuestionario_id,
            docente_id=session.get('user_id', 1),  # Por ahora usar 1 como default
            tipo_sala=tipo_sala,
            max_participantes=max_participantes,
            tiempo_respuesta=tiempo_respuesta
        )

        if request.is_json:
            return jsonify({'success': True, 'sala_id': sala_id, 'message': 'Sala creada exitosamente'})
        flash('¡Sala creada exitosamente!', 'success')
        return redirect(url_for('monitorear_sala', sala_id=sala_id))
    except Exception as e:
        if request.is_json:
            return jsonify({'success': False, 'error': str(e)}), 400
        flash(f'Error al crear la sala: {str(e)}', 'error')
        return redirect(url_for('error_sistema_page'))

@app.route('/mis-salas')
@login_required
@docente_required
def mis_salas():
    try:
        # Por ahora mostrar todas las salas, luego filtrar por docente
        salas = controlador_salas.obtener_salas()
        return render_template('MisSalas.html', salas=salas)
    except Exception as e:
        flash(f'Error al obtener las salas: {str(e)}', 'error')
        return redirect(url_for('error_sistema_page'))

@app.route('/sala/<string:codigo>/unirse', methods=['GET', 'POST'])
def unirse_sala_route(codigo):
    if request.method == 'GET':
        return render_template('UnirseASala.html')

    data = request.get_json(silent=True) or request.form.to_dict()
    try:
        nombre_participante = data.get('nombre_participante', '').strip()

        if not nombre_participante:
            raise ValueError("El nombre del participante es requerido")

        # Buscar la sala por código
        sala = controlador_salas.obtener_sala_por_codigo(codigo)
        if not sala:
            raise ValueError("Sala no encontrada")

        if sala['estado'] != 'esperando':
            raise ValueError("La sala no está disponible para unirse")

        # Verificar límite de participantes
        participantes_actuales = controlador_salas.obtener_participantes_sala(sala['id'])
        if len(participantes_actuales) >= sala['max_participantes']:
            raise ValueError("La sala ha alcanzado el límite de participantes")

        # Agregar el participante a la sala
        participante_id = controlador_salas.agregar_participante_sala(sala['id'], nombre_participante)

        session['participante_id'] = participante_id
        session['sala_id'] = sala['id']
        session['nombre_participante'] = nombre_participante

        if request.is_json:
            return jsonify({'success': True, 'participante_id': participante_id})
        return redirect(url_for('esperar_juego', sala_id=sala['id']))
    except Exception as e:
        if request.is_json:
            return jsonify({'success': False, 'error': str(e)}), 400
        flash(f'Error al unirse a la sala: {str(e)}', 'error')
        return redirect(url_for('unirse_sala_route', codigo=codigo))

@app.route('/sala/<int:sala_id>/monitorear')
@app.route('/monitorear-sala/<int:sala_id>')
@login_required
@docente_required
def monitorear_sala(sala_id):
    try:
        print(f"DEBUG: monitorear_sala - sala_id: {sala_id}")

        # Usar la función simple que ya existe
        sala = obtener_sala_por_id_simple(sala_id)
        print(f"DEBUG: Sala obtenida: {sala}")

        if not sala:
            print("DEBUG: Sala no encontrada")
            flash('Sala no encontrada', 'error')
            return redirect(url_for('mis_cuestionarios'))

        # Obtener participantes reales de la sala
        participantes = controlador_salas.obtener_participantes_sala(sala['id'])
        print(f"DEBUG: Participantes: {participantes}")

        # Verificar si existe el template MonitoreoJuego.html
        try:
            return render_template('MonitoreoJuego.html', sala=sala, participantes=participantes)
        except Exception as template_error:
            print(f"DEBUG: Error de template: {template_error}")
            # Template temporal hasta que se implemente MonitoreoJuego.html
            return render_template('MonitoreoJuego.html',
                                 sala=sala,
                                 participantes=participantes)

    except Exception as e:
        print(f"DEBUG: Error en monitorear_sala: {str(e)}")
        import traceback
        traceback.print_exc()
        flash(f'Error al cargar la sala: {str(e)}', 'error')
        return redirect(url_for('mis_cuestionarios'))

@app.route('/sala/<int:sala_id>/iniciar', methods=['POST'])
@login_required
@docente_required
def iniciar_sala(sala_id):
    try:
        print(f"\n{'='*80}")
        print(f"🎮 DOCENTE INICIANDO JUEGO - Sala ID: {sala_id}")
        print(f"{'='*80}\n")
        
        resultado = controlador_juego.iniciar_juego_sala(sala_id)
        
        print(f"Resultado de iniciar_juego_sala: {resultado}")

        if not resultado:
            raise ValueError("No se pudo iniciar la sala")

        # Verificar que el estado cambió correctamente
        sala_verificar = obtener_sala_por_id_simple(sala_id)
        print(f"✅ Estado actual de la sala después de iniciar: {sala_verificar.get('estado')} (debe ser 'en_curso')")
        print(f"{'='*80}\n")

        if request.is_json:
            return jsonify({'success': True, 'message': 'Sala iniciada exitosamente'})
        flash('¡Sala iniciada exitosamente!', 'success')
        return redirect(url_for('monitorear_sala', sala_id=sala_id))
    except Exception as e:
        print(f"❌ ERROR al iniciar sala: {e}")
        import traceback
        traceback.print_exc()
        if request.is_json:
            return jsonify({'success': False, 'error': str(e)}), 400
        flash(f'Error al iniciar la sala: {str(e)}', 'error')
        return redirect(url_for('error_sistema_page'))


@app.route('/sala/<int:sala_id>/configurar-grupos', methods=['POST'])
def configurar_grupos_sala(sala_id):
    """Permite al docente habilitar grupos y definir cuántos grupos crear"""
    try:
        # Solo docentes pueden configurar
        if session.get('usuario_tipo') != 'docente':
            return jsonify({'success': False, 'error': 'No autorizado'}), 403

        data = request.get_json(silent=True) or request.form.to_dict()
        num_grupos = int(data.get('num_grupos', 0))
        nombres = data.get('nombres') or None
        if isinstance(nombres, str):
            # aceptar lista en JSON string
            try:
                import json as _json
                nombres = _json.loads(nombres)
            except Exception:
                nombres = None

        if num_grupos <= 0:
            return jsonify({'success': False, 'error': 'El número de grupos debe ser mayor a 0'}), 400

        ids = controlador_salas.crear_grupos_sala(sala_id, num_grupos, nombres=nombres)
        return jsonify({'success': True, 'grupos_creados': len(ids), 'ids': ids})
    except Exception as e:
        print(f"ERROR configurar_grupos_sala: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'error': 'Error del servidor'}), 500


@app.route('/api/sala/<int:sala_id>/grupos')
def api_obtener_grupos_sala(sala_id):
    try:
        print(f"🔍 [API] Obteniendo grupos para sala {sala_id}")
        grupos = controlador_salas.obtener_grupos_sala(sala_id)
        print(f"📦 [API] Grupos obtenidos: {len(grupos)} grupos")
        print(f"📋 [API] Detalle de grupos: {grupos}")
        return jsonify({'success': True, 'grupos': grupos})
    except Exception as e:
        print(f"❌ ERROR api_obtener_grupos_sala: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/participante/<int:participante_id>/asignar-grupo', methods=['POST'])
def api_asignar_participante_grupo(participante_id):
    """Asignar participante a grupo - Permite asignación después de unirse"""
    try:
        data = request.get_json(silent=True) or request.form.to_dict()
        id_grupo = int(data.get('id_grupo'))

        print(f"🎯 Asignando participante {participante_id} al grupo {id_grupo}")

        ok = controlador_salas.asignar_participante_grupo(participante_id, id_grupo)
        if ok:
            print(f"✅ Grupo asignado correctamente")
            return jsonify({'success': True})
        
        print(f"❌ No se pudo asignar al grupo")
        return jsonify({'success': False, 'error': 'No se pudo asignar al grupo'}), 400
    except Exception as e:
        print(f"❌ ERROR api_asignar_participante_grupo: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'error': 'Error del servidor'}), 500

@app.route('/sala/<int:sala_id>/cerrar', methods=['POST'])
def cerrar_sala(sala_id):
    try:
        resultado = finalizar_sala(sala_id)

        if not resultado:
            raise ValueError("No se pudo cerrar la sala")

        if request.is_json:
            return jsonify({'success': True, 'message': 'Sala cerrada exitosamente'})
        flash('¡Sala cerrada exitosamente!', 'success')
        return redirect(url_for('mis_salas'))
    except Exception as e:
        if request.is_json:
            return jsonify({'success': False, 'error': str(e)}), 400
        flash(f'Error al cerrar la sala: {str(e)}', 'error')
        return redirect(url_for('error_sistema_page'))

@app.route('/sala/<int:sala_id>/juego')
def juego_sala(sala_id):
    """Página del juego en tiempo real"""
    try:
        # Verificar que la sala existe
        conexion = obtener_conexion()
        cursor = conexion.cursor()

        cursor.execute('''
            SELECT id_sala, pin_sala, id_cuestionario, estado, modo_juego
            FROM salas_juego
            WHERE id_sala = %s
        ''', (sala_id,))

        sala_data = cursor.fetchone()
        cursor.close()
        conexion.close()

        if not sala_data:
            flash('Sala no encontrada', 'error')
            return redirect(url_for('error_sistema_page'))

        # Obtener información completa de la sala incluyendo tiene_docente
        sala_info = obtener_sala_por_id_simple(sala_id)

        sala = {
            'id': sala_data[0],
            'pin_sala': sala_data[1],
            'id_cuestionario': sala_data[2],
            'estado': sala_data[3],
            'modo_juego': sala_data[4],
            'tiene_docente': sala_info.get('tiene_docente', False)  # Agregar detección de modo
        }

        # Determinar qué vista mostrar según el tipo de usuario
        if session.get('usuario_tipo') == 'docente':
            return render_template('ControlJuegoDocente.html', sala=sala)
        else:
            return render_template('JuegoEstudiante.html', sala=sala)

    except Exception as e:
        print(f"ERROR juego_sala: {e}")
        import traceback
        traceback.print_exc()
        flash(f'Error: {str(e)}', 'error')
        return redirect(url_for('error_sistema_page'))

# ==================== RUTAS DEL SISTEMA DE JUEGO EN TIEMPO REAL ====================

@app.route('/api/sala/<int:sala_id>/pregunta-actual')
def obtener_pregunta_actual(sala_id):
    """API para obtener la pregunta actual que se está mostrando"""
    try:
        pregunta = controlador_juego.obtener_pregunta_actual_sala(sala_id)

        if not pregunta:
            return jsonify({'success': False, 'error': 'No hay pregunta activa'}), 404

        # Obtener configuración de tiempo de la sala (por defecto 30 segundos)
        conexion = obtener_conexion()
        cursor = conexion.cursor()
        cursor.execute("SELECT tiempo_por_pregunta FROM salas_juego WHERE id_sala = %s", (sala_id,))
        resultado = cursor.fetchone()
        tiempo_pregunta = resultado[0] if resultado and resultado[0] else 30
        conexion.close()

        # Agregar tiempo de pregunta a la respuesta
        pregunta['tiempo_limite'] = tiempo_pregunta

        return jsonify({
            'success': True,
            'pregunta': pregunta
        })
    except Exception as e:
        print(f"ERROR obtener_pregunta_actual: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/sala/<int:sala_id>/responder', methods=['POST'])
def responder_pregunta_juego(sala_id):
    """Registra la respuesta de un participante a la pregunta actual"""
    try:
        data = request.get_json()

        participante_id = session.get('participante_id')
        print(f"\n📝 RESPONDER PREGUNTA:")
        print(f"   Sala ID: {sala_id}")
        print(f"   Participante ID: {participante_id}")
        print(f"   Data recibida: {data}")

        if not participante_id:
            return jsonify({'success': False, 'error': 'No hay sesión de participante'}), 401

        id_pregunta = int(data.get('id_pregunta'))
        id_opcion = int(data.get('id_opcion'))
        tiempo_respuesta = float(data.get('tiempo_respuesta'))

        print(f"   ID Pregunta: {id_pregunta}")
        print(f"   ID Opción: {id_opcion}")
        print(f"   Tiempo respuesta: {tiempo_respuesta}s")

        resultado = controlador_juego.registrar_respuesta_participante(
            participante_id=participante_id,
            sala_id=sala_id,
            id_pregunta=id_pregunta,
            id_opcion_seleccionada=id_opcion,
            tiempo_respuesta=tiempo_respuesta
        )

        print(f"   ✅ Resultado: es_correcta={resultado['es_correcta']}, puntaje={resultado['puntaje_obtenido']}")

        # CAMBIO: Ya no avanzar automáticamente en ningún modo
        # El estudiante siempre debe esperar a que el docente avance
        return jsonify({
            'success': True,
            'resultado': resultado,
            'mensaje': 'Respuesta registrada. Esperando a que el docente avance.'
        })

    except Exception as e:
        print(f"ERROR responder_pregunta_juego: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/sala/<int:sala_id>/siguiente-pregunta', methods=['POST'])
def avanzar_pregunta(sala_id):
    """Avanza a la siguiente pregunta (docente o estudiante en modo individual)"""
    try:
        # Permitir tanto a docentes como a estudiantes avanzar
        # (Estudiantes solo avanzan después de responder)
        usuario_tipo = session.get('usuario_tipo')

        if usuario_tipo not in ['docente', 'estudiante']:
            return jsonify({'success': False, 'error': 'No autorizado'}), 403

        # Si es estudiante, verificar que haya respondido la pregunta actual
        if usuario_tipo == 'estudiante':
            participante_id = session.get('participante_id')
            if not participante_id:
                return jsonify({'success': False, 'error': 'No hay sesión de participante'}), 401

            # Verificar que el estudiante haya respondido la pregunta actual
            conexion = obtener_conexion()
            try:
                with conexion.cursor() as cursor:
                    # Obtener el número de pregunta actual y el id_cuestionario
                    cursor.execute('''
                        SELECT ejs.pregunta_actual, s.id_cuestionario
                        FROM estado_juego_sala ejs
                        JOIN salas_juego s ON ejs.id_sala = s.id_sala
                        WHERE ejs.id_sala = %s
                    ''', (sala_id,))

                    estado_result = cursor.fetchone()
                    if not estado_result:
                        return jsonify({
                            'success': False,
                            'error': 'No se pudo obtener el estado de la sala'
                        }), 500

                    # Usar índices numéricos en lugar de claves de diccionario
                    num_pregunta_actual = estado_result[0]  # pregunta_actual
                    id_cuestionario = estado_result[1]      # id_cuestionario

                    # Obtener el id_pregunta correspondiente al número de pregunta actual
                    cursor.execute('''
                        SELECT p.id_pregunta
                        FROM cuestionario_preguntas cp
                        JOIN preguntas p ON cp.id_pregunta = p.id_pregunta
                        WHERE cp.id_cuestionario = %s
                        ORDER BY cp.orden
                        LIMIT %s, 1
                    ''', (id_cuestionario, num_pregunta_actual - 1))

                    pregunta_result = cursor.fetchone()
                    if not pregunta_result:
                        return jsonify({
                            'success': False,
                            'error': 'No se pudo obtener la pregunta actual'
                        }), 500

                    id_pregunta_actual = pregunta_result[0]  # id_pregunta

                    # Verificar si el estudiante respondió esta pregunta
                    cursor.execute('''
                        SELECT COUNT(*) as count
                        FROM respuestas_participantes
                        WHERE id_participante = %s
                        AND id_sala = %s
                        AND id_pregunta = %s
                    ''', (participante_id, sala_id, id_pregunta_actual))

                    result = cursor.fetchone()
                    if not result or result[0] == 0:  # count es el primer elemento
                        return jsonify({
                            'success': False,
                            'error': 'Debes responder la pregunta actual antes de avanzar'
                        }), 403
            finally:
                conexion.close()

        hay_mas = controlador_juego.avanzar_siguiente_pregunta(sala_id)

        return jsonify({
            'success': True,
            'hay_mas_preguntas': hay_mas,
            'message': 'Siguiente pregunta' if hay_mas else 'Juego finalizado'
        })
    except Exception as e:
        print(f"ERROR avanzar_pregunta: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/sala/<int:sala_id>/ranking')
def obtener_ranking(sala_id):
    """Obtiene el ranking actual de la sala"""
    try:
        ranking = controlador_juego.obtener_ranking_sala(sala_id)

        return jsonify({
            'success': True,
            'ranking': ranking
        })
    except Exception as e:
        print(f"ERROR obtener_ranking: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/sala/<int:sala_id>/estadisticas-pregunta')
def obtener_estadisticas_pregunta(sala_id):
    """Obtiene estadísticas de cuántos han respondido la pregunta actual"""
    try:
        stats = controlador_juego.obtener_estadisticas_pregunta_actual(sala_id)

        return jsonify({
            'success': True,
            'estadisticas': stats
        })
    except Exception as e:
        print(f"ERROR obtener_estadisticas_pregunta: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/sala/<int:sala_id>/detalle-respuestas')
def obtener_detalle_respuestas(sala_id):
    """Obtiene el detalle de qué estudiantes han respondido la pregunta actual"""
    try:
        detalle = controlador_juego.obtener_detalle_respuestas_estudiantes(sala_id)

        if not detalle:
            return jsonify({
                'success': False,
                'error': 'No se pudo obtener el detalle de respuestas'
            }), 404

        return jsonify({
            'success': True,
            'detalle': detalle
        })
    except Exception as e:
        print(f"ERROR obtener_detalle_respuestas: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/sala/<int:sala_id>/finalizar', methods=['POST'])
@login_required
@docente_required
def finalizar_juego(sala_id):
    """Finaliza el juego y redirige a resultados"""
    try:
        # Solo docentes pueden finalizar
        if session.get('usuario_tipo') != 'docente':
            return jsonify({'success': False, 'error': 'No autorizado'}), 403

        resultado = controlador_juego.finalizar_juego_sala(sala_id)

        if not resultado:
            raise ValueError("No se pudo finalizar el juego")

        if request.is_json:
            return jsonify({
                'success': True,
                'message': 'Juego finalizado',
                'redirect': url_for('ver_resultados_juego', sala_id=sala_id)
            })

        return redirect(url_for('ver_resultados_juego', sala_id=sala_id))
    except Exception as e:
        print(f"ERROR finalizar_juego: {e}")
        import traceback
        traceback.print_exc()
        if request.is_json:
            return jsonify({'success': False, 'error': str(e)}), 500
        flash(f'Error al finalizar el juego: {str(e)}', 'error')
        return redirect(url_for('error_sistema_page'))

@app.route('/sala/<int:sala_id>/resultados')
def ver_resultados_juego(sala_id):
    """Muestra los resultados finales del juego"""
    try:
        # Obtener información de la sala
        sala = obtener_sala_por_id_simple(sala_id)
        if not sala:
            flash('Sala no encontrada', 'error')
            return redirect(url_for('error_sistema_page'))

        # Obtener ranking completo
        ranking = controlador_juego.obtener_ranking_sala(sala_id)

        # Obtener cuestionario
        conexion = obtener_conexion()
        cursor = conexion.cursor()
        cursor.execute('''
            SELECT c.id_cuestionario, c.titulo, c.descripcion
            FROM cuestionarios c
            JOIN salas_juego s ON c.id_cuestionario = s.id_cuestionario
            WHERE s.id_sala = %s
        ''', (sala_id,))
        cuestionario_data = cursor.fetchone()
        cursor.close()
        conexion.close()

        # Obtener recompensas otorgadas en este juego
        recompensas_otorgadas = []
        try:
            if cuestionario_data:
                id_cuestionario = cuestionario_data[0]
                # Obtener recompensas asignadas a los participantes de esta sala
                conexion = obtener_conexion()
                cursor = conexion.cursor()
                cursor.execute('''
                    SELECT
                        r.nombre AS recompensa,
                        r.tipo,
                        p.nombre_participante,
                        rk.posicion,
                        ro.fecha_otorgacion
                    FROM recompensas_otorgadas ro
                    JOIN recompensas r ON ro.id_recompensa = r.id_recompensa
                    JOIN usuarios u ON ro.id_estudiante = u.id_usuario
                    JOIN participantes_sala p ON p.id_usuario = u.id_usuario AND p.id_sala = %s
                    JOIN ranking_sala rk ON rk.id_participante = p.id_participante AND rk.id_sala = %s
                    ORDER BY rk.posicion ASC
                ''', (sala_id, sala_id))

                for row in cursor.fetchall():
                    recompensas_otorgadas.append({
                        'recompensa': row[0],
                        'tipo': row[1],
                        'nombre_participante': row[2],
                        'posicion': row[3],
                        'fecha': row[4]
                    })
                cursor.close()
                conexion.close()
        except Exception as e:
            print(f"⚠️ Error al obtener recompensas otorgadas: {e}")
            recompensas_otorgadas = []

        # Si es estudiante, obtener su resultado personal
        resultado_personal = None
        if session.get('usuario_tipo') == 'estudiante':
            participante_id = session.get('participante_id')
            if participante_id:
                resultado_personal = controlador_juego.obtener_resultado_participante(sala_id, participante_id)

        # Normalizar nombres de campos para consistencia con el frontend
        ranking_normalizado = []
        for jugador in ranking:
            ranking_normalizado.append({
                'posicion': jugador.get('posicion'),
                'nombre_participante': jugador.get('nombre_completo') or jugador.get('nombre_participante', 'Jugador'),
                'puntaje_total': jugador.get('puntos_totales') or jugador.get('puntaje_total', 0),
                'respuestas_correctas': jugador.get('respuestas_correctas', 0),
                'tiempo_total_respuestas': jugador.get('tiempo_total_respuestas', 0),
                'numero_grupo': jugador.get('numero_grupo'),
                'id_participante': jugador.get('id_participante')
            })

        return render_template('ResultadosJuego.html',
            sala=sala,
            ranking=ranking_normalizado,
            cuestionario_titulo=cuestionario_data[1] if cuestionario_data else 'Cuestionario',
            cuestionario_descripcion=cuestionario_data[2] if cuestionario_data else '',
            resultado_personal=resultado_personal,
            recompensas_otorgadas=recompensas_otorgadas,
            es_docente=(session.get('usuario_tipo') == 'docente')
        )
    except Exception as e:
        print(f"ERROR ver_resultados_juego: {e}")
        import traceback
        traceback.print_exc()
        flash(f'Error al cargar resultados: {str(e)}', 'error')
        return redirect(url_for('error_sistema_page'))

# Rutas para respuestas y participaciones
@app.route('/responder-pregunta', methods=['POST'])
def responder_pregunta():
    data = request.get_json(silent=True) or request.form.to_dict()
    try:
        pregunta_id = int(data.get('pregunta_id'))
        participante_id = int(data.get('participante_id'))
        sala_id = int(data.get('sala_id'))
        id_opcion_seleccionada = int(data.get('id_opcion'))
        tiempo_respuesta = float(data.get('tiempo_respuesta', 0))

        # Registrar respuesta usando la función correcta del controlador
        resultado = controlador_juego.registrar_respuesta_participante(
            participante_id=participante_id,
            sala_id=sala_id,
            id_pregunta=pregunta_id,
            id_opcion_seleccionada=id_opcion_seleccionada,
            tiempo_respuesta=tiempo_respuesta
        )

        return jsonify({
            'success': True,
            'es_correcta': resultado['es_correcta'],
            'puntaje_obtenido': resultado['puntaje_obtenido'],
            'tiempo_respuesta': resultado['tiempo_respuesta'],
            'message': 'Respuesta registrada exitosamente'
        })
    except Exception as e:
        print(f"❌ Error en responder_pregunta: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 400

@app.route('/resultados-participacion/<int:participacion_id>')
def ver_resultados_participacion(participacion_id):
    try:
        resultados = obtener_resultados_participacion(participacion_id)
        return render_template('ResultadosJuego.html', resultados=resultados)
    except Exception as e:
        flash(f'Error al obtener los resultados: {str(e)}', 'error')
        return redirect(url_for('error_sistema_page'))

@app.route('/ranking-sala/<int:sala_id>')
def ver_ranking_sala(sala_id):
    try:
        ranking = obtener_ranking_sala(sala_id)
        return jsonify({'success': True, 'ranking': ranking})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 400

# Rutas de administración y sistema
@app.route('/admin/usuarios')
def admin_usuarios():
    try:
        usuarios = obtener_todos_usuarios()
        return render_template('AdminUsuarios.html', usuarios=usuarios)
    except Exception as e:
        flash(f'Error al cargar usuarios: {str(e)}', 'error')
        return redirect(url_for('error_sistema_page'))

@app.route('/admin/estadisticas')
def admin_estadisticas():
    try:
        estadisticas = obtener_estadisticas_sistema()
        return render_template('Estadisticas.html', estadisticas=estadisticas)
    except Exception as e:
        flash(f'Error al cargar estadísticas: {str(e)}', 'error')
        return redirect(url_for('error_sistema_page'))

@app.route('/perfil/<int:user_id>')
def perfil_usuario(user_id):
    try:
        usuario = obtener_usuario_por_id(user_id)
        if not usuario:
            flash('Usuario no encontrado', 'error')
            return redirect(url_for('maestra'))

        # Obtener estadísticas del usuario
        estadisticas = obtener_estadisticas_usuario(user_id)

        return render_template('PerfilUsuario.html', usuario=usuario, estadisticas=estadisticas)
    except Exception as e:
        flash(f'Error al cargar el perfil: {str(e)}', 'error')
        return redirect(url_for('error_sistema_page'))

@app.route('/configuracion', methods=['GET', 'POST'])
def configuracion_sistema_old():
    if request.method == 'GET':
        try:
            configuracion = obtener_configuracion_sistema()
            return render_template('ConfiguracionSistema.html', configuracion=configuracion)
        except Exception as e:
            flash(f'Error al cargar configuración: {str(e)}', 'error')
            return redirect(url_for('error_sistema_page'))

    # POST - actualizar configuración
    try:
        data = request.get_json(silent=True) or request.form.to_dict()
        actualizar_configuracion_sistema(data)

        if request.is_json:
            return jsonify({'success': True, 'message': 'Configuración actualizada exitosamente'})
        flash('¡Configuración actualizada exitosamente!', 'success')
        return redirect(url_for('configuracion_sistema'))
    except Exception as e:
        if request.is_json:
            return jsonify({'success': False, 'error': str(e)}), 400
        flash(f'Error al actualizar configuración: {str(e)}', 'error')
        return redirect(url_for('error_sistema_page'))

# Rutas de juego tiempo real
@app.route('/juego/<int:sala_id>/estudiante')
def juego_estudiante(sala_id):
    try:
        # Verificar que el estudiante está en la sala
        participante_id = session.get('participante_id')
        if not participante_id:
            flash('Debes unirte a una sala primero', 'error')
            return redirect(url_for('unirse_sala_route', codigo=''))

        sala = obtener_sala_por_id_simple(sala_id)
        if not sala:
            flash('Sala no encontrada', 'error')
            return redirect(url_for('maestra'))

        # Verificar si es modo automático basándose en el PIN
        pin_sala = sala.get('pin_sala', '')
        modo_automatico = es_sala_automatica(pin_sala)

        return render_template('JuegoEstudiante.html',
                             sala=sala,
                             participante_id=participante_id,
                             modo_automatico=modo_automatico)
    except Exception as e:
        flash(f'Error al cargar el juego: {str(e)}', 'error')
        return redirect(url_for('error_sistema_page'))

@app.route('/juego/<int:sala_id>/siguiente-pregunta')
def siguiente_pregunta(sala_id):
    try:
        pregunta = obtener_siguiente_pregunta_sala(sala_id)
        if not pregunta:
            return jsonify({'success': False, 'finished': True, 'message': 'No hay más preguntas'})

        return jsonify({'success': True, 'pregunta': pregunta})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 400

@app.route('/esperar-juego/<int:sala_id>')
def esperar_juego(sala_id):
    try:
        sala = obtener_sala_por_id_simple(sala_id)
        if not sala:
            flash('Sala no encontrada', 'error')
            return redirect(url_for('maestra'))

        return render_template('EsperarJuego.html', sala=sala)
    except Exception as e:
        flash(f'Error: {str(e)}', 'error')
        return redirect(url_for('error_sistema_page'))

# ========== APIs para actualizaciones en tiempo real ==========

@app.route('/api/sala/<int:sala_id>/participantes')
def api_obtener_participantes(sala_id):
    """API para obtener lista de participantes en tiempo real"""
    try:
        participantes = controlador_salas.obtener_participantes_sala(sala_id)

        # Obtener también el estado de la sala para verificar si el juego comenzó
        conexion = obtener_conexion()
        cursor = conexion.cursor()
        cursor.execute('SELECT estado FROM salas_juego WHERE id_sala = %s', (sala_id,))
        sala_data = cursor.fetchone()
        cursor.close()
        conexion.close()

        estado_sala = sala_data[0] if sala_data else 'esperando'

        # Log para depuración
        print(f"📊 API participantes - Sala {sala_id}: Estado = '{estado_sala}', Participantes = {len(participantes)}")

        return jsonify({
            'success': True,
            'participantes': participantes,
            'total': len(participantes),
            'sala': {
                'estado': estado_sala
            }
        })
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/sala/<int:sala_id>/estado')
def api_obtener_estado_sala(sala_id):
    """API para obtener el estado actual de la sala"""
    try:
        sala = obtener_sala_por_id_simple(sala_id)
        if not sala:
            return jsonify({'success': False, 'error': 'Sala no encontrada'}), 404

        participantes = controlador_salas.obtener_participantes_sala(sala_id)
        
        # Asegurar que el estado siempre tenga un valor
        estado_actual = sala.get('estado') or 'esperando'
        
        print(f"🔍 [API ESTADO] Sala {sala_id}: estado='{estado_actual}', participantes={len(participantes)}")

        return jsonify({
            'success': True,
            'estado': estado_actual,
            'total_participantes': len(participantes),
            'sala': {
                'id': sala.get('id'),
                'pin_sala': sala.get('pin_sala'),
                'estado': estado_actual,
                'max_participantes': sala.get('max_participantes', 30)
            }
        })
    except Exception as e:
        print(f"❌ [API ESTADO] Error: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/sala/verificar/<string:pin>')
def api_verificar_sala_por_pin(pin):
    """API para verificar si una sala existe y está disponible"""
    try:
        # Validar formato del PIN
        if not pin or len(pin) != 6 or not pin.isdigit():
            return jsonify({'success': False, 'error': 'PIN inválido'}), 400

        # Buscar sala por PIN
        sala = controlador_salas.obtener_sala_por_codigo(pin)

        if not sala:
            return jsonify({'success': False, 'error': 'No existe una sala con ese PIN'}), 404

        # Obtener participantes actuales
        participantes = controlador_salas.obtener_participantes_sala(sala['id'])

        return jsonify({
            'success': True,
            'sala': {
                'id': sala.get('id'),
                'id_sala': sala.get('id'),  # Agregar también como id_sala
                'pin_sala': sala.get('pin_sala'),
                'estado': sala.get('estado'),
                'max_participantes': sala.get('max_participantes', 30),
                'participantes_actuales': len(participantes)
            }
        })
    except Exception as e:
        print(f"ERROR en api_verificar_sala_por_pin: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'error': 'Error interno del servidor'}), 500

@app.route('/api/participantes/<string:pin>')
def api_obtener_participantes_por_pin(pin):
    """API para obtener lista de participantes usando el PIN de la sala (para MonitoreoJuego.html)"""
    try:
        print(f"🔍 API: Obteniendo participantes para PIN: {pin}")

        # Validar formato del PIN
        if not pin or len(pin) != 6 or not pin.isdigit():
            print(f"❌ PIN inválido: {pin}")
            return jsonify({'success': False, 'error': 'PIN inválido'}), 400

        # Buscar sala por PIN
        sala = controlador_salas.obtener_sala_por_codigo(pin)

        if not sala:
            print(f"❌ Sala no encontrada con PIN: {pin}")
            return jsonify({'success': False, 'error': 'Sala no encontrada'}), 404

        # Obtener participantes
        participantes = controlador_salas.obtener_participantes_sala(sala['id'])

        print(f"✅ Participantes encontrados: {len(participantes)}")

        # Formatear participantes para el frontend
        participantes_formateados = []
        for p in participantes:
            participantes_formateados.append({
                'id': p['id_participante'],
                'nombre': p['nombre_participante'],
                'estado': p['estado'],
                'fecha_union': p['fecha_union'].isoformat() if p['fecha_union'] else None,
                'id_grupo': p.get('id_grupo'),  # ID del grupo si está asignado
                'nombre_grupo': p.get('nombre_grupo')  # Nombre del grupo si está asignado
            })

        return jsonify({
            'success': True,
            'participantes': participantes_formateados,
            'total': len(participantes_formateados)
        })

    except Exception as e:
        print(f"❌ ERROR en api_obtener_participantes_por_pin: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'error': 'Error del servidor'}), 500

@app.route('/unirse-juego', methods=['GET', 'POST'])
def unirse_juego():
    """Ruta mejorada para unirse a un juego con PIN"""
    if request.method == 'GET':
        return render_template('UnirseJuego.html')

    try:
        data = request.get_json(silent=True) or request.form.to_dict()
        pin_sala = data.get('pin_sala', '').strip()
        nombre_estudiante = data.get('nombre_estudiante', '').strip()

        print(f"\n{'='*80}")
        print(f"🎮 ESTUDIANTE INTENTANDO UNIRSE")
        print(f"PIN: {pin_sala}")
        print(f"Nombre: {nombre_estudiante}")
        print(f"Datos recibidos: {data}")
        print(f"{'='*80}\n")

        # Validaciones
        if not pin_sala:
            print("❌ Error: PIN vacío")
            return jsonify({'success': False, 'error': 'El PIN de la sala es requerido'}), 400

        if not nombre_estudiante:
            print("❌ Error: Nombre vacío")
            return jsonify({'success': False, 'error': 'Tu nombre es requerido'}), 400

        if len(pin_sala) != 6 or not pin_sala.isdigit():
            print(f"❌ Error: PIN inválido (longitud: {len(pin_sala)})")
            return jsonify({'success': False, 'error': 'El PIN debe ser de 6 dígitos'}), 400

        # Buscar la sala por PIN
        print(f"🔍 Buscando sala con PIN: {pin_sala}")
        sala = controlador_salas.obtener_sala_por_codigo(pin_sala)
        print(f"📋 Sala encontrada: {sala}")

        if not sala:
            print(f"❌ Error: No se encontró sala con PIN {pin_sala}")
            return jsonify({'success': False, 'error': 'Sala no encontrada. Verifica el PIN'}), 404

        print(f"✅ Sala encontrada - ID: {sala['id']}, Estado: {sala['estado']}")

        # Verificar que la sala está en estado 'esperando'
        if sala['estado'] != 'esperando':
            if sala['estado'] == 'en_curso':
                return jsonify({'success': False, 'error': 'La sala ya está en curso. No puedes unirte'}), 400
            elif sala['estado'] == 'finalizada':
                return jsonify({'success': False, 'error': 'La sala ha finalizado'}), 400
            else:
                return jsonify({'success': False, 'error': f'La sala no está disponible (estado: {sala["estado"]}'}), 400

        # Verificar límite de participantes
        print(f"📊 Obteniendo participantes actuales de sala {sala['id']}...")
        participantes_actuales = controlador_salas.obtener_participantes_sala(sala['id'])
        max_participantes = sala.get('max_participantes') or 30  # Default 30 si es None

        print(f"👥 Participantes actuales: {len(participantes_actuales)}/{max_participantes}")
        print(f"   Lista: {participantes_actuales}")

        if max_participantes and len(participantes_actuales) >= max_participantes:
            print(f"❌ Error: Sala llena")
            return jsonify({'success': False, 'error': f'La sala está llena ({max_participantes} participantes máximo)'}), 400

        # Agregar el participante a la sala
        id_usuario = session.get('usuario_id')  # Si está logueado
        print(f"➕ Agregando participante: {nombre_estudiante} (usuario_id: {id_usuario})")

        try:
            participante_id = controlador_salas.agregar_participante_sala(sala['id'], nombre_estudiante, id_usuario)
        except ValueError as ve:
            print(f"❌ Error al agregar: {ve}")
            return jsonify({'success': False, 'error': str(ve)}), 400

        print(f"✅ Participante agregado exitosamente!")
        print(f"   - ID Participante: {participante_id}")
        print(f"   - Sala ID: {sala['id']}")
        print(f"   - PIN Sala: {pin_sala}")

        # Guardar en sesión
        session['participante_id'] = participante_id
        session['sala_id'] = sala['id']
        session['nombre_participante'] = nombre_estudiante
        session['pin_sala'] = pin_sala

        redirect_url = url_for('sala_espera', sala_id=sala['id'])
        print(f"   - URL Redirección: {redirect_url}")
        print(f"{'='*80}\n")

        return jsonify({
            'success': True,
            'participante_id': participante_id,
            'sala_id': sala['id'],
            'redirect_url': redirect_url,
            'message': f'¡Te has unido exitosamente! Espera a que el docente inicie el juego.'
        })

    except ValueError as ve:
        print(f"DEBUG: ValueError: {ve}")
        return jsonify({'success': False, 'error': str(ve)}), 400
    except Exception as e:
        print(f"DEBUG: Error al unirse a sala: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'error': 'Error interno del servidor'}), 500

@app.route('/sala-espera/<int:sala_id>')
def sala_espera(sala_id):
    """Página de espera para estudiantes"""
    try:
        # Verificar que el estudiante esté en la sala
        if session.get('sala_id') != sala_id:
            flash('No estás registrado en esta sala', 'error')
            return redirect(url_for('unirse_juego'))

        sala = obtener_sala_por_id_simple(sala_id)
        if not sala:
            flash('Sala no encontrada', 'error')
            return redirect(url_for('unirse_juego'))

        # Obtener título del cuestionario
        cuestionario = obtener_cuestionario_por_id_simple(sala['id_cuestionario'])
        if cuestionario:
            sala['cuestionario_titulo'] = cuestionario[1]  # El título está en índice 1

        # Obtener información del participante
        nombre_participante = session.get('nombre_participante', 'Participante')
        participante_id = session.get('participante_id')
        pin_sala = session.get('pin_sala', sala.get('pin_sala', ''))

        return render_template('SalaEspera.html',
                             sala=sala,
                             nombre_participante=nombre_participante,
                             participante_id=participante_id,
                             pin_sala=pin_sala)
    except Exception as e:
        print(f"ERROR en sala_espera: {e}")
        import traceback
        traceback.print_exc()
        flash(f'Error: {str(e)}', 'error')
        return redirect(url_for('unirse_juego'))

# Rutas para juegos
@app.route('/crear-cuestionario', methods=['GET', 'POST'])
@login_required
@docente_required
def crear_cuestionario_route():
    if request.method == 'POST':
        data = request.get_json(silent=True) or request.form.to_dict()
        try:
            titulo = data.get('titulo', '').strip()
            descripcion = data.get('descripcion', '').strip()
            id_docente = session.get('usuario_id', 1)  # Usar usuario de sesión si está disponible

            if not titulo:
                raise ValueError("El título es requerido")

            # Debug: Verificar los datos antes de crear
            print(f"DEBUG: Creando cuestionario - titulo: {titulo}, id_docente: {id_docente}")

            cuestionario_id = controlador_cuestionarios.crear_cuestionario(titulo, descripcion, id_docente)

            print(f"DEBUG: Cuestionario creado con ID: {cuestionario_id}")

            if request.is_json:
                return jsonify({'success': True, 'cuestionario_id': cuestionario_id, 'message': 'Cuestionario creado exitosamente'})

            # FLUJO CORREGIDO: Después de crear el cuestionario, redirigir a agregar preguntas
            flash('¡Cuestionario creado exitosamente! Ahora puedes agregar preguntas.', 'success')
            return redirect(url_for('agregar_preguntas', cuestionario_id=cuestionario_id))

        except Exception as e:
            print(f"DEBUG: Error al crear cuestionario: {str(e)}")
            if request.is_json:
                return jsonify({'success': False, 'error': str(e)}), 400
            flash(f'Error al crear el cuestionario: {str(e)}', 'error')
            return redirect(url_for('error_sistema_page'))

    return render_template('CrearCuestionario.html')

# Nueva ruta: Agregar preguntas a un cuestionario recién creado
@app.route('/agregar-preguntas/<int:cuestionario_id>', methods=['GET', 'POST'])
def agregar_preguntas(cuestionario_id):
    print(f"DEBUG: *** agregar_preguntas - Método: {request.method}, cuestionario_id: {cuestionario_id}")

    # Si es POST, procesar el formulario de agregar pregunta
    if request.method == 'POST':
        print(f"DEBUG: *** Procesando POST - Form data: {dict(request.form)}")
        try:
            # Verificar que el usuario es docente
            if session.get('usuario_tipo') != 'docente':
                print(f"DEBUG: Usuario no es docente: {session.get('usuario_tipo')}")
                flash('Solo los docentes pueden agregar preguntas', 'error')
                return redirect(url_for('agregar_preguntas', cuestionario_id=cuestionario_id))

            # Obtener datos del formulario
            enunciado = request.form.get('question_text', '').strip()
            tipo = request.form.get('question_type', 'opcion_multiple')

            print(f"DEBUG: Enunciado: '{enunciado}', Tipo: '{tipo}'")

            if not enunciado:
                flash('El enunciado de la pregunta es requerido', 'error')
                return redirect(url_for('agregar_preguntas', cuestionario_id=cuestionario_id))

            # Verificar que el cuestionario pertenece al docente
            cuestionario = controlador_cuestionarios.obtener_cuestionario_por_id_simple(cuestionario_id)
            if not cuestionario:
                flash('Cuestionario no encontrado', 'error')
                return redirect(url_for('mis_cuestionarios'))

            id_docente_actual = session.get('usuario_id', 1)
            if cuestionario[3] != id_docente_actual:
                flash('No tienes permisos para agregar preguntas a este cuestionario', 'error')
                return redirect(url_for('mis_cuestionarios'))

            # Crear la pregunta
            pregunta_id = controlador_preguntas.crear_pregunta(enunciado, tipo, cuestionario_id)
            print(f"DEBUG: Resultado crear_pregunta: {pregunta_id}")

            if pregunta_id:
                print(f"DEBUG: Pregunta {pregunta_id} creada exitosamente")

                # Si es opción múltiple, agregar las opciones
                if tipo == 'opcion_multiple':
                    opciones = []
                    opcion_correcta = request.form.get('correct_answer', '').strip()
                    print(f"DEBUG: Opción correcta seleccionada: '{opcion_correcta}'")

                    for i in range(0, 4):  # option_0, option_1, option_2, option_3
                        opcion_texto = request.form.get(f'option_{i}', '').strip()
                        if opcion_texto:
                            # Verificar si esta opción es la correcta (índice como string)
                            es_correcta = (str(i) == opcion_correcta)
                            opciones.append({
                                'texto': opcion_texto,
                                'es_correcta': es_correcta
                            })
                            print(f"DEBUG: Opción {i}: '{opcion_texto}' (correcta: {es_correcta})")

                    # Crear las opciones de respuesta
                    for opcion in opciones:
                        resultado = controlador_preguntas.crear_opcion_respuesta(pregunta_id, opcion['texto'], opcion['es_correcta'])
                        print(f"DEBUG: Opción creada: {resultado}")

                flash('Pregunta agregada exitosamente', 'success')
            else:
                flash('Error al crear la pregunta', 'error')

            return redirect(url_for('agregar_preguntas', cuestionario_id=cuestionario_id))

        except Exception as e:
            print(f"DEBUG: Error procesando pregunta: {str(e)}")
            import traceback
            print(f"DEBUG: Traceback: {traceback.format_exc()}")
            flash(f'Error al procesar la pregunta: {str(e)}', 'error')
            return redirect(url_for('agregar_preguntas', cuestionario_id=cuestionario_id))

    # Si es GET, mostrar el formulario
    try:
        # Verificar que el cuestionario existe y pertenece al usuario actual
        cuestionario = controlador_cuestionarios.obtener_cuestionario_por_id_simple(cuestionario_id)
        print(f"DEBUG: agregar_preguntas GET - cuestionario obtenido: {cuestionario}")

        if not cuestionario:
            print(f"DEBUG: Cuestionario {cuestionario_id} no encontrado")
            flash('Cuestionario no encontrado', 'error')
            return redirect(url_for('mis_cuestionarios'))

        # Verificar que el cuestionario pertenece al usuario actual
        id_docente_actual = session.get('usuario_id', 1)
        if cuestionario[3] != id_docente_actual:  # cuestionario[3] es id_docente
            print(f"DEBUG: Usuario {id_docente_actual} no tiene permisos para cuestionario del docente {cuestionario[3]}")
            flash('No tienes permisos para editar este cuestionario', 'error')
            return redirect(url_for('mis_cuestionarios'))

        # Obtener las preguntas existentes del cuestionario
        preguntas = []
        try:
            preguntas = controlador_preguntas.obtener_preguntas_por_cuestionario(cuestionario_id)
        except Exception as e:
            print(f"DEBUG: Error obteniendo preguntas: {e}")
            preguntas = []  # En caso de error, asumimos que no hay preguntas aún

        # Convertir cuestionario (tupla) a diccionario para el template
        cuestionario_data = {
            'id_cuestionario': cuestionario[0],
            'titulo': cuestionario[1],
            'descripcion': cuestionario[2],
            'id_docente': cuestionario[3],
            'fecha_creacion': cuestionario[4],
            'fecha_programada': cuestionario[5] if len(cuestionario) > 5 else None,
            'fecha_publicacion': cuestionario[6] if len(cuestionario) > 6 else None,
            'estado': cuestionario[7] if len(cuestionario) > 7 else 'borrador'
        }

        print(f"DEBUG: Mostrando GenerarPreguntas.html con cuestionario: {cuestionario_data}")
        print(f"DEBUG: Número de preguntas existentes: {len(preguntas)}")

        # Usar GenerarPreguntas.html que ya existe
        return render_template('GenerarPreguntas.html',
                             cuestionario=cuestionario_data,
                             preguntas=preguntas)

    except Exception as e:
        print(f"DEBUG: Error en agregar_preguntas GET: {str(e)}")
        flash(f'Error al cargar el cuestionario: {str(e)}', 'error')
        return redirect(url_for('mis_cuestionarios'))


# Nueva ruta: Crear sala específicamente para un cuestionario
@app.route('/crear-sala-cuestionario/<int:cuestionario_id>', methods=['GET', 'POST'])
def crear_sala_para_cuestionario(cuestionario_id):
    try:
        print(f"DEBUG: crear_sala_para_cuestionario - cuestionario_id: {cuestionario_id}, method: {request.method}")

        # Verificar que el cuestionario existe y tiene preguntas
        cuestionario = obtener_cuestionario_por_id_simple(cuestionario_id)
        print(f"DEBUG: Cuestionario obtenido: {cuestionario}")

        if not cuestionario:
            flash('Cuestionario no encontrado', 'error')
            return redirect(url_for('mis_cuestionarios'))

        # Verificar que el cuestionario tiene preguntas
        preguntas = obtener_preguntas_por_cuestionario_simple(cuestionario_id)
        print(f"DEBUG: Preguntas encontradas: {len(preguntas) if preguntas else 0}")

        if not preguntas or len(preguntas) == 0:
            flash('No se puede crear una sala para un cuestionario sin preguntas. Agrega preguntas primero.', 'error')
            return redirect(url_for('agregar_preguntas', cuestionario_id=cuestionario_id))

        if request.method == 'POST':
            print("DEBUG: Procesando POST para crear sala")
            data = request.get_json(silent=True) or request.form.to_dict()
            print(f"DEBUG: Datos recibidos: {data}")

            nombre_sala = data.get('nombre_sala', f"Sala - {cuestionario[1]}")  # cuestionario[1] es el titulo
            capacidad_maxima = int(data.get('capacidad_maxima', 50))
            modo_juego = data.get('modo_juego', 'individual')
            num_grupos = int(data.get('num_grupos', 3)) if modo_juego == 'grupo' else 0
            id_docente = session.get('usuario_id', 1)

            print(f"DEBUG: Creando sala - nombre: {nombre_sala}, capacidad: {capacidad_maxima}, modo: {modo_juego}, grupos: {num_grupos}, docente: {id_docente}")

            # Crear la sala
            try:
                sala_id, codigo_sala = crear_sala_simple(cuestionario_id)
                print(f"DEBUG: Sala creada con ID: {sala_id}, código: {codigo_sala}")

                if sala_id and codigo_sala:
                    # Si es modo grupo, crear los grupos
                    if modo_juego == 'grupo' and num_grupos > 0:
                        try:
                            crear_grupos_para_sala(sala_id, num_grupos)
                            print(f"DEBUG: {num_grupos} grupos creados para sala {sala_id}")
                        except Exception as grupos_error:
                            print(f"DEBUG: Error al crear grupos: {grupos_error}")
                            # No fallar si los grupos no se crean, la sala ya existe

                    # Para peticiones POST, siempre devolver JSON (ya que el frontend espera JSON)
                    print("DEBUG: Devolviendo respuesta JSON exitosa")
                    return jsonify({
                        'success': True,
                        'sala_id': sala_id,
                        'codigo_sala': codigo_sala,
                        'capacidad': capacidad_maxima,
                        'modo_juego': modo_juego,
                        'num_grupos': num_grupos if modo_juego == 'grupo' else 0,
                        'message': 'Sala creada exitosamente'
                    })
                else:
                    error_msg = 'No se pudo crear la sala - función crear_sala_simple retornó valores nulos'
                    print(f"DEBUG: Error - {error_msg}")
                    print("DEBUG: Devolviendo respuesta JSON de error")
                    return jsonify({
                        'success': False,
                        'error': error_msg
                    }), 500

            except Exception as sala_error:
                error_msg = f'Error al crear la sala: {str(sala_error)}'
                print(f"DEBUG: Exception en creación de sala: {error_msg}")
                print("DEBUG: Devolviendo respuesta JSON de excepción")
                return jsonify({
                    'success': False,
                    'error': error_msg
                }), 500

        # GET: Mostrar formulario para crear sala
        print("DEBUG: Mostrando formulario GET")
        cuestionario_data = {
            'id_cuestionario': cuestionario[0],
            'titulo': cuestionario[1],
            'descripcion': cuestionario[2],
            'num_preguntas': len(preguntas)
        }
        print(f"DEBUG: Datos del cuestionario para template: {cuestionario_data}")

        return render_template('CrearSalaEspecifica.html', cuestionario=cuestionario_data)

    except Exception as e:
        print(f"DEBUG: Error en crear_sala_para_cuestionario: {str(e)}")
        import traceback
        traceback.print_exc()
        flash(f'Error al crear la sala: {str(e)}', 'error')
        return redirect(url_for('mis_cuestionarios'))

# Ruta para publicar un cuestionario
@app.route('/publicar-cuestionario/<int:cuestionario_id>', methods=['POST'])
def publicar_cuestionario_route(cuestionario_id):
    """Publica un cuestionario cambiando su estado a 'publicado'"""
    print(f"\n{'='*80}")
    print(f"📢 PUBLICAR CUESTIONARIO - ID: {cuestionario_id}")
    print(f"Sesión: logged_in={session.get('logged_in')}, tipo={session.get('usuario_tipo')}, user_id={session.get('usuario_id')}")
    print(f"{'='*80}\n")

    # Verificar sesión manualmente (sin decorador @login_required)
    if 'logged_in' not in session or not session.get('logged_in'):
        print("❌ Usuario no autenticado")
        return jsonify({'success': False, 'error': 'No has iniciado sesión'}), 401

    if session.get('usuario_tipo') != 'docente':
        print(f"❌ Usuario no es docente: {session.get('usuario_tipo')}")
        return jsonify({'success': False, 'error': 'Solo los docentes pueden publicar cuestionarios'}), 403

    try:
        id_docente = session.get('usuario_id')
        print(f"✅ Llamando a publicar_cuestionario({cuestionario_id}, {id_docente})")

        # Llamar a la función del controlador
        success, mensaje = controlador_cuestionarios.publicar_cuestionario(cuestionario_id, id_docente)

        print(f"Resultado: success={success}, mensaje={mensaje}")

        if success:
            print("✅ Publicación exitosa")
            return jsonify({'success': True, 'message': mensaje}), 200
        else:
            print(f"❌ Publicación fallida: {mensaje}")
            return jsonify({'success': False, 'error': mensaje}), 400

    except Exception as e:
        print(f"❌ ERROR CRÍTICO en publicar_cuestionario_route: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'error': f'Error del servidor: {str(e)}'}), 500

# Ruta para despublicar un cuestionario
@app.route('/despublicar-cuestionario/<int:cuestionario_id>', methods=['POST'])
@login_required
def despublicar_cuestionario_route(cuestionario_id):
    try:
        if session.get('usuario_tipo') != 'docente':
            return jsonify({'success': False, 'error': 'Solo los docentes pueden despublicar cuestionarios'}), 403

        id_docente = session.get('usuario_id')

        # Usar la nueva función despublicar_cuestionario
        success, mensaje = controlador_cuestionarios.despublicar_cuestionario(cuestionario_id, id_docente)

        if success:
            return jsonify({
                'success': True,
                'message': mensaje
            })
        else:
            return jsonify({
                'success': False,
                'error': mensaje
            }), 400

    except Exception as e:
        print(f"DEBUG: Error en despublicar_cuestionario: {str(e)}")
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/mis-cuestionarios')
@login_required
@docente_required
def mis_cuestionarios():
    try:
        id_docente = session.get('usuario_id', 1)  # Usar usuario de sesión si está disponible
        print(f"DEBUG: Obteniendo cuestionarios para docente ID: {id_docente}")

        cuestionarios = obtener_cuestionarios_por_docente_simple(id_docente)
        print(f"DEBUG: Cuestionarios encontrados: {len(cuestionarios) if cuestionarios else 0}")

        print(f"DEBUG: Cuestionarios: {cuestionarios}")
        return render_template('MisCuestionarios.html', cuestionarios=cuestionarios)
    except Exception as e:
        print(f"DEBUG: Error al obtener cuestionarios: {str(e)}")
        flash(f'Error al obtener cuestionarios: {str(e)}', 'error')
        return redirect(url_for('error_sistema_page'))

@app.route('/cuestionario/<int:cuestionario_id>')
def ver_cuestionario(cuestionario_id):
    try:
        cuestionario = controlador_cuestionarios.obtener_cuestionario_por_id(cuestionario_id)
        if not cuestionario:
            flash('Cuestionario no encontrado', 'error')
            return redirect(url_for('mis_cuestionarios'))

        preguntas = controlador_preguntas.obtener_preguntas_por_cuestionario(cuestionario_id)
        return render_template('VerCuestionario.html', cuestionario=cuestionario, preguntas=preguntas)
    except Exception as e:
        flash(f'Error al obtener cuestionario: {str(e)}', 'error')
        return redirect(url_for('error_sistema_page'))

@app.route('/editar-cuestionario/<int:cuestionario_id>')
@login_required
@docente_required
def editar_cuestionario(cuestionario_id):
    try:
        print(f"DEBUG: Editando cuestionario ID: {cuestionario_id}")

        # Verificar que el usuario es docente
        if session.get('usuario_tipo') != 'docente':
            flash('Solo los docentes pueden editar cuestionarios', 'error')
            return redirect(url_for('dashboard_docente'))

        # Usar la función que devuelve diccionario
        cuestionario = controlador_cuestionarios.obtener_cuestionario_por_id(cuestionario_id)
        print(f"DEBUG: Cuestionario obtenido (diccionario): {cuestionario}")

        if not cuestionario:
            print(f"DEBUG: Cuestionario {cuestionario_id} no encontrado")
            flash('Cuestionario no encontrado', 'error')
            return redirect(url_for('mis_cuestionarios'))

        # Verificar que el cuestionario pertenece al docente actual
        id_docente_actual = session.get('usuario_id', 1)
        if cuestionario.get('id_docente') != id_docente_actual:
            print(f"DEBUG: Usuario {id_docente_actual} no tiene permisos para cuestionario del docente {cuestionario.get('id_docente')}")
            flash('No tienes permisos para editar este cuestionario', 'error')
            return redirect(url_for('mis_cuestionarios'))

        print(f"DEBUG: Obteniendo preguntas para cuestionario {cuestionario_id}")
        preguntas = []
        try:
            preguntas = controlador_preguntas.obtener_preguntas_por_cuestionario(cuestionario_id)
            print(f"DEBUG: Preguntas obtenidas: {len(preguntas) if preguntas else 0}")
        except Exception as e:
            print(f"DEBUG: Error obteniendo preguntas: {e}")
            preguntas = []

        return render_template('EditarCuestionario.html',
                             cuestionario=cuestionario,
                             preguntas=preguntas,
                             cuestionario_id=cuestionario_id)
    except Exception as e:
        print(f"DEBUG: Error en editar_cuestionario: {str(e)}")
        print(f"DEBUG: Tipo de error: {type(e)}")
        import traceback
        print(f"DEBUG: Traceback: {traceback.format_exc()}")
        flash(f'Error al obtener cuestionario: {str(e)}', 'error')
        return redirect(url_for('error_sistema_page'))

@app.route('/editar-cuestionario/<int:cuestionario_id>', methods=['POST'])
def actualizar_cuestionario_route(cuestionario_id):
    data = request.get_json(silent=True) or request.form.to_dict()
    try:
        titulo = data.get('titulo', '').strip()
        descripcion = data.get('descripcion', '').strip()

        if not titulo:
            raise ValueError("El título es requerido")

        resultado = controlador_cuestionarios.actualizar_cuestionario(cuestionario_id, titulo, descripcion)

        if not resultado:
            raise ValueError("No se pudo actualizar el cuestionario")

        # Verificar si es AJAX por el header, no solo por request.is_json
        is_ajax = request.headers.get('X-Requested-With') == 'XMLHttpRequest' or request.is_json

        if is_ajax:
            return jsonify({'success': True, 'message': 'Cuestionario actualizado exitosamente'})
        flash('¡Cuestionario actualizado exitosamente!', 'success')
        return redirect(url_for('mis_cuestionarios'))
    except Exception as e:
        is_ajax = request.headers.get('X-Requested-With') == 'XMLHttpRequest' or request.is_json

        if is_ajax:
            return jsonify({'success': False, 'error': str(e)}), 400
        flash(f'Error al actualizar el cuestionario: {str(e)}', 'error')
        return redirect(url_for('error_sistema_page'))

@app.route('/eliminar-cuestionario/<int:cuestionario_id>', methods=['POST'])
def eliminar_cuestionario_route(cuestionario_id):
    try:
        resultado = controlador_cuestionarios.eliminar_cuestionario(cuestionario_id)

        if not resultado:
            raise ValueError("No se pudo eliminar el cuestionario")

        if request.is_json:
            return jsonify({'success': True, 'message': 'Cuestionario eliminado exitosamente'})
        flash('¡Cuestionario eliminado exitosamente!', 'success')
        return redirect(url_for('mis_cuestionarios'))
    except Exception as e:
        if request.is_json:
            return jsonify({'success': False, 'error': str(e)}), 400
        flash(f'Error al eliminar el cuestionario: {str(e)}', 'error')
        return redirect(url_for('error_sistema_page'))

# Rutas para preguntas
@app.route('/crear-pregunta/<int:cuestionario_id>', methods=['POST'])
def crear_pregunta_route(cuestionario_id):
    data = request.get_json(silent=True) or request.form.to_dict()
    try:
        enunciado = data.get('enunciado', '').strip()
        tipo = data.get('tipo_pregunta', 'opcion_multiple')  # Corregido: usar 'tipo_pregunta'
        puntaje_base = int(data.get('puntaje_base', 1))
        tiempo_limite = int(data.get('tiempo_limite', 30))

        # Obtener opciones del formulario
        opciones_textos = request.form.getlist('opciones[]')
        respuesta_correcta_index = int(data.get('respuesta_correcta', 0))

        if not enunciado:
            raise ValueError("El enunciado es requerido")

        # Crear la pregunta
        pregunta_id = controlador_preguntas.crear_pregunta(enunciado, tipo, puntaje_base, tiempo_limite)

        # Crear las opciones si hay opciones de texto
        if opciones_textos and len(opciones_textos) > 0:
            for i, texto_opcion in enumerate(opciones_textos):
                if texto_opcion.strip():  # Solo crear si no está vacío
                    es_correcta = (i == respuesta_correcta_index)
                    controlador_opciones.crear_opcion_respuesta(pregunta_id, texto_opcion.strip(), es_correcta)

        # Obtener el siguiente orden para la pregunta en el cuestionario
        preguntas_existentes = controlador_preguntas.obtener_preguntas_por_cuestionario(cuestionario_id)
        orden = len(preguntas_existentes) + 1

        # Agregar la pregunta al cuestionario
        controlador_preguntas.agregar_pregunta_a_cuestionario(cuestionario_id, pregunta_id, orden)

        if request.is_json:
            return jsonify({'success': True, 'pregunta_id': pregunta_id, 'message': 'Pregunta creada exitosamente'})
        flash('¡Pregunta creada exitosamente!', 'success')
        return redirect(url_for('agregar_preguntas', cuestionario_id=cuestionario_id))
    except Exception as e:
        print(f"DEBUG: Error en crear_pregunta_route: {str(e)}")
        if request.is_json:
            return jsonify({'success': False, 'error': str(e)}), 400
        flash(f'Error al crear la pregunta: {str(e)}', 'error')
        return redirect(url_for('agregar_preguntas', cuestionario_id=cuestionario_id))

@app.route('/editar-pregunta/<int:pregunta_id>', methods=['POST'])
def editar_pregunta_route(pregunta_id):
    data = request.get_json(silent=True) or request.form.to_dict()
    try:
        enunciado = data.get('enunciado', '').strip()
        tipo = data.get('tipo', 'opcion_multiple')
        puntaje_base = int(data.get('puntaje_base', 1))
        tiempo_limite = int(data.get('tiempo_limite', 30))
        opciones_data = data.get('opciones', [])

        if not enunciado:
            raise ValueError("El enunciado es requerido")

        # Actualizar la pregunta
        resultado = controlador_preguntas.actualizar_pregunta(pregunta_id, enunciado, tipo, puntaje_base, tiempo_limite)

        if not resultado:
            raise ValueError("No se pudo actualizar la pregunta")

        # Actualizar las opciones si es de opción múltiple
        if tipo == 'opcion_multiple' and opciones_data:
            actualizar_opciones_multiple(pregunta_id, opciones_data)

        if request.is_json:
            return jsonify({'success': True, 'message': 'Pregunta actualizada exitosamente'})
        flash('¡Pregunta actualizada exitosamente!', 'success')
        return redirect(request.referrer or url_for('mis_cuestionarios'))
    except Exception as e:
        if request.is_json:
            return jsonify({'success': False, 'error': str(e)}), 400
        flash(f'Error al actualizar la pregunta: {str(e)}', 'error')
        return redirect(url_for('error_sistema_page'))

@app.route('/pregunta/<int:pregunta_id>/obtener', methods=['GET'])
@login_required
def obtener_pregunta_route(pregunta_id):
    """Obtener los datos de una pregunta para edición"""
    try:
        # Verificar que el usuario es docente
        if session.get('usuario_tipo') != 'docente':
            return jsonify({'success': False, 'error': 'Solo los docentes pueden obtener preguntas'}), 403

        # Obtener la pregunta
        pregunta = controlador_preguntas.obtener_pregunta_por_id(pregunta_id)

        if not pregunta:
            return jsonify({'success': False, 'error': 'Pregunta no encontrada'}), 404

        # Obtener las opciones de la pregunta
        opciones = controlador_opciones.obtener_opciones_por_pregunta(pregunta_id)

        pregunta['opciones'] = opciones

        return jsonify({
            'success': True,
            'pregunta': pregunta
        })

    except Exception as e:
        print(f"Error al obtener pregunta: {str(e)}")
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/pregunta/<int:pregunta_id>/eliminar', methods=['POST'])
@login_required
def eliminar_pregunta_route(pregunta_id):
    try:
        # Verificar que el usuario es docente
        if session.get('usuario_tipo') != 'docente':
            return jsonify({'success': False, 'error': 'Solo los docentes pueden eliminar preguntas'}), 403

        resultado = controlador_preguntas.eliminar_pregunta(pregunta_id)

        if not resultado:
            raise ValueError("No se pudo eliminar la pregunta")

        if request.is_json or request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return jsonify({'success': True, 'message': 'Pregunta eliminada exitosamente'})
        flash('¡Pregunta eliminada exitosamente!', 'success')
        return redirect(request.referrer or url_for('mis_cuestionarios'))
    except Exception as e:
        if request.is_json or request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return jsonify({'success': False, 'error': str(e)}), 400
        flash(f'Error al eliminar la pregunta: {str(e)}', 'error')
        return redirect(url_for('error_sistema_page'))

@app.route('/cuestionario/<int:id_cuestionario>/descargar-plantilla', methods=['GET'])
@login_required
def descargar_plantilla_excel(id_cuestionario):
    """Descargar plantilla de Excel para importar preguntas"""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        from io import BytesIO

        # Verificar que el usuario es docente
        if session.get('usuario_tipo') != 'docente':
            flash('Solo los docentes pueden descargar plantillas', 'error')
            return redirect(url_for('dashboard_docente'))

        # Verificar que el cuestionario existe y pertenece al docente
        cuestionario = controlador_cuestionarios.obtener_cuestionario_por_id(id_cuestionario)
        if not cuestionario or cuestionario.get('id_docente') != session.get('usuario_id'):
            flash('No tienes permiso para acceder a este cuestionario', 'error')
            return redirect(url_for('mis_cuestionarios'))

        # Crear libro de Excel
        wb = Workbook()
        ws = wb.active
        ws.title = "Preguntas"

        # Configurar encabezados
        headers = ['Pregunta', 'Opción A', 'Opción B', 'Opción C', 'Opción D', 'Respuesta Correcta', 'Tiempo (segundos)']
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")

        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')

        # Agregar fila de instrucciones
        instruction_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
        instruction_font = Font(italic=True, color="7F6000")

        instructions = [
            "Escribe aquí la pregunta",
            "Primera opción de respuesta",
            "Segunda opción de respuesta",
            "(Opcional) Tercera opción",
            "(Opcional) Cuarta opción",
            "Escribe la letra de la opción correcta: A, B, C o D",
            "Tiempo en segundos (ej: 30)"
        ]

        for col_num, instruction in enumerate(instructions, 1):
            cell = ws.cell(row=2, column=col_num)
            cell.value = instruction
            cell.fill = instruction_fill
            cell.font = instruction_font

        # Agregar ejemplo
        example_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
        example = [
            "¿Cuál es la capital de Francia?",
            "Madrid",
            "París",
            "Roma",
            "Berlín",
            "B",
            "30"
        ]

        for col_num, value in enumerate(example, 1):
            cell = ws.cell(row=3, column=col_num)
            cell.value = value
            cell.fill = example_fill

        # Agregar filas vacías para que el usuario complete
        for row in range(4, 24):
            for col in range(1, 8):
                ws.cell(row=row, column=col)

        # Ajustar ancho de columnas
        column_widths = [50, 30, 30, 30, 30, 20, 20]
        for col_num, width in enumerate(column_widths, 1):
            ws.column_dimensions[chr(64 + col_num)].width = width

        # Guardar en memoria
        excel_file = BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)

        # Crear nombre de archivo
        filename = f"plantilla_preguntas_{cuestionario['titulo'].replace(' ', '_')}.xlsx"

        return send_file(
            excel_file,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )

    except Exception as e:
        print(f"Error al generar plantilla Excel: {str(e)}")
        flash(f'Error al generar la plantilla: {str(e)}', 'error')
        return redirect(url_for('editar_cuestionario', cuestionario_id=id_cuestionario))

@app.route('/cuestionario/<int:id_cuestionario>/importar-preguntas', methods=['POST'])
@login_required
def importar_preguntas_excel(id_cuestionario):
    """Importar preguntas desde un archivo Excel"""
    import tempfile
    import os

    temp_file_path = None
    try:
        from openpyxl import load_workbook

        # Verificar que el usuario es docente
        if session.get('usuario_tipo') != 'docente':
            return jsonify({'success': False, 'error': 'Solo los docentes pueden importar preguntas'}), 403

        # Verificar que el cuestionario existe y pertenece al docente
        cuestionario = controlador_cuestionarios.obtener_cuestionario_por_id(id_cuestionario)
        if not cuestionario or cuestionario.get('id_docente') != session.get('usuario_id'):
            return jsonify({'success': False, 'error': 'No tienes permiso para acceder a este cuestionario'}), 403

        # Verificar que se envió un archivo
        if 'excel_file' not in request.files:
            return jsonify({'success': False, 'error': 'No se envió ningún archivo'}), 400

        file = request.files['excel_file']

        if file.filename == '':
            return jsonify({'success': False, 'error': 'No se seleccionó ningún archivo'}), 400

        if not file.filename.endswith(('.xlsx', '.xls')):
            return jsonify({'success': False, 'error': 'El archivo debe ser de formato Excel (.xlsx o .xls)'}), 400

        # Guardar el archivo temporalmente
        # Esto es necesario en PythonAnywhere y otros servidores
        temp_fd, temp_file_path = tempfile.mkstemp(suffix='.xlsx')
        os.close(temp_fd)  # Cerrar el file descriptor
        file.save(temp_file_path)

        # Cargar el archivo Excel desde el archivo temporal
        wb = load_workbook(temp_file_path)
        ws = wb.active

        print(f"DEBUG: Archivo Excel cargado. Hojas disponibles: {wb.sheetnames}")
        print(f"DEBUG: Hoja activa: {ws.title}")
        print(f"DEBUG: Total de filas en la hoja: {ws.max_row}")
        print(f"DEBUG: Total de columnas en la hoja: {ws.max_column}")

        preguntas_importadas = []
        errores = []

        # Mostrar las primeras 5 filas para depuración
        print("\n=== CONTENIDO DEL EXCEL ===")
        for i, row in enumerate(ws.iter_rows(min_row=1, max_row=5, values_only=True), start=1):
            print(f"Fila {i}: {row}")
        print("=========================\n")

        # Procesar filas (empezando desde la fila 4 para saltar encabezados, instrucciones y ejemplo)
        filas_procesadas = 0
        for row_num, row in enumerate(ws.iter_rows(min_row=4, values_only=True), start=4):
            filas_procesadas += 1
            print(f"\nDEBUG: Procesando fila {row_num}: {row}")

            # Saltar filas vacías
            if not row[0] or str(row[0]).strip() == '':
                print(f"DEBUG: Fila {row_num} vacía o sin pregunta, saltando...")
                continue

            try:
                # Extraer datos de la fila
                texto_pregunta = str(row[0]).strip()
                opcion_a = str(row[1]).strip() if row[1] else ''
                opcion_b = str(row[2]).strip() if row[2] else ''
                opcion_c = str(row[3]).strip() if row[3] else ''
                opcion_d = str(row[4]).strip() if row[4] else ''
                respuesta_correcta = str(row[5]).strip().upper() if row[5] else ''

                # Leer tiempo y asegurar que sea un entero en segundos
                if row[6]:
                    # Convertir a float primero por si acaso tiene decimales, luego a int
                    tiempo_raw = float(row[6])
                    # Si el valor es muy grande (> 600), probablemente está en minutos, convertir a segundos
                    tiempo = int(tiempo_raw) if tiempo_raw <= 600 else int(tiempo_raw / 60)
                    print(f"DEBUG: Tiempo leído del Excel: {row[6]} -> convertido a: {tiempo} segundos")
                else:
                    tiempo = cuestionario.get('tiempo_limite_pregunta', 30)
                    print(f"DEBUG: Tiempo no especificado, usando default: {tiempo} segundos")

                # Validaciones
                if not texto_pregunta:
                    errores.append(f"Fila {row_num}: La pregunta no puede estar vacía")
                    continue

                if not opcion_a or not opcion_b:
                    errores.append(f"Fila {row_num}: Se requieren al menos 2 opciones (A y B)")
                    continue

                if respuesta_correcta not in ['A', 'B', 'C', 'D']:
                    errores.append(f"Fila {row_num}: La respuesta correcta debe ser A, B, C o D")
                    continue

                # Validar que la respuesta correcta corresponde a una opción existente
                opciones_disponibles = []
                if opcion_a: opciones_disponibles.append('A')
                if opcion_b: opciones_disponibles.append('B')
                if opcion_c: opciones_disponibles.append('C')
                if opcion_d: opciones_disponibles.append('D')

                if respuesta_correcta not in opciones_disponibles:
                    errores.append(f"Fila {row_num}: La respuesta correcta '{respuesta_correcta}' no tiene opción asociada")
                    continue

                if tiempo < 5 or tiempo > 300:
                    errores.append(f"Fila {row_num}: El tiempo debe estar entre 5 y 300 segundos")
                    continue

                # Crear la pregunta con la firma correcta
                print(f"DEBUG: Creando pregunta: '{texto_pregunta}'")
                pregunta_id = controlador_preguntas.crear_pregunta(
                    enunciado=texto_pregunta,
                    tipo='opcion_multiple',
                    cuestionario_id=id_cuestionario,
                    puntaje_base=1,
                    tiempo_limite=tiempo
                )

                print(f"DEBUG: Pregunta creada con ID: {pregunta_id}")

                if not pregunta_id:
                    errores.append(f"Fila {row_num}: Error al crear la pregunta en la base de datos")
                    continue

                # Eliminar opciones antiguas si la pregunta ya existía (para actualizarlas)
                conexion_temp = obtener_conexion()
                with conexion_temp.cursor() as cursor_temp:
                    cursor_temp.execute("DELETE FROM opciones_respuesta WHERE id_pregunta = %s", (pregunta_id,))
                    conexion_temp.commit()
                conexion_temp.close()
                print(f"DEBUG: Opciones antiguas eliminadas para pregunta {pregunta_id}")

                # Crear las opciones
                opciones_data = []
                if opcion_a:
                    opciones_data.append({
                        'texto': opcion_a,
                        'es_correcta': 1 if respuesta_correcta == 'A' else 0
                    })
                if opcion_b:
                    opciones_data.append({
                        'texto': opcion_b,
                        'es_correcta': 1 if respuesta_correcta == 'B' else 0
                    })
                if opcion_c:
                    opciones_data.append({
                        'texto': opcion_c,
                        'es_correcta': 1 if respuesta_correcta == 'C' else 0
                    })
                if opcion_d:
                    opciones_data.append({
                        'texto': opcion_d,
                        'es_correcta': 1 if respuesta_correcta == 'D' else 0
                    })

                print(f"DEBUG: Creando {len(opciones_data)} opciones para pregunta {pregunta_id}")

                # Guardar las opciones
                opciones_creadas = True
                for idx, opcion_data in enumerate(opciones_data):
                    resultado = controlador_opciones.crear_opcion(
                        id_pregunta=pregunta_id,
                        texto_opcion=opcion_data['texto'],
                        es_correcta=opcion_data['es_correcta'],
                        explicacion=''
                    )
                    print(f"DEBUG: Opción {idx+1} '{opcion_data['texto']}' creada: {resultado}")
                    if not resultado:
                        opciones_creadas = False
                        break

                if not opciones_creadas:
                    # Si falla la creación de opciones, eliminar la pregunta
                    print(f"DEBUG: Error al crear opciones, eliminando pregunta {pregunta_id}")
                    controlador_preguntas.eliminar_pregunta(pregunta_id)
                    errores.append(f"Fila {row_num}: Error al crear las opciones de respuesta")
                    continue

                # La pregunta ya fue asociada al cuestionario por crear_pregunta()
                # Solo agregamos a la lista de importadas
                preguntas_importadas.append({
                    'id_pregunta': pregunta_id,
                    'texto': texto_pregunta,
                    'orden': len(preguntas_importadas) + 1
                })

                print(f"DEBUG: Pregunta importada exitosamente: {texto_pregunta}")

            except Exception as e:
                print(f"DEBUG: Error al procesar fila {row_num}: {str(e)}")
                errores.append(f"Fila {row_num}: {str(e)}")
                continue

        print(f"\n=== RESUMEN DE IMPORTACIÓN ===")
        print(f"Filas procesadas: {filas_procesadas}")
        print(f"Preguntas importadas: {len(preguntas_importadas)}")
        print(f"Errores encontrados: {len(errores)}")
        print("============================\n")

        # Preparar respuesta
        total_importadas = len(preguntas_importadas)

        if total_importadas == 0 and len(errores) == 0:
            return jsonify({
                'success': False,
                'error': 'No se encontraron preguntas válidas en el archivo. Asegúrate de seguir el formato de la plantilla.'
            }), 400

        return jsonify({
            'success': True,
            'total_importadas': total_importadas,
            'errores': errores,
            'preguntas': preguntas_importadas
        })

    except Exception as e:
        print(f"Error al importar preguntas: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'error': f'Error al procesar el archivo: {str(e)}'}), 500
    finally:
        # Limpiar el archivo temporal
        if temp_file_path and os.path.exists(temp_file_path):
            try:
                os.remove(temp_file_path)
                print(f"DEBUG: Archivo temporal eliminado: {temp_file_path}")
            except Exception as e:
                print(f"DEBUG: Error al eliminar archivo temporal: {str(e)}")

@app.route('/pregunta/<int:pregunta_id>/editar', methods=['POST'])
@login_required
def editar_pregunta_ajax(pregunta_id):
    """Editar una pregunta existente"""
    try:
        print(f"DEBUG: Editando pregunta ID: {pregunta_id}")

        # Verificar que el usuario es docente
        if session.get('usuario_tipo') != 'docente':
            return jsonify({'success': False, 'error': 'Solo los docentes pueden editar preguntas'}), 403

        # Obtener datos del formulario
        enunciado = request.form.get('enunciado', '').strip()
        tiempo = int(request.form.get('tiempo', 30))
        opciones_json = request.form.get('opciones', '[]')

        print(f"DEBUG: Datos recibidos - enunciado: '{enunciado}', tiempo: {tiempo}")
        print(f"DEBUG: Opciones JSON: {opciones_json}")

        if not enunciado:
            return jsonify({'success': False, 'error': 'El enunciado es requerido'}), 400

        # Parsear opciones
        import json
        opciones = json.loads(opciones_json)

        if not opciones or len(opciones) < 2:
            return jsonify({'success': False, 'error': 'Debe haber al menos 2 opciones'}), 400

        # Verificar que hay una opción correcta
        if not any(opcion.get('es_correcta') for opcion in opciones):
            return jsonify({'success': False, 'error': 'Debe seleccionar una respuesta correcta'}), 400

        # Actualizar la pregunta
        resultado = controlador_preguntas.actualizar_pregunta(
            pregunta_id,
            enunciado,
            'opcion_multiple',
            1,  # puntaje_base
            tiempo
        )

        if not resultado:
            return jsonify({'success': False, 'error': 'No se pudo actualizar la pregunta'}), 500

        # Eliminar opciones anteriores
        controlador_opciones.eliminar_opciones_pregunta(pregunta_id)

        # Crear nuevas opciones
        for opcion in opciones:
            controlador_opciones.crear_opcion(
                pregunta_id,
                opcion['texto_opcion'],
                opcion['es_correcta'],
                ''  # explicacion
            )

        # Obtener la pregunta actualizada con sus opciones
        pregunta_actualizada = controlador_preguntas.obtener_pregunta_por_id(pregunta_id)
        opciones_actualizadas = controlador_opciones.obtener_opciones_por_pregunta(pregunta_id)

        pregunta_data = {
            'id_pregunta': pregunta_actualizada['id_pregunta'],
            'enunciado': pregunta_actualizada['enunciado'],
            'tiempo': pregunta_actualizada.get('tiempo_limite', 30),
            'opciones': opciones_actualizadas
        }

        return jsonify({
            'success': True,
            'message': 'Pregunta actualizada exitosamente',
            'pregunta': pregunta_data
        })

    except json.JSONDecodeError as e:
        print(f"DEBUG: Error parsing JSON: {str(e)}")
        return jsonify({'success': False, 'error': 'Formato de opciones inválido'}), 400
    except Exception as e:
        print(f"DEBUG: Error en editar_pregunta_ajax: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/pregunta/<int:pregunta_id>/opciones')
def obtener_opciones_pregunta(pregunta_id):
    try:
        opciones = controlador_opciones.obtener_opciones_por_pregunta(pregunta_id)
        return jsonify({'success': True, 'opciones': opciones})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 400

@app.route('/pregunta/<int:pregunta_id>/obtener')
@login_required
def obtener_pregunta(pregunta_id):
    """Obtener los datos de una pregunta específica"""
    try:
        print(f"DEBUG: Obteniendo pregunta ID: {pregunta_id}")

        # Obtener la pregunta
        pregunta = controlador_preguntas.obtener_pregunta_por_id(pregunta_id)
        if not pregunta:
            return jsonify({'success': False, 'error': 'Pregunta no encontrada'}), 404

        # Obtener las opciones
        opciones = controlador_opciones.obtener_opciones_por_pregunta(pregunta_id)

        # Formatear la respuesta
        pregunta_data = {
            'id_pregunta': pregunta['id_pregunta'],
            'enunciado': pregunta['enunciado'],
            'tipo': pregunta.get('tipo', 'opcion_multiple'),
            'tiempo': pregunta.get('tiempo_limite', 30),
            'opciones': opciones
        }

        print(f"DEBUG: Pregunta encontrada: {pregunta_data}")
        return jsonify({'success': True, 'pregunta': pregunta_data})

    except Exception as e:
        print(f"DEBUG: Error en obtener_pregunta: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/cuestionario/<int:cuestionario_id>/agregar-pregunta', methods=['POST'])
@login_required
def agregar_pregunta_cuestionario(cuestionario_id):
    """Agregar una nueva pregunta a un cuestionario"""
    try:
        print(f"DEBUG: agregar_pregunta_cuestionario - cuestionario_id: {cuestionario_id}")

        # Verificar que el usuario es docente
        if session.get('usuario_tipo') != 'docente':
            return jsonify({'success': False, 'error': 'Solo los docentes pueden agregar preguntas'}), 403

        # Verificar que el cuestionario existe y pertenece al docente
        cuestionario = controlador_cuestionarios.obtener_cuestionario_por_id(cuestionario_id)
        if not cuestionario:
            return jsonify({'success': False, 'error': 'Cuestionario no encontrado'}), 404

        id_docente_actual = session.get('usuario_id')
        if cuestionario.get('id_docente') != id_docente_actual:
            return jsonify({'success': False, 'error': 'No tienes permisos para agregar preguntas a este cuestionario'}), 403

        # Obtener datos del formulario
        enunciado = request.form.get('enunciado', '').strip()
        tiempo = request.form.get('tiempo', 30)
        opciones_json = request.form.get('opciones', '[]')

        print(f"DEBUG: Datos recibidos - enunciado: '{enunciado}', tiempo: {tiempo}")
        print(f"DEBUG: Opciones JSON: {opciones_json}")

        if not enunciado:
            return jsonify({'success': False, 'error': 'El enunciado de la pregunta es requerido'}), 400

        # Parsear opciones
        import json
        try:
            opciones = json.loads(opciones_json)
        except:
            return jsonify({'success': False, 'error': 'Formato de opciones inválido'}), 400

        if len(opciones) < 2:
            return jsonify({'success': False, 'error': 'Debe haber al menos 2 opciones'}), 400

        # Verificar que hay una opción correcta
        tiene_correcta = any(opcion.get('es_correcta', False) for opcion in opciones)
        if not tiene_correcta:
            return jsonify({'success': False, 'error': 'Debe seleccionar una respuesta correcta'}), 400

        # Crear la pregunta (ya la asocia al cuestionario internamente)
        pregunta_id = controlador_preguntas.crear_pregunta(
            enunciado=enunciado,
            tipo='opcion_multiple',
            cuestionario_id=cuestionario_id,
            puntaje_base=1,
            tiempo_limite=int(tiempo)
        )

        print(f"DEBUG: Pregunta creada con ID: {pregunta_id}")

        if not pregunta_id:
            return jsonify({'success': False, 'error': 'Error al crear la pregunta'}), 500

        # Crear las opciones
        for i, opcion in enumerate(opciones):
            texto_opcion = opcion.get('texto_opcion', '').strip()
            es_correcta = opcion.get('es_correcta', False)

            if texto_opcion:
                controlador_opciones.crear_opcion(
                    id_pregunta=pregunta_id,
                    texto_opcion=texto_opcion,
                    es_correcta=es_correcta,
                    explicacion=''
                )
                print(f"DEBUG: Opción {i+1} creada: '{texto_opcion}' (correcta: {es_correcta})")

        # Obtener la pregunta completa para devolverla
        pregunta_completa = controlador_preguntas.obtener_pregunta_por_id(pregunta_id)
        opciones_pregunta = controlador_opciones.obtener_opciones_por_pregunta(pregunta_id)

        pregunta_response = {
            'id_pregunta': pregunta_id,
            'enunciado': enunciado,
            'tiempo': tiempo,
            'opciones': [{
                'texto_opcion': opc.get('texto_opcion', ''),
                'es_correcta': opc.get('es_correcta', False)
            } for opc in opciones_pregunta]
        }

        return jsonify({
            'success': True,
            'message': 'Pregunta agregada exitosamente',
            'pregunta': pregunta_response
        })

    except Exception as e:
        print(f"DEBUG: Error en agregar_pregunta_cuestionario: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 500




@app.route('/pregunta/<int:pregunta_id>/eliminar', methods=['POST'])
@login_required
def eliminar_pregunta_cuestionario(pregunta_id):
    """Eliminar una pregunta"""
    try:
        print(f"DEBUG: eliminar_pregunta_cuestionario - pregunta_id: {pregunta_id}")

        # Verificar que el usuario es docente
        if session.get('usuario_tipo') != 'docente':
            return jsonify({'success': False, 'error': 'Solo los docentes pueden eliminar preguntas'}), 403

        # Verificar que la pregunta existe
        pregunta = controlador_preguntas.obtener_pregunta_por_id(pregunta_id)
        if not pregunta:
            return jsonify({'success': False, 'error': 'Pregunta no encontrada'}), 404

        # Eliminar la pregunta (esto también eliminará las opciones por CASCADE)
        resultado = controlador_preguntas.eliminar_pregunta(pregunta_id)

        if resultado:
            return jsonify({
                'success': True,
                'message': 'Pregunta eliminada exitosamente'
            })
        else:
            return jsonify({'success': False, 'error': 'Error al eliminar la pregunta'}), 500

    except Exception as e:
        print(f"DEBUG: Error en eliminar_pregunta_cuestionario: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/juego-estudiante/<pin_sala>')
def juego_estudiante_por_pin(pin_sala):
    # Aquí iría la lógica para obtener las preguntas del juego
    return render_template('JuegoEstudiante.html', pin_sala=pin_sala)

@app.route('/monitoreo-juego/<pin_sala>')
def monitoreo_juego(pin_sala):
    # Aquí iría la lógica para obtener el estado actual del juego
    return render_template('MonitoreoJuego.html', pin_sala=pin_sala)

@app.route('/resultados-juego/<pin_sala>')
def resultados_juego(pin_sala):
    # Aquí iría la lógica para obtener los resultados finales
    return render_template('ResultadosJuego.html', pin_sala=pin_sala)

###############################
# RUTAS CRUD PARA EL DASHBOARD
###############################

#-----INICIO-USUARIO-----#
@app.route('/usuario')
def usuario():
    usuarios = controlador_usuario.obtener_usuarios()
    return render_template('dashboard/usuario.html', usuarios=usuarios)

@app.route('/registrar_usuario')
def registrar_usuario():
    return render_template('dashboard/registrar_usuario.html')

@app.route('/insertar_usuario', methods=['POST'])
def insertar_usuario():
    nombre = request.form['nombre']
    apellidos = request.form['apellidos']
    email = request.form['email']
    password = request.form['password']
    tipo_usuario = request.form['tipo_usuario']
    controlador_usuario.crear_usuario(
        nombre=nombre,
        apellidos=apellidos,
        email=email,
        password=password,
        tipo_usuario=tipo_usuario
    )
    return redirect(url_for('usuario'))

@app.route('/eliminar_usuario/<int:id>')
def eliminar_usuario(id):
    controlador_usuario.eliminar_usuario(id)
    return redirect(url_for('usuario'))

@app.route('/modificar_usuario', methods=['POST'])
def modificar_usuario():
    id = request.form['id']
    usuario = controlador_usuario.obtener_usuario_por_id(id)
    return render_template('dashboard/modificar_usuario.html', usuario=usuario)

@app.route('/actualizar_usuario', methods=['POST'])
def actualizar_usuario():
    id = request.form['id']
    nombre = request.form['nombre']
    apellidos = request.form['apellidos']
    email = request.form['email']
    tipo_usuario = request.form['tipo_usuario']
    controlador_usuario.actualizar_usuario(id, nombre, apellidos, email, tipo_usuario)
    return redirect(url_for('usuario'))

@app.route('/suspender_usuario', methods=['POST'])
def suspender_usuario():
    id = request.form['id']
    controlador_usuario.cambiar_estado_usuario(id, 'suspendido')
    return redirect(url_for('usuario'))

@app.route('/activar_usuario', methods=['POST'])
def activar_usuario():
    id = request.form['id']
    controlador_usuario.cambiar_estado_usuario(id, 'activo')
    return redirect(url_for('usuario'))
#-----FIN-USUARIO-----#

#-----INICIO-CUESTIONARIO-----#
@app.route('/cuestionario')
def cuestionario():
    cuestionarios = controlador_cuestionarios.obtener_cuestionarios()
    return render_template('dashboard/cuestionario.html', cuestionarios=cuestionarios)

@app.route('/registrar_cuestionario')
def registrar_cuestionario():
    docentes = controlador_usuario.obtener_usuarios_por_tipo('docente')
    return render_template('dashboard/registrar_cuestionario.html', docentes=docentes)

@app.route('/insertar_cuestionario', methods=['POST'])
def insertar_cuestionario():
    titulo = request.form['titulo']
    descripcion = request.form['descripcion']
    id_docente = request.form['id_docente']
    categoria = request.form.get('categoria', '')
    tiempo_limite = request.form.get('tiempo_limite', 0)
    max_intentos = request.form.get('max_intentos', 1)

    controlador_cuestionarios.crear_cuestionario(
        titulo=titulo,
        descripcion=descripcion,
        id_docente=id_docente,
        categoria=categoria,
        tiempo_limite=tiempo_limite,
        max_intentos=max_intentos
    )
    return redirect(url_for('cuestionario'))

@app.route('/eliminar_cuestionario/<int:id>')
def eliminar_cuestionario(id):
    controlador_cuestionarios.eliminar_cuestionario(id)
    return redirect(url_for('cuestionario'))

@app.route('/modificar_cuestionario', methods=['POST'])
def modificar_cuestionario():
    id = request.form['id']
    cuestionario = controlador_cuestionarios.obtener_cuestionario_por_id(id)
    docentes = controlador_usuario.obtener_usuarios_por_tipo('docente')
    return render_template('dashboard/modificar_cuestionario.html', cuestionario=cuestionario, docentes=docentes)

@app.route('/actualizar_cuestionario', methods=['POST'])
def actualizar_cuestionario():
    id = request.form['id']
    titulo = request.form['titulo']
    descripcion = request.form['descripcion']
    id_docente = request.form['id_docente']
    categoria = request.form.get('categoria', '')
    tiempo_limite = request.form.get('tiempo_limite', 0)
    max_intentos = request.form.get('max_intentos', 1)

    controlador_cuestionarios.actualizar_cuestionario(id, titulo, descripcion, id_docente, categoria, tiempo_limite, max_intentos)
    return redirect(url_for('cuestionario'))

@app.route('/publicar_cuestionario', methods=['POST'])
def publicar_cuestionario_dashboard():
    id = request.form['id']
    controlador_cuestionarios.cambiar_estado_cuestionario(id, 'publicado')
    return redirect(url_for('cuestionario'))

@app.route('/despublicar_cuestionario', methods=['POST'])
def despublicar_cuestionario():
    id = request.form['id']
    controlador_cuestionarios.cambiar_estado_cuestionario(id, 'borrador')
    return redirect(url_for('cuestionario'))
#-----FIN-CUESTIONARIO-----#

#-----INICIO-PREGUNTA-----#
@app.route('/pregunta')
def pregunta():
    preguntas = controlador_preguntas.obtener_preguntas()
    return render_template('dashboard/pregunta.html', preguntas=preguntas)

@app.route('/registrar_pregunta')
def registrar_pregunta():
    cuestionarios = controlador_cuestionarios.obtener_cuestionarios()
    return render_template('dashboard/registrar_pregunta.html', cuestionarios=cuestionarios)

@app.route('/insertar_pregunta', methods=['POST'])
def insertar_pregunta():
    enunciado = request.form['enunciado']
    tipo = request.form['tipo']
    puntaje_base = request.form.get('puntaje_base', 1)
    tiempo_limite = request.form.get('tiempo_limite', 30)
    archivo_multimedia = request.form.get('archivo_multimedia', '')
    id_cuestionario = request.form.get('id_cuestionario')

    pregunta_id = controlador_preguntas.crear_pregunta(
        enunciado=enunciado,
        tipo=tipo,
        puntaje_base=puntaje_base,
        tiempo_limite=tiempo_limite,
        archivo_multimedia=archivo_multimedia
    )

    # Si se especificó un cuestionario, agregar la pregunta
    if id_cuestionario:
        orden = controlador_cuestionario_preguntas.contar_preguntas_cuestionario(id_cuestionario) + 1
        controlador_preguntas.agregar_pregunta_a_cuestionario(id_cuestionario, pregunta_id, orden)

    return redirect(url_for('pregunta'))

@app.route('/eliminar_pregunta/<int:id>')
def eliminar_pregunta(id):
    controlador_preguntas.eliminar_pregunta(id)
    return redirect(url_for('pregunta'))

@app.route('/modificar_pregunta', methods=['POST'])
def modificar_pregunta():
    id = request.form['id']
    pregunta = controlador_preguntas.obtener_pregunta_por_id(id)
    cuestionarios = controlador_cuestionarios.obtener_cuestionarios()
    return render_template('dashboard/modificar_pregunta.html', pregunta=pregunta, cuestionarios=cuestionarios)

@app.route('/actualizar_pregunta', methods=['POST'])
def actualizar_pregunta():
    id = request.form['id']
    enunciado = request.form['enunciado']
    tipo = request.form['tipo']
    puntaje_base = request.form.get('puntaje_base', 1)
    tiempo_limite = request.form.get('tiempo_limite', 30)
    archivo_multimedia = request.form.get('archivo_multimedia', '')

    controlador_preguntas.actualizar_pregunta(id, enunciado, tipo, puntaje_base, tiempo_limite, archivo_multimedia)
    return redirect(url_for('pregunta'))
#-----FIN-PREGUNTA-----#

#-----INICIO-OPCION-----#
@app.route('/opcion')
def opcion():
    opciones = controlador_opciones.obtener_opciones()
    return render_template('dashboard/opcion.html', opciones=opciones)

@app.route('/registrar_opcion')
def registrar_opcion():
    preguntas = controlador_preguntas.obtener_preguntas()
    return render_template('dashboard/registrar_opcion.html', preguntas=preguntas)

@app.route('/insertar_opcion', methods=['POST'])
def insertar_opcion():
    id_pregunta = request.form['id_pregunta']
    texto_opcion = request.form['texto_opcion']
    es_correcta = 'es_correcta' in request.form
    explicacion = request.form.get('explicacion', '')

    controlador_opciones.crear_opcion(
        id_pregunta=id_pregunta,
        texto_opcion=texto_opcion,
        es_correcta=es_correcta,
        explicacion=explicacion
    )
    return redirect(url_for('opcion'))

@app.route('/eliminar_opcion/<int:id>')
def eliminar_opcion(id):
    controlador_opciones.eliminar_opcion(id)
    return redirect(url_for('opcion'))

@app.route('/modificar_opcion', methods=['POST'])
def modificar_opcion():
    id = request.form['id']
    opcion = controlador_opciones.obtener_opcion_por_id(id)
    preguntas = controlador_preguntas.obtener_preguntas()
    return render_template('dashboard/modificar_opcion.html', opcion=opcion, preguntas=preguntas)

@app.route('/actualizar_opcion', methods=['POST'])
def actualizar_opcion():
    id = request.form['id']
    id_pregunta = request.form['id_pregunta']
    texto_opcion = request.form['texto_opcion']
    es_correcta = 'es_correcta' in request.form
    explicacion = request.form.get('explicacion', '')

    controlador_opciones.actualizar_opcion(id, id_pregunta, texto_opcion, es_correcta, explicacion)
    return redirect(url_for('opcion'))
#-----FIN-OPCION-----#

#-----INICIO-SALA-----#
@app.route('/sala')
def sala():
    salas = controlador_salas.obtener_salas()
    return render_template('dashboard/sala.html', salas=salas)

@app.route('/registrar_sala')
def registrar_sala():
    cuestionarios = controlador_cuestionarios.obtener_cuestionarios_publicados()
    docentes = controlador_usuario.obtener_usuarios_por_tipo('docente')
    return render_template('dashboard/registrar_sala.html', cuestionarios=cuestionarios, docentes=docentes)

@app.route('/insertar_sala', methods=['POST'])
def insertar_sala():
    nombre = request.form['nombre']
    id_cuestionario = request.form['id_cuestionario']
    id_docente = request.form['id_docente']
    tipo_sala = request.form.get('tipo_sala', 'individual')
    max_participantes = request.form.get('max_participantes', 30)
    tiempo_respuesta = request.form.get('tiempo_respuesta', 30)

    controlador_salas.crear_sala(
        nombre=nombre,
        id_cuestionario=id_cuestionario,
        id_docente=id_docente,
        tipo_sala=tipo_sala,
        max_participantes=max_participantes,
        tiempo_respuesta=tiempo_respuesta
    )
    return redirect(url_for('sala'))

@app.route('/eliminar_sala/<int:id>')
def eliminar_sala(id):
    controlador_salas.eliminar_sala(id)
    return redirect(url_for('sala'))

@app.route('/modificar_sala', methods=['POST'])
def modificar_sala():
    id = request.form['id']
    sala = controlador_salas.obtener_sala_por_id(id)
    cuestionarios = controlador_cuestionarios.obtener_cuestionarios_publicados()
    docentes = controlador_usuario.obtener_usuarios_por_tipo('docente')
    return render_template('dashboard/modificar_sala.html', sala=sala, cuestionarios=cuestionarios, docentes=docentes)

@app.route('/actualizar_sala', methods=['POST'])
def actualizar_sala():
    id = request.form['id']
    nombre = request.form['nombre']
    id_cuestionario = request.form['id_cuestionario']
    id_docente = request.form['id_docente']
    tipo_sala = request.form.get('tipo_sala', 'individual')
    max_participantes = request.form.get('max_participantes', 30)
    tiempo_respuesta = request.form.get('tiempo_respuesta', 30)

    controlador_salas.actualizar_sala(id, nombre, id_cuestionario, id_docente, tipo_sala, max_participantes, tiempo_respuesta)
    return redirect(url_for('sala'))

@app.route('/iniciar_sala', methods=['POST'])
def iniciar_sala_dashboard():
    id = request.form['id']
    controlador_salas.cambiar_estado_sala(id, 'activa')
    return redirect(url_for('sala'))

@app.route('/cerrar_sala', methods=['POST'])
def cerrar_sala_dashboard():
    id = request.form['id']
    controlador_salas.cambiar_estado_sala(id, 'finalizada')
    return redirect(url_for('sala'))
#-----FIN-SALA-----#

#-----INICIO-PARTICIPACION-----#
@app.route('/participacion')
def participacion():
    participaciones = controlador_participaciones.obtener_participaciones()
    return render_template('dashboard/participacion.html', participaciones=participaciones)

@app.route('/ver_participacion/<int:id>')
def ver_participacion(id):
    participacion = controlador_participaciones.obtener_participacion_por_id(id)
    respuestas = controlador_respuestas.obtener_respuestas_por_participacion(id)
    return render_template('dashboard/ver_participacion.html', participacion=participacion, respuestas=respuestas)

@app.route('/finalizar_participacion', methods=['POST'])
def finalizar_participacion():
    id = request.form['id']
    puntaje_total = request.form.get('puntaje_total', 0)
    controlador_participaciones.finalizar_participacion(id, puntaje_total)
    return redirect(url_for('participacion'))

@app.route('/eliminar_participacion/<int:id>')
def eliminar_participacion(id):
    controlador_participaciones.eliminar_participacion(id)
    return redirect(url_for('participacion'))
#-----FIN-PARTICIPACION-----#

#-----INICIO-RESPUESTA-----#
@app.route('/respuesta')
def respuesta():
    respuestas = controlador_respuestas.obtener_respuestas()
    return render_template('dashboard/respuesta.html', respuestas=respuestas)

@app.route('/calificar_respuesta', methods=['POST'])
def calificar_respuesta():
    id = request.form['id']
    puntaje_obtenido = request.form['puntaje_obtenido']
    es_correcta = 'es_correcta' in request.form
    controlador_respuestas.calificar_respuesta(id, puntaje_obtenido, es_correcta)
    return redirect(url_for('respuesta'))

@app.route('/ver_respuesta/<int:id>')
def ver_respuesta(id):
    respuesta = controlador_respuestas.obtener_respuesta_por_id(id)
    return render_template('dashboard/ver_respuesta.html', respuesta=respuesta)

@app.route('/eliminar_respuesta/<int:id>')
def eliminar_respuesta(id):
    controlador_respuestas.eliminar_respuesta(id)
    return redirect(url_for('respuesta'))
#-----FIN-RESPUESTA-----#

#-----INICIO-RANKING-----#
@app.route('/ranking')
def ranking():
    ranking_global = controlador_ranking.obtener_ranking_global()
    return render_template('dashboard/ranking.html', ranking=ranking_global)

@app.route('/ranking_cuestionario/<int:id_cuestionario>')
def ranking_cuestionario(id_cuestionario):
    ranking = controlador_ranking.obtener_ranking_cuestionario(id_cuestionario)
    cuestionario = controlador_cuestionarios.obtener_cuestionario_por_id(id_cuestionario)
    return render_template('dashboard/ranking_cuestionario.html', ranking=ranking, cuestionario=cuestionario)

@app.route('/ranking_sala/<int:id_sala>')
def ranking_sala(id_sala):
    ranking = controlador_ranking.obtener_ranking_sala(id_sala)
    sala = controlador_salas.obtener_sala_por_id(id_sala)
    return render_template('dashboard/ranking_sala.html', ranking=ranking, sala=sala)

@app.route('/actualizar_ranking_cuestionario', methods=['POST'])
def actualizar_ranking_cuestionario():
    id_cuestionario = request.form['id_cuestionario']
    controlador_ranking.actualizar_ranking_completo(id_cuestionario)
    return redirect(url_for('ranking_cuestionario', id_cuestionario=id_cuestionario))

@app.route('/estadisticas_ranking')
def estadisticas_ranking():
    estadisticas = controlador_ranking.obtener_estadisticas_ranking(1)  # Ejemplo con ID 1
    return render_template('dashboard/estadisticas_ranking.html', estadisticas=estadisticas)
#-----FIN-RANKING-----#

# ----- RECOMPENSAS GENERALES (ADMIN / DASHBOARD) -----
@app.route('/recompensas')
def obtener_todas_recompensas():
    recompensas = controlador_recompensas.obtener_todas_recompensas()
    return render_template('dashboard/recompensa.html', recompensas=recompensas)


@app.route('/registrar_recompensa')
def registrar_recompensa():
    return render_template('dashboard/registrar_recompensa.html')


@app.route('/insertar_recompensa_admin', methods=['POST'])
def insertar_recompensa_admin():
    nombre = request.form['nombre']
    descripcion = request.form['descripcion']
    puntos_requeridos = request.form['puntos_requeridos']
    tipo = request.form['tipo']

    # Si no hay cuestionario asociado (admin), usa NULL
    controlador_recompensas.insertar_recompensa(nombre, descripcion, puntos_requeridos, tipo, None)
    return redirect(url_for('obtener_todas_recompensas'))


@app.route('/eliminar_recompensa_admin/<int:id>')
def eliminar_recompensa_admin(id):
    controlador_recompensas.eliminar_recompensa(id)
    return redirect(url_for('obtener_todas_recompensas'))


@app.route('/modificar_recompensa', methods=['POST'])
def modificar_recompensa():
    id = request.form['id']
    recompensa = controlador_recompensas.obtener_recompensa_por_id(id)
    return render_template('dashboard/modificar_recompensa.html', recompensa=recompensa)


@app.route('/actualizar_recompensa', methods=['POST'])
def actualizar_recompensa():
    id = request.form['id']
    nombre = request.form['nombre']
    descripcion = request.form['descripcion']
    puntos_requeridos = request.form['puntos_requeridos']
    tipo = request.form['tipo']

    controlador_recompensas.actualizar_recompensa(id, nombre, descripcion, puntos_requeridos, tipo)
    return redirect(url_for('obtener_todas_recompensas'))


@app.route('/otorgar_recompensa', methods=['POST'])
def otorgar_recompensa():
    id_estudiante = request.form['id_estudiante']
    id_recompensa = request.form['id_recompensa']
    controlador_recompensas.otorgar_recompensa(id_estudiante, id_recompensa)
    return redirect(url_for('obtener_todas_recompensas'))


@app.route('/revocar_recompensa', methods=['POST'])
def revocar_recompensa():
    id_estudiante = request.form['id_estudiante']
    id_recompensa = request.form['id_recompensa']
    controlador_recompensas.revocar_recompensa(id_estudiante, id_recompensa)
    return redirect(url_for('obtener_todas_recompensas'))


# ----- GESTIONAR RECOMPENSAS (POR CUESTIONARIO) -----
@app.route('/gestionar_recompensas/<int:id_cuestionario>')
def gestionar_recompensas(id_cuestionario):
    """
    Vista principal para gestionar recompensas de un cuestionario específico.
    """
    return render_template('GestionarRecompensas.html', id_cuestionario=id_cuestionario)


@app.route('/recompensas_cuestionario/<int:id_cuestionario>')
def recompensas_cuestionario_public(id_cuestionario):
    """
    Endpoint público (sin verificación de token) que devuelve las recompensas
    asociadas a un cuestionario. Se usa desde las templates/JS del dashboard
    para evitar colisiones con rutas API autenticadas.
    """
    try:
        recompensas = controlador_recompensas.obtener_recompensas_por_cuestionario(id_cuestionario)
        return jsonify({'success': True, 'recompensas': recompensas})
    except Exception as e:
        print(f"❌ Error al obtener recompensas (público) del cuestionario {id_cuestionario}: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/recompensas/<int:id_cuestionario>')
def api_recompensas_por_cuestionario(id_cuestionario):
    """
    Devuelve las recompensas asociadas a un cuestionario específico.
    """
    try:
        recompensas = controlador_recompensas.obtener_recompensas_por_cuestionario(id_cuestionario)
        return jsonify({'success': True, 'recompensas': recompensas})
    except Exception as e:
        print(f"❌ Error al obtener recompensas del cuestionario {id_cuestionario}: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/insertar_recompensa', methods=['POST'])
def insertar_recompensa_por_cuestionario():
    try:
        nombre = request.form.get('nombre')
        descripcion = request.form.get('descripcion')
        puntos = request.form.get('puntos_requeridos')
        tipo = request.form.get('tipo')
        id_cuestionario = request.form.get('id_cuestionario')

        print(f"📥 Recibiendo recompensa: nombre={nombre}, puntos={puntos}, tipo={tipo}, cuestionario={id_cuestionario}")

        if not all([nombre, puntos, tipo, id_cuestionario]):
            print(f"❌ Datos incompletos: nombre={nombre}, puntos={puntos}, tipo={tipo}, cuestionario={id_cuestionario}")
            return jsonify({'success': False, 'error': 'Datos incompletos'}), 400

        controlador_recompensas.insertar_recompensa(
            nombre, descripcion, puntos, tipo, id_cuestionario
        )
        
        print(f"✅ Recompensa insertada exitosamente: {nombre} - {tipo}")
        return jsonify({'success': True, 'message': 'Recompensa creada exitosamente'}), 200

    except Exception as e:
        print(f"❌ Error al insertar recompensa: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500



@app.route('/eliminar_recompensa/<int:id_recompensa>', methods=['DELETE'])
def eliminar_recompensa_por_cuestionario(id_recompensa):
    """Elimina una recompensa por su ID (usada desde la UI).

    Responde JSON {'success': True} o {'success': False, 'error': ...}.
    """
    try:
        controlador_recompensas.eliminar_recompensa(id_recompensa)
        return jsonify({'success': True})
    except Exception as e:
        print(f"❌ Error al eliminar recompensa {id_recompensa}: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500

#-----FIN RECOMPESA-----#
#-----INICIO-ROL-----#
@app.route('/rol')
def rol():
    roles = controlador_roles.obtener_roles()
    return render_template('dashboard/rol.html', roles=roles)

@app.route('/registrar_rol')
def registrar_rol():
    return render_template('dashboard/registrar_rol.html')

@app.route('/insertar_rol', methods=['POST'])
def insertar_rol():
    nombre_rol = request.form['nombre_rol']
    descripcion = request.form['descripcion']

    controlador_roles.crear_rol(
        nombre_rol=nombre_rol,
        descripcion=descripcion
    )
    return redirect(url_for('rol'))

@app.route('/eliminar_rol/<int:id>')
def eliminar_rol(id):
    controlador_roles.eliminar_rol(id)
    return redirect(url_for('rol'))

@app.route('/modificar_rol', methods=['POST'])
def modificar_rol():
    id = request.form['id']
    rol = controlador_roles.obtener_rol_por_id(id)
    return render_template('dashboard/modificar_rol.html', rol=rol)

@app.route('/actualizar_rol', methods=['POST'])
def actualizar_rol():
    id = request.form['id']
    nombre_rol = request.form['nombre_rol']
    descripcion = request.form['descripcion']

    controlador_roles.actualizar_rol(id, nombre_rol, descripcion)
    return redirect(url_for('rol'))

@app.route('/asignar_rol', methods=['POST'])
def asignar_rol():
    id_usuario = request.form['id_usuario']
    id_rol = request.form['id_rol']
    controlador_roles.asignar_rol_usuario(id_usuario, id_rol)
    return redirect(url_for('rol'))

@app.route('/remover_rol', methods=['POST'])
def remover_rol():
    id_usuario = request.form['id_usuario']
    id_rol = request.form['id_rol']
    controlador_roles.remover_rol_usuario(id_usuario, id_rol)
    return redirect(url_for('rol'))

@app.route('/usuarios_por_rol/<int:id_rol>')
def usuarios_por_rol(id_rol):
    usuarios = controlador_roles.obtener_usuarios_por_rol(id_rol)
    rol = controlador_roles.obtener_rol_por_id(id_rol)
    return render_template('dashboard/usuarios_por_rol.html', usuarios=usuarios, rol=rol)
#-----FIN-ROL-----#

###############################
# RUTAS API ADICIONALES
###############################

@app.route('/api/cuestionarios_por_docente/<int:id_docente>')
def api_cuestionarios_por_docente(id_docente):
    try:
        cuestionarios = controlador_cuestionarios.obtener_cuestionarios_por_docente(id_docente)
        return jsonify({'success': True, 'cuestionarios': cuestionarios})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 400

@app.route('/api/preguntas_por_cuestionario/<int:id_cuestionario>')
def api_preguntas_por_cuestionario(id_cuestionario):
    try:
        preguntas = controlador_preguntas.obtener_preguntas_por_cuestionario(id_cuestionario)
        return jsonify({'success': True, 'preguntas': preguntas})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 400

@app.route('/api/cuestionario/<int:cuestionario_id>/salas')
def obtener_salas_por_cuestionario(cuestionario_id):
    """Obtiene todas las salas creadas con un cuestionario específico"""
    try:
        conexion = obtener_conexion()
        cursor = conexion.cursor()

        cursor.execute('''
            SELECT
                s.id_sala,
                s.pin_sala,
                s.estado,
                s.modo_juego,
                s.fecha_creacion,
                s.total_preguntas,
                COUNT(DISTINCT p.id_participante) as total_participantes
            FROM salas_juego s
            LEFT JOIN participantes_sala p ON s.id_sala = p.id_sala
            WHERE s.id_cuestionario = %s
            GROUP BY s.id_sala
            ORDER BY s.fecha_creacion DESC
        ''', (cuestionario_id,))

        salas = []
        for row in cursor.fetchall():
            salas.append({
                'id_sala': row[0],
                'pin_sala': row[1],
                'estado': row[2],
                'modo_juego': row[3],
                'fecha_creacion': row[4].isoformat() if row[4] else None,
                'total_preguntas': row[5],
                'total_participantes': row[6]
            })

        cursor.close()
        conexion.close()

        return jsonify({
            'success': True,
            'salas': salas
        })
    except Exception as e:
        print(f"ERROR obtener_salas_por_cuestionario: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/jugar-individual/<int:cuestionario_id>', methods=['POST'])
@login_required
def jugar_individual(cuestionario_id):
    """Crear una sala automática individual para que el estudiante juegue solo"""
    try:
        if session.get('usuario_tipo') != 'estudiante':
            return jsonify({'success': False, 'error': 'Solo estudiantes pueden jugar en modo individual'}), 403

        import random
        import string
        import pymysql.cursors

        conexion = obtener_conexion()
        cursor = conexion.cursor(pymysql.cursors.DictCursor)

        # Verificar que el cuestionario existe y está publicado
        cursor.execute('''
            SELECT id_cuestionario, titulo
            FROM cuestionarios
            WHERE id_cuestionario = %s AND estado = 'publicado'
        ''', (cuestionario_id,))

        cuestionario = cursor.fetchone()
        if not cuestionario:
            cursor.close()
            conexion.close()
            return jsonify({'success': False, 'error': 'Cuestionario no encontrado o no está publicado'}), 404

        # Obtener preguntas del cuestionario
        cursor.execute('''
            SELECT COUNT(*) as total
            FROM cuestionario_preguntas
            WHERE id_cuestionario = %s
        ''', (cuestionario_id,))

        total_preguntas = cursor.fetchone()['total']

        if total_preguntas == 0:
            cursor.close()
            conexion.close()
            return jsonify({'success': False, 'error': 'El cuestionario no tiene preguntas'}), 400

        # Generar PIN único de 8 caracteres alfanuméricos (para modo automático/individual)
        # Formato: AUTO-XXXX donde X son letras y números
        while True:
            codigo_aleatorio = ''.join(random.choices(string.ascii_uppercase + string.digits, k=4))
            pin_sala = f'AUTO{codigo_aleatorio}'
            cursor.execute('SELECT id_sala FROM salas_juego WHERE pin_sala = %s', (pin_sala,))
            if not cursor.fetchone():
                break

        # Crear sala automática en modo individual
        cursor.execute('''
            INSERT INTO salas_juego
            (id_cuestionario, pin_sala, estado, modo_juego, total_preguntas, fecha_creacion, grupos_habilitados, num_grupos)
            VALUES (%s, %s, 'en_curso', 'individual', %s, NOW(), 0, 0)
        ''', (cuestionario_id, pin_sala, total_preguntas))

        id_sala = cursor.lastrowid

        # Crear participante automáticamente
        nombre_participante = f"{session['usuario_nombre']} {session['usuario_apellidos']}"
        cursor.execute('''
            INSERT INTO participantes_sala
            (id_sala, id_usuario, nombre_participante, estado, fecha_union)
            VALUES (%s, %s, %s, 'jugando', NOW())
        ''', (id_sala, session['usuario_id'], nombre_participante))

        id_participante = cursor.lastrowid

        # Crear entrada en ranking_sala
        cursor.execute('''
            INSERT INTO ranking_sala
            (id_participante, id_sala, puntaje_total, respuestas_correctas, tiempo_total_respuestas, posicion)
            VALUES (%s, %s, 0, 0, 0, 1)
        ''', (id_participante, id_sala))

        # Crear estado del juego (empezar en pregunta 1)
        cursor.execute('''
            INSERT INTO estado_juego_sala
            (id_sala, pregunta_actual, estado_pregunta, tiempo_inicio_pregunta)
            VALUES (%s, 1, 'mostrando', NOW())
        ''', (id_sala,))

        conexion.commit()
        cursor.close()
        conexion.close()

        # Guardar en sesión
        session['sala_actual'] = id_sala
        session['participante_id'] = id_participante

        return jsonify({
            'success': True,
            'message': 'Sala individual creada exitosamente',
            'sala_id': id_sala,
            'pin_sala': pin_sala,
            'redirect': url_for('juego_estudiante', sala_id=id_sala)
        })

    except Exception as e:
        print(f"ERROR en jugar_individual: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/cuestionarios/<int:cuestionario_id>/publicar', methods=['POST'])
@jwt_or_session_required
def api_publicar_cuestionario(cuestionario_id):
    """Publicar un cuestionario vía API - Requiere autenticación"""
    try:
        conexion = obtener_conexion()
        cursor = conexion.cursor()

        cursor.execute('''
            UPDATE cuestionarios
            SET estado = 'publicado'
            WHERE id_cuestionario = %s
        ''', (cuestionario_id,))

        conexion.commit()
        cursor.close()
        conexion.close()

        return jsonify({'success': True, 'message': 'Cuestionario publicado'})
    except Exception as e:
        print(f"ERROR api_publicar_cuestionario: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/cuestionarios/<int:cuestionario_id>/despublicar', methods=['POST'])
@jwt_or_session_required
def api_despublicar_cuestionario(cuestionario_id):
    """Despublicar un cuestionario vía API - Requiere autenticación"""
    try:
        conexion = obtener_conexion()
        cursor = conexion.cursor()

        cursor.execute('''
            UPDATE cuestionarios
            SET estado = 'borrador'
            WHERE id_cuestionario = %s
        ''', (cuestionario_id,))

        conexion.commit()
        cursor.close()
        conexion.close()

        return jsonify({'success': True, 'message': 'Cuestionario despublicado'})
    except Exception as e:
        print(f"ERROR api_despublicar_cuestionario: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/docente/<int:docente_id>/ranking-tiempo-real', methods=['GET'])
def obtener_ranking_tiempo_real_docente(docente_id):
    """Obtiene el ranking en tiempo real de todas las salas activas del docente"""
    try:
        import pymysql.cursors
        conexion = obtener_conexion()
        cursor = conexion.cursor(pymysql.cursors.DictCursor)

        # Obtener todas las salas en curso o esperando del docente
        query = '''
            SELECT
                s.id_sala,
                s.pin_sala,
                s.estado,
                c.titulo as cuestionario_titulo,
                c.id_cuestionario
            FROM salas_juego s
            INNER JOIN cuestionarios c ON s.id_cuestionario = c.id_cuestionario
            WHERE c.id_docente = %s
            AND s.estado IN ('esperando', 'en_curso')
            ORDER BY s.fecha_creacion DESC
        '''

        cursor.execute(query, (docente_id,))
        salas = cursor.fetchall()

        # Para cada sala, obtener el ranking actual
        resultados = []
        for sala in salas:
            # Obtener el top 3 del ranking de cada sala
            ranking_query = '''
                SELECT
                    r.id_participante,
                    p.nombre_participante,
                    g.nombre_grupo,
                    r.puntaje_total,
                    r.respuestas_correctas,
                    r.tiempo_total_respuestas,
                    r.posicion,
                    p.id_usuario
                FROM ranking_sala r
                INNER JOIN participantes_sala p ON r.id_participante = p.id_participante
                LEFT JOIN grupos_sala g ON p.id_grupo = g.id_grupo
                WHERE r.id_sala = %s
                ORDER BY r.posicion ASC
                LIMIT 3
            '''

            cursor.execute(ranking_query, (sala['id_sala'],))
            ranking = cursor.fetchall()

            if ranking:  # Solo incluir salas con participantes
                resultados.append({
                    'sala_id': sala['id_sala'],
                    'pin_sala': sala['pin_sala'],
                    'estado': sala['estado'],
                    'cuestionario_titulo': sala['cuestionario_titulo'],
                    'cuestionario_id': sala['id_cuestionario'],
                    'ranking': ranking
                })

        cursor.close()
        conexion.close()

        return jsonify({
            'success': True,
            'salas': resultados,
            'total_salas_activas': len(resultados)
        })

    except Exception as e:
        print(f"ERROR obtener_ranking_tiempo_real_docente: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/opciones_por_pregunta/<int:id_pregunta>')
def api_opciones_por_pregunta(id_pregunta):
    try:
        opciones = controlador_opciones.obtener_opciones_por_pregunta(id_pregunta)
        return jsonify({'success': True, 'opciones': opciones})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 400

@app.route('/api/estadisticas_sistema')
def api_estadisticas_sistema():
    try:
        stats = {
            'total_usuarios': len(controlador_usuario.obtener_usuarios()),
            'total_cuestionarios': len(controlador_cuestionarios.obtener_cuestionarios()),
            'total_preguntas': len(controlador_preguntas.obtener_preguntas()),
            'total_salas': len(controlador_salas.obtener_salas()),
            'total_participaciones': len(controlador_participaciones.obtener_participaciones())
        }
        return jsonify({'success': True, 'estadisticas': stats})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 400

@app.route('/api/validar_email/<email>')
def api_validar_email(email):
    try:
        usuario = controlador_usuario.obtener_usuario_por_email(email)
        existe = usuario is not None
        return jsonify({'success': True, 'existe': existe})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 400

@app.route('/api/buscar_usuarios/<termino>')
def api_buscar_usuarios(termino):
    try:
        usuarios = controlador_usuario.buscar_usuarios(termino)
        return jsonify({'success': True, 'usuarios': usuarios})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 400

@app.route('/api/recompensas_disponibles/<int:id_estudiante>')
def api_recompensas_disponibles(id_estudiante):
    try:
        recompensas = controlador_recompensas.verificar_recompensas_disponibles(id_estudiante)
        return jsonify({'success': True, 'recompensas': recompensas})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 400

@app.route('/api/otorgar_recompensas_automaticas/<int:id_estudiante>', methods=['POST'])
@jwt_or_session_required
def api_otorgar_recompensas_automaticas(id_estudiante):
    """Otorgar recompensas automáticas - Requiere autenticación"""
    try:
        recompensas_otorgadas = controlador_recompensas.otorgar_recompensas_automaticas(id_estudiante)
        return jsonify({'success': True, 'recompensas_otorgadas': recompensas_otorgadas})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 400

# APIs para manejo de salas en tiempo real
@app.route('/api/sala/<int:sala_id>/usuarios', methods=['GET'])
def api_obtener_usuarios_sala(sala_id):
    """Obtener usuarios conectados a una sala"""
    try:
        # Verificar que la sala existe
        sala = controlador_salas.obtener_sala_por_id(sala_id)
        if not sala:
            return jsonify({'success': False, 'error': 'Sala no encontrada'}), 404

        # Obtener usuarios de la sala (simulado por ahora)
        # En una implementación real, esto vendría de WebSockets o base de datos
        usuarios = controlador_salas.obtener_participantes_sala(sala_id)

        usuarios_formateados = []
        if usuarios:
            for usuario in usuarios:
                usuarios_formateados.append({
                    'id': usuario[0],
                    'nombre': usuario[1],
                    'estado': 'conectado',
                    'puntuacion': usuario[3] if len(usuario) > 3 else 0
                })

        return jsonify({
            'success': True,
            'usuarios': usuarios_formateados
        })

    except Exception as e:
        print(f"DEBUG: Error en api_obtener_usuarios_sala: {str(e)}")
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/sala/<int:sala_id>/iniciar', methods=['POST'])
@jwt_or_session_required
def api_iniciar_juego_sala(sala_id):
    """Iniciar el juego en una sala - Requiere autenticación"""
    try:
        # Verificar que la sala existe
        sala = controlador_salas.obtener_sala_por_id(sala_id)
        if not sala:
            return jsonify({'success': False, 'error': 'Sala no encontrada'}), 404

        # Verificar que hay usuarios conectados
        usuarios = controlador_salas.obtener_participantes_sala(sala_id)
        if not usuarios or len(usuarios) == 0:
            return jsonify({'success': False, 'error': 'No hay estudiantes conectados'}), 400

        # Iniciar juego usando el controlador (esto establece estado 'en_curso')
        success = controlador_juego.iniciar_juego_sala(sala_id)

        if success:
            return jsonify({
                'success': True,
                'message': 'Juego iniciado exitosamente'
            })
        else:
            return jsonify({'success': False, 'error': 'Error al iniciar el juego'}), 500

    except Exception as e:
        print(f"DEBUG: Error en api_iniciar_juego_sala: {str(e)}")
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/sala/<int:sala_id>/cerrar', methods=['POST'])
@jwt_or_session_required
def api_cerrar_sala(sala_id):
    """Cerrar una sala - Requiere autenticación"""
    try:
        # Verificar que la sala existe
        sala = controlador_salas.obtener_sala_por_id(sala_id)
        if not sala:
            return jsonify({'success': False, 'error': 'Sala no encontrada'}), 404

        # Cambiar estado de la sala a 'finalizada'
        success = controlador_salas.actualizar_estado_sala(sala_id, 'finalizada')

        if success:
            return jsonify({
                'success': True,
                'message': 'Sala cerrada exitosamente'
            })
        else:
            return jsonify({'success': False, 'error': 'Error al cerrar la sala'}), 500

    except Exception as e:
        print(f"DEBUG: Error en api_cerrar_sala: {str(e)}")
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/exportar-resultados/<int:sala_id>/excel', methods=['POST'])
@jwt_or_session_required
def exportar_resultados_excel(sala_id):
    """Exportar resultados a Excel - Requiere autenticación"""
    try:
        # Verificar que openpyxl está instalado
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        except ImportError:
            return jsonify({'success': False, 'error': 'Librería openpyxl no instalada. Ejecuta: pip install openpyxl'}), 500

        data = request.get_json()
        ranking = data.get('ranking', [])
        total_preguntas = data.get('totalPreguntas', 0)
        cuestionario_titulo = data.get('cuestionarioTitulo', 'Cuestionario')

        # Crear workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Resultados Brain RUSH"

        # Estilos
        header_fill = PatternFill(start_color="667eea", end_color="667eea", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=12)
        title_font = Font(bold=True, size=16, color="667eea")
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # Título
        ws.merge_cells('A1:G1')
        title_cell = ws['A1']
        title_cell.value = f"📊 Resultados - {cuestionario_titulo}"
        title_cell.font = title_font
        title_cell.alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[1].height = 30

        # Fecha
        ws.merge_cells('A2:G2')
        fecha_cell = ws['A2']
        fecha_cell.value = f"Fecha: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
        fecha_cell.alignment = Alignment(horizontal='center')
        ws.row_dimensions[2].height = 20

        # Encabezados
        headers = ['Posición', 'Nombre', 'Resp. Correctas', 'Total Preguntas', 'Precisión', 'Puntuación', 'Tiempo (s)']
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=4, column=col)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border

        ws.row_dimensions[4].height = 25

        # Datos
        for idx, player in enumerate(ranking, start=5):
            precision = round((player['respuestas_correctas'] / total_preguntas) * 100) if total_preguntas > 0 else 0

            # Soportar ambos formatos: nombre_participante (frontend) o nombre_completo (backend)
            nombre = player.get('nombre_participante') or player.get('nombre_completo', 'Jugador')
            puntaje = player.get('puntaje_total') or player.get('puntos_totales', 0)

            row_data = [
                player['posicion'],
                nombre,
                player['respuestas_correctas'],
                total_preguntas,
                f"{precision}%",
                puntaje,
                round(player['tiempo_total_respuestas'], 2)
            ]

            for col, value in enumerate(row_data, start=1):
                cell = ws.cell(row=idx, column=col)
                cell.value = value
                cell.border = border
                cell.alignment = Alignment(horizontal='center', vertical='center')

                # Colorear top 3
                if player['posicion'] <= 3:
                    colors = {1: "FFD700", 2: "C0C0C0", 3: "CD7F32"}
                    cell.fill = PatternFill(start_color=colors[player['posicion']],
                                          end_color=colors[player['posicion']],
                                          fill_type="solid")

        # Ajustar anchos de columna
        column_widths = [12, 25, 18, 18, 15, 15, 15]
        for col, width in enumerate(column_widths, start=1):
            ws.column_dimensions[chr(64 + col)].width = width

        # Guardar en memoria
        excel_file = BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)

        return send_file(
            excel_file,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=f'resultados_brain_rush_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        )

    except Exception as e:
        print(f"ERROR exportar_excel: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/exportar-resultados/<int:sala_id>/onedrive', methods=['POST'])
@jwt_or_session_required
def exportar_resultados_onedrive(sala_id):
    """Exportar resultados a OneDrive usando sistema centralizado y enviar email con link - Requiere autenticación"""
    try:
        # Verificar autenticación
        if 'usuario_id' not in session:
            return jsonify({
                'success': False,
                'error': 'Usuario no autenticado'
            }), 401

        # Obtener datos de la solicitud
        data = request.get_json()
        ranking = data.get('ranking', [])
        total_preguntas = data.get('totalPreguntas', 0)
        cuestionario_titulo = data.get('cuestionarioTitulo', 'Cuestionario')

        # Obtener datos del usuario (quien solicita la exportación)
        conexion = obtener_conexion()
        cursor = conexion.cursor()
        cursor.execute("""
            SELECT email, nombre, apellidos
            FROM usuarios WHERE id_usuario = %s
        """, (session['usuario_id'],))
        usuario = cursor.fetchone()
        cursor.close()
        conexion.close()

        if not usuario:
            return jsonify({
                'success': False,
                'error': 'Usuario no encontrado'
            }), 404

        email_usuario = usuario[0]
        nombre_completo = f"{usuario[1]} {usuario[2]}"

        # Crear archivo Excel en memoria
        import tempfile
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill

        wb = Workbook()
        ws = wb.active
        ws.title = "Resultados"

        # Título del cuestionario
        ws.merge_cells('A1:H1')
        titulo_cell = ws['A1']
        titulo_cell.value = f"Resultados - {cuestionario_titulo}"
        titulo_cell.font = Font(size=16, bold=True, color="FFFFFF")
        titulo_cell.fill = PatternFill(start_color="667eea", end_color="667eea", fill_type="solid")
        titulo_cell.alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[1].height = 30

        # Encabezados
        headers = ['Posición', 'Estudiante', 'Grupo', 'Puntaje', 'Correctas', 'Incorrectas',
                   'Precisión (%)', 'Tiempo Total']
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=2, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')

        # Datos del ranking
        for row_idx, participante in enumerate(ranking, 3):
            ws.cell(row=row_idx, column=1, value=participante.get('posicion', row_idx-2))
            # Soportar ambos formatos de nombre
            nombre = participante.get('nombre_participante') or participante.get('nombre_completo', 'N/A')
            ws.cell(row=row_idx, column=2, value=nombre)
            ws.cell(row=row_idx, column=3, value=participante.get('nombre_grupo', 'Sin grupo'))
            # Soportar ambos formatos de puntaje
            puntaje = participante.get('puntaje_total') or participante.get('puntos_totales', 0)
            ws.cell(row=row_idx, column=4, value=puntaje)
            ws.cell(row=row_idx, column=5, value=participante.get('respuestas_correctas', 0))

            incorrectas = total_preguntas - participante.get('respuestas_correctas', 0)
            ws.cell(row=row_idx, column=6, value=incorrectas)

            if total_preguntas > 0:
                precision = (participante.get('respuestas_correctas', 0) / total_preguntas) * 100
                ws.cell(row=row_idx, column=7, value=f"{precision:.1f}%")
            else:
                ws.cell(row=row_idx, column=7, value="0%")

            tiempo = participante.get('tiempo_total_respuestas', 0)
            if tiempo:
                minutos = int(tiempo) // 60
                segundos = int(tiempo) % 60
                ws.cell(row=row_idx, column=8, value=f"{minutos}m {segundos}s")
            else:
                ws.cell(row=row_idx, column=8, value="0m 0s")

        # Ajustar anchos de columna
        column_widths = [12, 25, 15, 10, 12, 12, 15, 15]
        for i, width in enumerate(column_widths, 1):
            ws.column_dimensions[chr(64 + i)].width = width

        # Guardar en archivo temporal para PythonAnywhere
        with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp_file:
            wb.save(tmp_file.name)
            tmp_file.seek(0)
            with open(tmp_file.name, 'rb') as f:
                excel_data = f.read()

        # Calcular estadísticas
        total_participantes = len(ranking)
        puntaje_maximo = max([p.get('puntaje_total', 0) for p in ranking]) if ranking else 0
        promedio_correctas = sum([p.get('respuestas_correctas', 0) for p in ranking]) / total_participantes if total_participantes > 0 else 0

        # Nombre del archivo
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"BrainRush_Resultados_{cuestionario_titulo.replace(' ', '_')}_{timestamp}.xlsx"

        # Subir a OneDrive usando el sistema centralizado
        print(f"📤 Subiendo resultados a OneDrive: {filename}")
        resultado_onedrive = subir_archivo_onedrive_sistema(filename, excel_data, carpeta="BrainRush")

        if not resultado_onedrive['success']:
            print(f"❌ Error al subir a OneDrive: {resultado_onedrive.get('error')}")
            return jsonify({
                'success': False,
                'error': f"Error al subir a OneDrive: {resultado_onedrive.get('error')}"
            }), 500

        share_link = resultado_onedrive['share_link']
        file_name = resultado_onedrive['file_name']

        print(f"✅ Archivo subido exitosamente a OneDrive")
        print(f"🔗 Link compartido: {share_link}")

        # SIEMPRE enviar email con el link
        email_sent = False
        if app.config.get('MAIL_ENABLED') and app.config.get('MAIL_USERNAME'):
            from flask_mail import Message
            from extensions import mail

            try:
                mensaje_html = f"""
                <html>
                <body style="font-family: Arial, sans-serif; max-width: 600px; margin: 0 auto; padding: 20px;">
                    <div style="background: linear-gradient(135deg, #667eea 0%, #764ba2 100%); padding: 30px; border-radius: 10px 10px 0 0; text-align: center;">
                        <h1 style="color: white; margin: 0;">🎓 Brain RUSH</h1>
                    </div>

                    <div style="background: #f9f9f9; padding: 30px; border-radius: 0 0 10px 10px;">
                        <h2 style="color: #333;">¡Los Resultados están listos! 📊</h2>

                        <p>Hola <strong>{nombre_completo}</strong>,</p>

                        <p>Los resultados del cuestionario "<strong>{cuestionario_titulo}</strong>" han sido exportados exitosamente a OneDrive.</p>

                        <div style="background: white; padding: 20px; border-radius: 8px; margin: 20px 0; border-left: 4px solid #667eea;">
                            <p style="margin: 0 0 10px 0; color: #666;">📁 Archivo:</p>
                            <p style="margin: 0; font-size: 16px; color: #333;"><strong>{file_name}</strong></p>
                        </div>

                        <div style="background: white; padding: 20px; border-radius: 8px; margin: 20px 0;">
                            <p style="margin: 0 0 15px 0; font-size: 14px; color: #666; text-align: center;">📊 <strong>Resumen del Cuestionario</strong></p>
                            <div style="display: flex; justify-content: space-around; text-align: center;">
                                <div>
                                    <p style="margin: 0; font-size: 24px; color: #667eea; font-weight: bold;">{total_participantes}</p>
                                    <p style="margin: 5px 0 0 0; font-size: 12px; color: #666;">Participantes</p>
                                </div>
                                <div>
                                    <p style="margin: 0; font-size: 24px; color: #667eea; font-weight: bold;">{puntaje_maximo}</p>
                                    <p style="margin: 5px 0 0 0; font-size: 12px; color: #666;">Puntaje Máximo</p>
                                </div>
                                <div>
                                    <p style="margin: 0; font-size: 24px; color: #667eea; font-weight: bold;">{promedio_correctas:.1f}</p>
                                    <p style="margin: 5px 0 0 0; font-size: 12px; color: #666;">Promedio Correctas</p>
                                </div>
                            </div>
                        </div>

                        <div style="text-align: center; margin: 30px 0;">
                            <a href="{share_link}"
                               style="display: inline-block; background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
                                      color: white; padding: 15px 40px; text-decoration: none; border-radius: 25px;
                                      font-weight: bold; font-size: 16px;">
                                📂 Ver Resultados en OneDrive
                            </a>
                        </div>

                        <p style="color: #666; font-size: 14px; margin-top: 30px;">
                            💡 <em>Puedes acceder a este archivo en cualquier momento desde tu carpeta "BrainRush" en OneDrive.</em>
                        </p>
                    </div>

                    <div style="text-align: center; margin-top: 20px; color: #999; font-size: 12px;">
                        <p>Brain RUSH - Sistema de Evaluación Interactiva</p>
                    </div>
                </body>
                </html>
                """

                msg = Message(
                    subject=f"📊 Resultados de {cuestionario_titulo} - Brain RUSH",
                    recipients=[email_usuario],
                    html=mensaje_html,
                    sender=app.config['MAIL_DEFAULT_SENDER']
                )

                mail.send(msg)
                email_sent = True
                print(f"✅ Email enviado exitosamente a {email_usuario}")

            except Exception as email_error:
                print(f"❌ Error enviando email: {email_error}")
                import traceback
                traceback.print_exc()

        # Retornar respuesta con toda la información
        return jsonify({
            'success': True,
            'message': f'Archivo exportado exitosamente a OneDrive',
            'file_name': file_name,
            'share_link': share_link,
            'email_sent': email_sent,
            'estadisticas': {
                'total_participantes': total_participantes,
                'puntaje_maximo': puntaje_maximo,
                'promedio_correctas': round(promedio_correctas, 1)
            }
        }), 200

    except Exception as e:
        print(f"❌ ERROR exportar_resultados_onedrive: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 500


def obtener_token_sistema_onedrive():
    """
    Obtener token de acceso del sistema OneDrive.
    Primero intenta usar el token del .env, si está expirado lo refresca.
    """
    try:
        # Obtener tokens del .env
        access_token = app.config.get('ONEDRIVE_ACCESS_TOKEN')
        refresh_token = app.config.get('ONEDRIVE_REFRESH_TOKEN')
        token_expires_str = app.config.get('ONEDRIVE_TOKEN_EXPIRES')

        # Convertir fecha de expiración si existe
        token_expires = None
        if token_expires_str:
            try:
                from datetime import datetime
                token_expires = datetime.fromisoformat(token_expires_str)
            except:
                pass

        # Si hay token y no ha expirado, usarlo
        if access_token and token_expires and datetime.now() < token_expires:
            print("✅ Usando token de OneDrive del .env (válido)")
            return access_token

        # Si hay refresh_token, intentar refrescar
        if refresh_token and MSAL_AVAILABLE:
            print("🔄 Token expirado, refrescando...")
            try:
                authority = f"https://login.microsoftonline.com/{app.config['AZURE_TENANT_ID']}"

                msal_app = msal.ConfidentialClientApplication(
                    app.config['AZURE_CLIENT_ID'],
                    authority=authority,
                    client_credential=app.config['AZURE_CLIENT_SECRET']
                )

                result = msal_app.acquire_token_by_refresh_token(
                    refresh_token,
                    scopes=app.config.get('ONEDRIVE_SCOPES', [
                        'Files.ReadWrite.All',
                        'offline_access'
                    ])
                )

                if "access_token" in result:
                    new_access_token = result['access_token']
                    new_refresh_token = result.get('refresh_token', refresh_token)
                    expires_in = result.get('expires_in', 3600)
                    new_token_expires = datetime.now() + timedelta(seconds=expires_in)

                    # Actualizar tokens en memoria
                    app.config['ONEDRIVE_ACCESS_TOKEN'] = new_access_token
                    app.config['ONEDRIVE_REFRESH_TOKEN'] = new_refresh_token
                    app.config['ONEDRIVE_TOKEN_EXPIRES'] = new_token_expires.isoformat()

                    # Actualizar archivo .env
                    actualizar_env_tokens(new_access_token, new_refresh_token, new_token_expires)

                    print("✅ Token refrescado exitosamente")
                    return new_access_token
                else:
                    error_desc = result.get("error_description", "Unknown error")
                    print(f"❌ Error refrescando token: {error_desc}")

            except Exception as e:
                print(f"⚠️ Error al refrescar token: {e}")

        # Si no hay tokens o falla el refresh, intentar Client Credentials Flow
        # (solo funciona con Application Permissions)
        if MSAL_AVAILABLE:
            print("🔄 Intentando obtener token con Client Credentials...")
            authority = f"https://login.microsoftonline.com/{app.config['AZURE_TENANT_ID']}"

            msal_app = msal.ConfidentialClientApplication(
                app.config['AZURE_CLIENT_ID'],
                authority=authority,
                client_credential=app.config['AZURE_CLIENT_SECRET']
            )

            result = msal_app.acquire_token_for_client(
                scopes=["https://graph.microsoft.com/.default"]
            )

            if "access_token" in result:
                print("✅ Token obtenido con Client Credentials")
                return result['access_token']
            else:
                error_desc = result.get("error_description", "Unknown error")
                print(f"❌ Error con Client Credentials: {error_desc}")

        print("❌ No se pudo obtener token de OneDrive")
        print("💡 Necesitas configurar ONEDRIVE_ACCESS_TOKEN y ONEDRIVE_REFRESH_TOKEN en .env")
        print("💡 O configurar Application Permissions en Azure AD")
        return None

    except Exception as e:
        print(f"ERROR en obtener_token_sistema_onedrive: {e}")
        import traceback
        traceback.print_exc()
        return None


def actualizar_env_tokens(access_token, refresh_token, token_expires):
    """Actualizar tokens en el archivo .env"""
    try:
        import os
        env_path = os.path.join(os.path.dirname(__file__), '.env')

        # Leer archivo .env
        with open(env_path, 'r', encoding='utf-8') as f:
            lines = f.readlines()

        # Actualizar líneas con nuevos tokens
        new_lines = []
        for line in lines:
            if line.startswith('ONEDRIVE_ACCESS_TOKEN='):
                new_lines.append(f'ONEDRIVE_ACCESS_TOKEN={access_token}\n')
            elif line.startswith('ONEDRIVE_REFRESH_TOKEN='):
                new_lines.append(f'ONEDRIVE_REFRESH_TOKEN={refresh_token}\n')
            elif line.startswith('ONEDRIVE_TOKEN_EXPIRES='):
                new_lines.append(f'ONEDRIVE_TOKEN_EXPIRES={token_expires.isoformat()}\n')
            else:
                new_lines.append(line)

        # Escribir archivo actualizado
        with open(env_path, 'w', encoding='utf-8') as f:
            f.writelines(new_lines)

        print("✅ Tokens actualizados en .env")

    except Exception as e:
        print(f"⚠️ No se pudo actualizar .env: {e}")
        print("   Los tokens están en memoria pero no se guardaron en disco")


def subir_archivo_onedrive_sistema(filename, file_data, carpeta="BrainRush"):
    """
    Subir archivo a OneDrive del sistema y crear link de compartición

    Args:
        filename: Nombre del archivo
        file_data: Datos binarios del archivo
        carpeta: Carpeta en OneDrive (default: BrainRush)

    Returns:
        dict: {'success': bool, 'web_url': str, 'share_link': str, 'error': str}
    """
    try:
        # Obtener token del sistema
        access_token = obtener_token_sistema_onedrive()

        if not access_token:
            return {
                'success': False,
                'error': 'No se pudo obtener token de acceso a OneDrive'
            }

        # 1. Verificar/crear carpeta BrainRush
        folder_url = f"https://graph.microsoft.com/v1.0/me/drive/root:/{carpeta}"
        headers = {'Authorization': f'Bearer {access_token}'}

        folder_response = requests.get(folder_url, headers=headers, timeout=10)

        if folder_response.status_code == 404:
            # Crear carpeta si no existe
            create_folder_url = "https://graph.microsoft.com/v1.0/me/drive/root/children"
            folder_data = {
                "name": carpeta,
                "folder": {},
                "@microsoft.graph.conflictBehavior": "rename"
            }
            requests.post(create_folder_url, headers={**headers, 'Content-Type': 'application/json'},
                         json=folder_data, timeout=10)
            print(f"✅ Carpeta '{carpeta}' creada en OneDrive")

        # 2. Subir archivo
        upload_url = f"https://graph.microsoft.com/v1.0/me/drive/root:/{carpeta}/{filename}:/content"
        upload_headers = {
            'Authorization': f'Bearer {access_token}',
            'Content-Type': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        }

        print(f"📤 Subiendo archivo a OneDrive: {filename}")
        upload_response = requests.put(upload_url, headers=upload_headers, data=file_data, timeout=30)

        if upload_response.status_code not in [200, 201]:
            print(f"❌ Error subiendo archivo: {upload_response.status_code} - {upload_response.text}")
            return {
                'success': False,
                'error': f'Error subiendo archivo: {upload_response.status_code}'
            }

        file_info = upload_response.json()
        file_id = file_info.get('id')
        web_url = file_info.get('webUrl')

        print(f"✅ Archivo subido exitosamente. ID: {file_id}")

        # 3. Crear link de compartición público
        share_url = f"https://graph.microsoft.com/v1.0/me/drive/items/{file_id}/createLink"
        share_data = {
            "type": "view",  # view = solo lectura, edit = edición
            "scope": "anonymous"  # anonymous = cualquiera con el link
        }
        share_headers = {
            'Authorization': f'Bearer {access_token}',
            'Content-Type': 'application/json'
        }

        share_response = requests.post(share_url, headers=share_headers, json=share_data, timeout=10)

        if share_response.status_code in [200, 201]:
            share_info = share_response.json()
            share_link = share_info.get('link', {}).get('webUrl')
            print(f"✅ Link de compartición creado: {share_link}")

            return {
                'success': True,
                'web_url': web_url,
                'share_link': share_link,
                'file_id': file_id,
                'file_name': filename
            }
        else:
            # Si falla la compartición, al menos devolver la URL del archivo
            print(f"⚠️ No se pudo crear link de compartición, pero archivo subido")
            return {
                'success': True,
                'web_url': web_url,
                'share_link': web_url,  # Usar web_url como fallback
                'file_id': file_id,
                'file_name': filename,
                'warning': 'Link de compartición no creado'
            }

    except Exception as e:
        print(f"ERROR en subir_archivo_onedrive_sistema: {e}")
        import traceback
        traceback.print_exc()
        return {
            'success': False,
            'error': str(e)
        }


def enviar_por_email_fallback(email_usuario, nombre_completo, cuestionario_titulo,
                               ranking, total_preguntas, filename, excel_data):
    """Función auxiliar para enviar por email cuando OneDrive falla"""
    try:
        from flask_mail import Message
        from extensions import mail

        if not app.config.get('MAIL_ENABLED') or not app.config.get('MAIL_USERNAME'):
            return jsonify({
                'success': False,
                'error': 'No se pudo subir a OneDrive y el correo no está configurado'
            }), 500

        mensaje_texto = f"""
Hola {nombre_completo},

No se pudo subir directamente a OneDrive, pero adjunto encontrarás los resultados de "{cuestionario_titulo}".

📊 Detalles:
- Total de participantes: {len(ranking)}
- Total de preguntas: {total_preguntas}
- Fecha de generación: {datetime.now().strftime("%d/%m/%Y %H:%M")}

💡 Puedes guardar este archivo en tu OneDrive, Google Drive o donde prefieras.

Saludos,
Sistema BrainRush
        """

        msg = Message(
            subject=f'Resultados de BrainRush - {cuestionario_titulo}',
            sender=app.config.get('MAIL_DEFAULT_SENDER', app.config.get('MAIL_USERNAME')),
            recipients=[email_usuario],
            body=mensaje_texto
        )

        msg.attach(
            filename,
            'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            excel_data
        )

        mail.send(msg)

        return jsonify({
            'success': True,
            'message': f'No se pudo subir a OneDrive. Resultados enviados por correo a {email_usuario}',
            'file_name': filename,
            'email_sent': True,
            'fallback': True
        }), 200

    except Exception as e:
        print(f"Error enviando email fallback: {e}")
        return jsonify({
            'success': False,
            'error': f'Error subiendo a OneDrive y enviando email: {str(e)}'
        }), 500

@app.route('/api/exportar-historial-estudiante/onedrive', methods=['POST'])
@login_required
@jwt_or_session_required
def exportar_historial_estudiante_onedrive():
    """Exportar historial completo del estudiante a OneDrive del sistema - Requiere autenticación"""
    try:
        usuario_id = session.get('usuario_id')

        # Obtener datos del usuario
        conexion = obtener_conexion()
        cursor = conexion.cursor()
        cursor.execute("""
            SELECT email, nombre, apellidos
            FROM usuarios WHERE id_usuario = %s
        """, (usuario_id,))
        usuario = cursor.fetchone()

        if not usuario:
            cursor.close()
            conexion.close()
            return jsonify({'success': False, 'error': 'Usuario no encontrado'}), 404

        email_usuario = usuario[0]
        nombre_completo = f"{usuario[1]} {usuario[2]}"

        # Obtener historial de participaciones del estudiante
        cursor.execute("""
            SELECT
                c.titulo as titulo_cuestionario,
                s.fecha_inicio,
                rs.puntaje_total,
                rs.respuestas_correctas,
                rp.total_preguntas,
                rs.posicion,
                rs.tiempo_total_respuestas,
                CONCAT(u_docente.nombre, ' ', u_docente.apellidos) as docente
            FROM participantes_sala ps
            INNER JOIN salas_juego s ON ps.id_sala = s.id
            INNER JOIN cuestionarios c ON s.id_cuestionario = c.id_cuestionario
            INNER JOIN ranking_sala rs ON ps.id_participante = rs.id_participante
            LEFT JOIN usuarios u_docente ON c.id_docente = u_docente.id_usuario
            LEFT JOIN (
                SELECT id_sala, COUNT(DISTINCT id_pregunta) as total_preguntas
                FROM respuestas_participantes
                GROUP BY id_sala
            ) rp ON s.id = rp.id_sala
            WHERE ps.id_usuario = %s
                AND s.estado = 'finalizado'
            ORDER BY s.fecha_inicio DESC
        """, (usuario_id,))

        participaciones = cursor.fetchall()
        cursor.close()
        conexion.close()

        if not participaciones:
            return jsonify({
                'success': False,
                'error': 'No tienes participaciones en tu historial'
            }), 404

        # Crear archivo Excel con el historial
        import io
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill

        wb = Workbook()
        ws = wb.active
        ws.title = "Mi Historial"

        # Título del documento
        ws.merge_cells('A1:H1')
        titulo_cell = ws['A1']
        titulo_cell.value = f"Historial de Participaciones - {nombre_completo}"
        titulo_cell.font = Font(size=16, bold=True, color="FFFFFF")
        titulo_cell.fill = PatternFill(start_color="667eea", end_color="667eea", fill_type="solid")
        titulo_cell.alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[1].height = 30

        # Encabezados
        headers = ['Cuestionario', 'Fecha', 'Puntaje', 'Correctas', 'Total Preguntas',
                   'Precisión (%)', 'Posición', 'Tiempo', 'Docente']
        header_fill = PatternFill(start_color="4ECDC4", end_color="4ECDC4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=2, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')

        # Datos
        for row_idx, participacion in enumerate(participaciones, 3):
            titulo_cuest = participacion[0]
            fecha_inicio = participacion[1]
            puntaje = participacion[2] or 0
            correctas = participacion[3] or 0
            total_pregs = participacion[4] or 10
            posicion = participacion[5] or 'N/A'
            tiempo = participacion[6] or 0
            docente = participacion[7] or 'N/A'

            ws.cell(row=row_idx, column=1, value=titulo_cuest)
            ws.cell(row=row_idx, column=2, value=fecha_inicio.strftime("%d/%m/%Y %H:%M") if fecha_inicio else 'N/A')
            ws.cell(row=row_idx, column=3, value=puntaje)
            ws.cell(row=row_idx, column=4, value=correctas)
            ws.cell(row=row_idx, column=5, value=total_pregs)

            # Calcular precisión
            if total_pregs > 0:
                precision = (correctas / total_pregs) * 100
                ws.cell(row=row_idx, column=6, value=f"{precision:.1f}%")
            else:
                ws.cell(row=row_idx, column=6, value="0%")

            ws.cell(row=row_idx, column=7, value=str(posicion))

            # Formatear tiempo
            minutos = tiempo // 60
            segundos = tiempo % 60
            ws.cell(row=row_idx, column=8, value=f"{minutos}m {segundos}s")

            ws.cell(row=row_idx, column=9, value=docente)

        # Agregar estadísticas al final
        ultima_fila = len(participaciones) + 4
        ws.merge_cells(f'A{ultima_fila}:I{ultima_fila}')
        stats_cell = ws[f'A{ultima_fila}']
        stats_cell.value = "📊 Estadísticas Generales"
        stats_cell.font = Font(size=14, bold=True, color="FFFFFF")
        stats_cell.fill = PatternFill(start_color="667eea", end_color="667eea", fill_type="solid")
        stats_cell.alignment = Alignment(horizontal='center', vertical='center')

        # Calcular estadísticas
        total_juegos = len(participaciones)
        total_puntaje = sum(p[2] or 0 for p in participaciones)
        promedio_puntaje = total_puntaje / total_juegos if total_juegos > 0 else 0
        total_correctas = sum(p[3] or 0 for p in participaciones)

        stats_data = [
            ['Total de Juegos:', total_juegos],
            ['Puntos Totales:', total_puntaje],
            ['Promedio por Juego:', f"{promedio_puntaje:.1f}"],
            ['Respuestas Correctas:', total_correctas]
        ]

        for i, (label, value) in enumerate(stats_data):
            fila = ultima_fila + 1 + i
            ws.cell(row=fila, column=1, value=label).font = Font(bold=True)
            ws.cell(row=fila, column=2, value=value)

        # Ajustar anchos de columna
        column_widths = [30, 18, 10, 10, 15, 12, 10, 12, 25]
        for i, width in enumerate(column_widths, 1):
            ws.column_dimensions[chr(64 + i)].width = width

        # Guardar en memoria
        excel_buffer = io.BytesIO()
        wb.save(excel_buffer)
        excel_buffer.seek(0)
        excel_data = excel_buffer.getvalue()

        # Nombre del archivo
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"BrainRush_MiHistorial_{nombre_completo.replace(' ', '_')}_{timestamp}.xlsx"

        # Usar sistema centralizado de OneDrive
        print(f"📤 Intentando subir historial a OneDrive del sistema para {nombre_completo}")

        resultado_onedrive = subir_archivo_onedrive_sistema(filename, excel_data, carpeta="BrainRush")

        if resultado_onedrive['success']:
            # Archivo subido exitosamente, enviar email con el link
            share_link = resultado_onedrive.get('share_link')
            file_name = resultado_onedrive.get('file_name')

            try:
                from flask_mail import Message
                from extensions import mail

                if app.config.get('MAIL_ENABLED'):
                    mensaje_html = f"""
                    <html>
                    <body style="font-family: Arial, sans-serif; max-width: 600px; margin: 0 auto; padding: 20px;">
                        <div style="background: linear-gradient(135deg, #667eea 0%, #764ba2 100%); padding: 30px; text-align: center; border-radius: 10px 10px 0 0;">
                            <h1 style="color: white; margin: 0; font-size: 2rem;">🎓 Brain RUSH</h1>
                        </div>

                        <div style="background: #f9f9f9; padding: 30px; border-radius: 0 0 10px 10px;">
                            <h2 style="color: #333; margin-top: 0;">¡Tu Historial está listo! 📊</h2>

                            <p style="color: #555; font-size: 1.1rem;">Hola <strong>{nombre_completo}</strong>,</p>

                            <p style="color: #666;">Tu historial de participaciones ha sido exportado exitosamente a OneDrive.</p>

                            <div style="background: white; padding: 20px; border-radius: 8px; margin: 20px 0; border-left: 4px solid #667eea;">
                                <p style="margin: 0 0 10px 0; color: #666;">📁 <strong>Archivo:</strong></p>
                                <p style="margin: 0; font-size: 14px; color: #333;">{file_name}</p>
                            </div>

                            <div style="background: #e8f4f8; padding: 15px; border-radius: 8px; margin: 20px 0;">
                                <p style="margin: 0; color: #555; font-size: 0.95rem;">
                                    <strong>📊 Estadísticas:</strong><br>
                                    Total de juegos: {total_juegos}<br>
                                    Puntos acumulados: {total_puntaje}<br>
                                    Promedio: {promedio_puntaje:.1f} pts/juego
                                </p>
                            </div>

                            <div style="text-align: center; margin: 30px 0;">
                                <a href="{share_link}"
                                   style="display: inline-block; background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
                                          color: white; padding: 15px 40px; text-decoration: none; border-radius: 25px;
                                          font-weight: bold; font-size: 1.1rem; box-shadow: 0 4px 15px rgba(102, 126, 234, 0.4);">
                                    📂 Ver mi Historial en OneDrive
                                </a>
                            </div>

                            <p style="color: #999; font-size: 0.9rem; text-align: center; margin-top: 30px;">
                                💡 <em>Puedes descargar y guardar este archivo para tus registros personales.</em>
                            </p>
                        </div>

                        <div style="text-align: center; margin-top: 20px; color: #999; font-size: 0.85rem;">
                            <p style="margin: 5px 0;">Brain RUSH - Sistema de Evaluación Interactiva</p>
                            <p style="margin: 5px 0;">Este archivo estará disponible en OneDrive</p>
                        </div>
                    </body>
                    </html>
                    """

                    msg = Message(
                        subject=f"📊 Tu Historial de Brain RUSH está listo",
                        recipients=[email_usuario],
                        html=mensaje_html,
                        sender=app.config['MAIL_DEFAULT_SENDER']
                    )

                    mail.send(msg)
                    print(f"✅ Email enviado a {email_usuario} con link de OneDrive")

                    return jsonify({
                        'success': True,
                        'message': 'Historial exportado a OneDrive y link enviado por correo',
                        'file_name': file_name,
                        'share_link': share_link,
                        'email_sent': True
                    }), 200
                else:
                    # Email no configurado, solo devolver el link
                    return jsonify({
                        'success': True,
                        'message': 'Historial exportado a OneDrive',
                        'file_name': file_name,
                        'share_link': share_link,
                        'email_sent': False
                    }), 200

            except Exception as email_error:
                print(f"⚠️ Error enviando email: {email_error}")
                # Aunque falle el email, el archivo está en OneDrive
                return jsonify({
                    'success': True,
                    'message': 'Historial exportado a OneDrive (email no enviado)',
                    'file_name': file_name,
                    'share_link': share_link,
                    'email_sent': False,
                    'warning': 'No se pudo enviar el email'
                }), 200
        else:
            # OneDrive falló, usar email como fallback
            print(f"❌ OneDrive falló: {resultado_onedrive.get('error')}")
            return enviar_historial_por_email(email_usuario, nombre_completo, filename, excel_data)

    except Exception as e:
        print(f"ERROR exportar_historial_estudiante: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 500


def enviar_historial_por_email(email_usuario, nombre_completo, filename, excel_data):
    """Enviar historial por email como fallback"""
    try:
        from flask_mail import Message
        from extensions import mail

        if not app.config.get('MAIL_ENABLED'):
            return jsonify({
                'success': False,
                'error': 'No se pudo subir a OneDrive y el correo no está configurado'
            }), 500

        msg = Message(
            subject='Tu Historial de Brain RUSH',
            sender=app.config.get('MAIL_DEFAULT_SENDER'),
            recipients=[email_usuario],
            body=f'Hola {nombre_completo},\n\nAdjunto tu historial de participaciones en Brain RUSH.'
        )

        msg.attach(filename, 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', excel_data)
        mail.send(msg)

        return jsonify({
            'success': True,
            'message': f'Historial enviado por correo a {email_usuario}',
            'file_name': filename,
            'email_sent': True,
            'fallback': True
        }), 200

    except Exception as e:
        print(f"Error enviando email: {e}")
        return jsonify({
            'success': False,
            'error': f'No se pudo subir a OneDrive ni enviar por email: {str(e)}'
        }), 500

@app.route('/auth/onedrive')
def onedrive_auth():
    """Iniciar flujo de autorización OAuth2 con OneDrive"""
    try:
        if 'usuario_id' not in session:
            flash('Debes iniciar sesión primero', 'warning')
            return redirect(url_for('login'))

        # Guardar usuario_id en el state para recuperarlo en el callback
        state = str(session['usuario_id'])

        # Configurar MSAL
        msal_app = msal.ConfidentialClientApplication(
            app.config['AZURE_CLIENT_ID'],
            authority=f"https://login.microsoftonline.com/{app.config['AZURE_TENANT_ID']}",
            client_credential=app.config['AZURE_CLIENT_SECRET']
        )

        # Generar URL de autorización
        auth_url = msal_app.get_authorization_request_url(
            scopes=app.config['ONEDRIVE_SCOPES'],
            state=state,
            redirect_uri=app.config['ONEDRIVE_REDIRECT_URI']
        )

        return redirect(auth_url)

    except Exception as e:
        print(f"Error en onedrive_auth: {e}")
        import traceback
        traceback.print_exc()
        flash(f'Error iniciando autorización: {str(e)}', 'danger')
        return redirect(url_for('dashboard'))


@app.route('/auth/onedrive-sistema')
def onedrive_auth_sistema():
    """
    Autorizar OneDrive para el SISTEMA (almacenar tokens en .env)
    TEMPORAL: Accesible por administradores Y docentes (para configuración inicial)
    """
    try:
        # Verificar que haya sesión
        if 'usuario_id' not in session:
            return """
            <html>
            <head><title>Error - Brain RUSH</title></head>
            <body style="font-family: Arial; padding: 40px; max-width: 600px; margin: 0 auto;">
                <h1 style="color: #ef4444;">❌ Error</h1>
                <p>Debes iniciar sesión para configurar OneDrive del sistema.</p>
                <a href="/login" style="color: #667eea;">Iniciar sesión</a>
            </body>
            </html>
            """

        # Verificar que sea admin O docente (temporal)
        conexion = obtener_conexion()
        cursor = conexion.cursor()
        cursor.execute("SELECT tipo_usuario FROM usuarios WHERE id_usuario = %s", (session['usuario_id'],))
        usuario = cursor.fetchone()
        cursor.close()
        conexion.close()

        if not usuario or usuario[0] not in ['admin', 'docente']:
            return """
            <html>
            <head><title>Acceso Denegado - Brain RUSH</title></head>
            <body style="font-family: Arial; padding: 40px; max-width: 600px; margin: 0 auto;">
                <h1 style="color: #ef4444;">🚫 Acceso Denegado</h1>
                <p>Solo los administradores y docentes pueden configurar OneDrive del sistema.</p>
                <a href="/" style="color: #667eea;">Volver al inicio</a>
            </body>
            </html>
            """

        # Usar state especial para indicar que es configuración del sistema
        state = "SISTEMA_CONFIG"

        # Configurar MSAL
        msal_app = msal.ConfidentialClientApplication(
            app.config['AZURE_CLIENT_ID'],
            authority=f"https://login.microsoftonline.com/{app.config['AZURE_TENANT_ID']}",
            client_credential=app.config['AZURE_CLIENT_SECRET']
        )

        # Generar URL de autorización
        auth_url = msal_app.get_authorization_request_url(
            scopes=app.config['ONEDRIVE_SCOPES'],
            state=state,
            redirect_uri=app.config['ONEDRIVE_REDIRECT_URI']
        )

        return redirect(auth_url)

    except Exception as e:
        print(f"Error en onedrive_auth_sistema: {e}")
        import traceback
        traceback.print_exc()
        return f"""
        <html>
        <head><title>Error - Brain RUSH</title></head>
        <body style="font-family: Arial; padding: 40px; max-width: 600px; margin: 0 auto;">
            <h1 style="color: #ef4444;">❌ Error</h1>
            <p>Error iniciando autorización: {str(e)}</p>
            <a href="/" style="color: #667eea;">Volver al inicio</a>
        </body>
        </html>
        """


@app.route('/callback/onedrive')
def onedrive_callback():
    """Callback de OAuth2 para OneDrive"""
    try:
        # Obtener el código de autorización
        code = request.args.get('code')
        state = request.args.get('state')
        error = request.args.get('error')

        if error:
            return f"""
            <html>
            <head><title>Error - Brain RUSH</title></head>
            <body style="font-family: Arial; padding: 40px; max-width: 600px; margin: 0 auto;">
                <h1 style="color: #ef4444;">❌ Error en Autorización</h1>
                <p>Error: {error}</p>
                <a href="/" style="color: #667eea;">Volver al inicio</a>
            </body>
            </html>
            """

        if not code:
            return """
            <html>
            <head><title>Error - Brain RUSH</title></head>
            <body style="font-family: Arial; padding: 40px; max-width: 600px; margin: 0 auto;">
                <h1 style="color: #ef4444;">❌ Error</h1>
                <p>No se recibió código de autorización</p>
                <a href="/" style="color: #667eea;">Volver al inicio</a>
            </body>
            </html>
            """

        # Configurar MSAL
        msal_app = msal.ConfidentialClientApplication(
            app.config['AZURE_CLIENT_ID'],
            authority=f"https://login.microsoftonline.com/{app.config['AZURE_TENANT_ID']}",
            client_credential=app.config['AZURE_CLIENT_SECRET']
        )

        # Intercambiar código por token
        result = msal_app.acquire_token_by_authorization_code(
            code,
            scopes=app.config['ONEDRIVE_SCOPES'],
            redirect_uri=app.config['ONEDRIVE_REDIRECT_URI']
        )

        if "access_token" in result:
            access_token = result['access_token']
            refresh_token = result.get('refresh_token', '')
            expires_in = result.get('expires_in', 3600)
            token_expires = datetime.now() + timedelta(seconds=expires_in)

            # Verificar si es configuración del sistema
            if state == "SISTEMA_CONFIG":
                # Guardar tokens en .env
                actualizar_env_tokens(access_token, refresh_token, token_expires)

                # También actualizar en memoria
                app.config['ONEDRIVE_ACCESS_TOKEN'] = access_token
                app.config['ONEDRIVE_REFRESH_TOKEN'] = refresh_token
                app.config['ONEDRIVE_TOKEN_EXPIRES'] = token_expires.isoformat()

                return f"""
                <html>
                <head>
                    <title>Configuración Exitosa - Brain RUSH</title>
                    <style>
                        body {{ font-family: Arial, sans-serif; padding: 40px; max-width: 800px; margin: 0 auto; }}
                        .success {{ background: #10b981; color: white; padding: 20px; border-radius: 8px; margin: 20px 0; }}
                        .code-block {{ background: #1f2937; color: #10b981; padding: 15px; border-radius: 8px; font-family: monospace; margin: 20px 0; overflow-x: auto; }}
                        .warning {{ background: #fef3c7; color: #92400e; padding: 15px; border-radius: 8px; margin: 20px 0; }}
                        h1 {{ color: #667eea; }}
                        a {{ color: #667eea; text-decoration: none; font-weight: bold; }}
                    </style>
                </head>
                <body>
                    <h1>✅ Configuración de OneDrive Exitosa</h1>

                    <div class="success">
                        <h2 style="margin-top: 0;">🎉 ¡Tokens Guardados!</h2>
                        <p>Los tokens de OneDrive del sistema han sido guardados en el archivo .env</p>
                    </div>

                    <h3>📋 Tokens Obtenidos:</h3>
                    <div class="code-block">
ONEDRIVE_ACCESS_TOKEN={access_token[:50]}...
ONEDRIVE_REFRESH_TOKEN={refresh_token[:50]}...
ONEDRIVE_TOKEN_EXPIRES={token_expires.isoformat()}
                    </div>

                    <div class="warning">
                        <h3 style="margin-top: 0;">⚠️ Importante</h3>
                        <ul>
                            <li>Los tokens se han guardado automáticamente en tu archivo .env</li>
                            <li>El access_token expira en {expires_in // 3600} hora(s)</li>
                            <li>El refresh_token se usará para obtener nuevos access_tokens automáticamente</li>
                            <li>Asegúrate de que el archivo .env esté en tu .gitignore</li>
                        </ul>
                    </div>

                    <h3>✅ Siguiente Paso:</h3>
                    <p>Los archivos exportados ahora se guardarán en la cuenta de OneDrive que acabas de autorizar.</p>

                    <p><a href="/">← Volver al inicio</a></p>
                </body>
                </html>
                """

            # Si no es configuración del sistema, es un usuario normal
            try:
                usuario_id = int(state)
            except:
                return """
                <html>
                <head><title>Error - Brain RUSH</title></head>
                <body style="font-family: Arial; padding: 40px; max-width: 600px; margin: 0 auto;">
                    <h1 style="color: #ef4444;">❌ Error</h1>
                    <p>Sesión inválida</p>
                    <a href="/login" style="color: #667eea;">Iniciar sesión</a>
                </body>
                </html>
                """

            # Guardar tokens en la base de datos para el usuario
            conexion = obtener_conexion()
            cursor = conexion.cursor()

            cursor.execute("""
                UPDATE usuarios
                SET onedrive_access_token = %s,
                    onedrive_refresh_token = %s,
                    onedrive_token_expires = %s
                WHERE id_usuario = %s
            """, (access_token, refresh_token, token_expires, usuario_id))

            conexion.commit()

            # Restaurar sesión
            cursor.execute("SELECT nombre, tipo_usuario FROM usuarios WHERE id_usuario = %s", (usuario_id,))
            usuario = cursor.fetchone()

            cursor.close()
            conexion.close()

            if usuario:
                # Restaurar la sesión
                session['usuario_id'] = usuario_id
                session['nombre_usuario'] = usuario[0]
                session['usuario_tipo'] = usuario[1]  # CORREGIDO: usuario_tipo (no tipo_usuario)

                flash('✅ OneDrive conectado exitosamente. Ahora puedes exportar resultados directamente.', 'success')
                return redirect(url_for('dashboard'))
            else:
                flash('Error restaurando sesión', 'danger')
                return redirect(url_for('login'))
        else:
            error_desc = result.get('error_description', 'Error desconocido')
            flash(f'Error al obtener token de OneDrive: {error_desc}', 'danger')
            return redirect(url_for('login'))

    except Exception as e:
        print(f"ERROR onedrive_callback: {e}")
        import traceback
        traceback.print_exc()
        flash(f'Error en callback de OneDrive: {str(e)}', 'danger')
        return redirect(url_for('dashboard'))

@app.route('/perfil', methods=['GET', 'POST'])
def perfil():
    if 'usuario_id' not in session:
        return redirect(url_for('login'))

    usuario_id = session['usuario_id']
    conexion = obtener_conexion()
    cursor = conexion.cursor()

    cursor.execute("""
        SELECT id_usuario, nombre, apellidos, email, tipo_usuario, fecha_registro, estado
        FROM usuarios
        WHERE id_usuario = %s
    """, (usuario_id,))
    fila = cursor.fetchone()
    columnas = [col[0] for col in cursor.description]
    usuario = dict(zip(columnas, fila)) if fila else None

    if not usuario:
        flash("No se encontró el usuario.", "danger")
        conexion.close()
        return redirect(url_for('login'))

    if request.method == 'POST':
        print("DEBUG /perfil - request.form:", dict(request.form))

    if request.method == 'POST' and request.form.get('form_type') == 'update':
        nombre = request.form.get('nombre', '').strip()
        apellidos = request.form.get('apellidos', '').strip()
        email = request.form.get('email', '').strip().lower()
        password = request.form.get('password', '').strip()

        # ===== VALIDACIÓN DE CORREO INSTITUCIONAL =====
        if email != usuario['email']:  # Solo validar si se está cambiando
            dominios_permitidos = ['@usat.pe', '@usat.edu.pe']
            email_valido = any(email.endswith(dominio) for dominio in dominios_permitidos)

            if not email_valido:
                flash('❌ Solo se permiten correos institucionales con dominio @usat.pe o @usat.edu.pe', 'danger')
                conexion.close()
                return redirect(url_for('perfil'))

            # Verificar que el email no esté en uso por otro usuario
            cursor.execute("SELECT id_usuario FROM usuarios WHERE email = %s AND id_usuario != %s", (email, usuario_id))
            if cursor.fetchone():
                flash('❌ Este correo ya está registrado por otro usuario', 'danger')
                conexion.close()
                return redirect(url_for('perfil'))

        # ===== VALIDACIÓN DE CONTRASEÑA FUERTE =====
        if password:
            # Validar longitud mínima
            if len(password) < 8:
                flash('❌ La contraseña debe tener al menos 8 caracteres', 'danger')
                conexion.close()
                return redirect(url_for('perfil'))

            # Validar que tenga mayúscula
            if not any(c.isupper() for c in password):
                flash('❌ La contraseña debe contener al menos una letra mayúscula', 'danger')
                conexion.close()
                return redirect(url_for('perfil'))

            # Validar que tenga minúscula
            if not any(c.islower() for c in password):
                flash('❌ La contraseña debe contener al menos una letra minúscula', 'danger')
                conexion.close()
                return redirect(url_for('perfil'))

            # Validar que tenga número
            if not any(c.isdigit() for c in password):
                flash('❌ La contraseña debe contener al menos un número', 'danger')
                conexion.close()
                return redirect(url_for('perfil'))

            # Validar contraseña única (no usada por otros)
            if not controlador_usuario.verificar_contrasena_unica(password, usuario_id):
                flash('❌ Esta contraseña ya está siendo utilizada por otro usuario. Por favor, elige una diferente.', 'danger')
                conexion.close()
                return redirect(url_for('perfil'))

        cambios = []
        valores = []

        if nombre != usuario['nombre']:
            cambios.append("nombre = %s")
            valores.append(nombre)
        if apellidos != usuario['apellidos']:
            cambios.append("apellidos = %s")
            valores.append(apellidos)
        if email != usuario['email']:
            cambios.append("email = %s")
            valores.append(email)
        if password:
            cambios.append("contraseña_hash = MD5(%s)")
            valores.append(password)

        if cambios:
            query = f"UPDATE usuarios SET {', '.join(cambios)} WHERE id_usuario = %s"
            valores.append(usuario_id)
            cursor.execute(query, tuple(valores))
            conexion.commit()
            flash('✅ Perfil actualizado correctamente.', 'success')
        else:
            flash('ℹ️ No se realizaron cambios.', 'info')

        conexion.close()
        return redirect(url_for('perfil'))

    if request.method == 'POST' and request.form.get('form_type') == 'delete':
        conexion.close()  # Cerrar la conexión actual antes de llamar a la función

        # Usar la función que elimina en cascada
        exito, mensaje = controlador_usuario.eliminar_usuario_completo(usuario_id)

        if exito:
            session.clear()
            flash('🗑️ Tu cuenta y todos tus datos han sido eliminados correctamente.', 'warning')
            return redirect(url_for('login'))
        else:
            flash(f'❌ Error al eliminar cuenta: {mensaje}', 'danger')
            return redirect(url_for('perfil'))

    conexion.close()
    return render_template('MiPerfil.html', usuario=usuario)

if __name__ == '__main__':
    # Verificar conexión e inicializar usuarios de prueba
    print("Iniciando Brain RUSH...")
    if verificar_conexion():
        print("✅ Conexión a base de datos exitosa")
        inicializar_usuarios_prueba()
    else:
        print("❌ Error de conexión a la base de datos")

    print("🚀 Iniciando servidor Flask en http://127.0.0.1:5000")
    app.run(debug=True, port=5000)